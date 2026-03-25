#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build an allowlist of car_week values from the current JP workbook and,
optionally, union it with the legacy workbook.
"""

from __future__ import annotations

import argparse
import datetime
import re
from pathlib import Path
from typing import Any, List, Optional

import pandas as pd


CARWEEK_RE = re.compile(r"(?P<car>\d{1,2})_(?P<start>\d{6})-(?P<end>\d+)")


def _norm(s: str) -> str:
    return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())


def _normalize_sheet_token(name: str) -> str:
    return re.sub(r"[\s_\-]+", "", str(name or "").strip()).casefold()


def _choose_sheet_name_from_workbook(
    wb: Any,
    requested: Optional[str],
    required_headers: Optional[List[str]] = None,
    prefer_keywords: Optional[List[str]] = None,
) -> str:
    names = list(getattr(wb, "sheetnames", []) or [])
    if not names:
        raise ValueError("Workbook has no sheets.")

    requested = str(requested or "").strip()
    if requested and requested in names:
        return requested

    requested_norm = _normalize_sheet_token(requested)
    for name in names:
        if requested_norm and _normalize_sheet_token(name) == requested_norm:
            return name

    required_headers = required_headers or []
    prefer_keywords = prefer_keywords or []
    best_name = names[0]
    best_score = -1

    for name in names:
        score = 0
        name_norm = _normalize_sheet_token(name)
        if requested_norm and requested_norm in name_norm:
            score += 20
        for kw in prefer_keywords:
            kw_norm = _normalize_sheet_token(kw)
            if kw_norm and kw_norm in name_norm:
                score += 8

        ws = wb[name]
        seen: set[str] = set()
        max_row = min(int(getattr(ws, "max_row", 0) or 0), 40)
        max_col = min(int(getattr(ws, "max_column", 0) or 0), 120)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                value = str(ws.cell(r, c).value or "").strip()
                if value:
                    seen.add(value)
        score += 12 * sum(1 for header in required_headers if header in seen)
        if score > best_score:
            best_name = name
            best_score = score

    return best_name


def _resolve_sheet_name(
    excel: Path,
    requested: Optional[str],
    required_headers: Optional[List[str]] = None,
    prefer_keywords: Optional[List[str]] = None,
) -> str:
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(excel, read_only=True, data_only=True)
    try:
        return _choose_sheet_name_from_workbook(
            wb,
            requested=requested,
            required_headers=required_headers,
            prefer_keywords=prefer_keywords,
        )
    finally:
        close = getattr(wb, "close", None)
        if callable(close):
            close()


def _looks_like_yymmdd(value: str) -> bool:
    if len(value) != 6 or not value.isdigit():
        return False
    try:
        datetime.datetime.strptime(value, "%y%m%d")
        return True
    except Exception:
        return False


def _repair_carweek_end(start_ymd: str, tail_digits: str) -> Optional[str]:
    digits = re.sub(r"\D", "", tail_digits or "")
    if not start_ymd or not digits:
        return None
    candidates = []
    if len(digits) >= 6:
        candidates.append(digits[:6])
    else:
        prefix = start_ymd[: max(0, 6 - len(digits))]
        candidates.append((prefix + digits)[-6:])
    for cand in candidates:
        if _looks_like_yymmdd(cand):
            return cand
    return candidates[0] if candidates else None


def normalize_carweek_token(raw: str) -> str:
    s = str(raw or "").strip().replace("\\", "/")
    m = CARWEEK_RE.search(s)
    if not m:
        return s
    car = m.group("car").zfill(2)
    start = m.group("start")
    end_raw = m.group("end")
    end = end_raw[:6] if _looks_like_yymmdd(end_raw[:6]) else _repair_carweek_end(start, end_raw)
    return f"{car}_{start}-{end}" if end else f"{car}_{start}-{end_raw[:6]}"


def _try_extract_from_pandas(excel: Path, sheet: str) -> Optional[List[str]]:
    """Fast path for sheets with a top-row car_week column."""
    try:
        df = pd.read_excel(excel, sheet_name=sheet)
    except Exception:
        return None

    cols = list(df.columns)
    if not cols:
        return None

    if "car_week" in df.columns:
        col = "car_week"
    else:
        norm = {_norm(c): c for c in cols}
        col = None
        for key in ("carweek", "carwk", "carweekid", "carweekname"):
            if key in norm:
                col = norm[key]
                break
        if col is None:
            for c in cols:
                token = _norm(c)
                if "car" in token and "week" in token:
                    col = c
                    break
    if col is None:
        return None

    vals = []
    for v in df[col].dropna().astype(str).tolist():
        token = normalize_carweek_token(v.strip())
        if token:
            vals.append(token)
    return sorted(set(vals)) or None


def _extract_from_jp_sheet(excel: Path, sheet: str) -> List[str]:
    """Fallback for JP sheets where the header row is not the first row."""
    from openpyxl import load_workbook  # lazy import

    folder_header = "\u30d5\u30a9\u30eb\u30c0"
    jp_keywords = ["\u96c6\u8a08", "\u9032\u6357", "progress", "summary"]
    wb = load_workbook(excel, data_only=True)
    try:
        sheet_name = _choose_sheet_name_from_workbook(
            wb,
            requested=sheet,
            required_headers=[folder_header, "HDD-No"],
            prefer_keywords=jp_keywords,
        )
        ws = wb[sheet_name]

        header_row = None
        col_folder = None
        wanted_tokens = {
            _normalize_sheet_token(folder_header),
            _normalize_sheet_token("folder"),
            _normalize_sheet_token("car_week"),
            _normalize_sheet_token("carweek"),
        }

        for r in range(1, 401):
            row_vals = [ws.cell(r, c).value for c in range(1, 80)]
            for c, value in enumerate(row_vals, start=1):
                token = _normalize_sheet_token(value)
                if token in wanted_tokens or ("car" in token and "week" in token):
                    header_row = r
                    col_folder = c
                    break
            if header_row is not None:
                break

        if header_row is None or col_folder is None:
            raise ValueError(f"Could not locate a folder/car_week column in sheet: {sheet_name}")

        carweeks = []
        for r in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(r, col_folder).value
            if value is None:
                continue
            token = normalize_carweek_token(str(value).strip())
            if token:
                carweeks.append(token)

        carweeks = sorted(set(carweeks))
        if not carweeks:
            raise ValueError(f"No car_week values found in sheet: {sheet_name}")
        return carweeks
    finally:
        close = getattr(wb, "close", None)
        if callable(close):
            close()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--old-excel", default=None)
    ap.add_argument("--old-sheet", default=None)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    excel = Path(args.excel)
    if not excel.exists():
        raise FileNotFoundError(excel)

    folder_header = "\u30d5\u30a9\u30eb\u30c0"
    jp_sheet_default = "\u96c6\u8a08\u8868"
    jp_keywords = ["\u96c6\u8a08", "\u9032\u6357", "progress", "summary"]
    sheet = _resolve_sheet_name(
        excel,
        args.sheet or jp_sheet_default,
        required_headers=[folder_header, "HDD-No"],
        prefer_keywords=jp_keywords,
    )

    carweeks = _try_extract_from_pandas(excel, sheet)
    if carweeks is None:
        carweeks = _extract_from_jp_sheet(excel, sheet)

    carweek_set = set(carweeks)
    if args.old_excel:
        old_excel = Path(args.old_excel)
        if old_excel.exists():
            try:
                old_sheet = _resolve_sheet_name(
                    old_excel,
                    args.old_sheet or "AllPoints_1018",
                    required_headers=["car_week"],
                    prefer_keywords=["allpoints", "1018", "legacy"],
                )
                df_old = pd.read_excel(old_excel, sheet_name=old_sheet)
                if "car_week" in df_old.columns:
                    extra = df_old["car_week"].dropna().astype(str).tolist()
                    for token in map(normalize_carweek_token, extra):
                        if token:
                            carweek_set.add(token)
            except Exception:
                pass

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    carweeks_out = sorted(carweek_set)
    out.write_text("\n".join(carweeks_out) + "\n", encoding="utf-8")

    print(f"Wrote {len(carweeks_out)} car_week entries -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
