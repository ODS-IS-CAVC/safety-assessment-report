#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_nearmiss_map_v67_richui_both_excels.py

Goal:
- Read BOTH excels:
    1) ニアミス抽出_進捗管理表.xlsx (sheet: 集計表)  [JP progress]
    2) route_nearmiss_analysis_with_dx_and_unknown_GROUPED_v12_all1018.xlsx (sheet: AllPoints_1018) [legacy 1018]
- DX parsing rule:
    - Only the FIRST '●' is main DX (dx_main)
    - '〇' are subcategories (dx_subs) and are NOT counted as dx_main
- Counts:
    - Keep ALL rows (~14k) in DATA.rows
    - Mark rows without coordinates as has_latlon=false
    - Map draws only has_latlon rows, but Summary/Analysis uses ALL rows.
- Label fallback:
    - Prefer existing road/location
    - Else: OSM cache (osm_label_cache.json)
    - Else: optional OSM reverse lookup (Nominatim) + writeback to osm_reverse_cache.json
"""

from __future__ import annotations

import argparse
import bisect
import csv
import datetime
import json
import math
import os
import re
import time
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

# --------------------------- small utils ---------------------------

def _s(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _is_mark(val: Any) -> bool:
    """Risk marker: accept any non-empty that isn't 0."""
    if val is None:
        return False
    s = str(val).strip()
    return s != "" and s != "0"

def _has_symbol(val: Any, sym: str) -> bool:
    if val is None:
        return False
    return sym in str(val)

def _safe_float(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None

def _safe_int(v: Any) -> Optional[int]:
    try:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        return int(float(s))
    except Exception:
        return None

def _is_missing_value(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    s = str(v).strip()
    return s == "" or s.lower() in {"nan", "none", "null"}

def _coerce_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "on"}

def _looks_like_dx_code(v: Any) -> bool:
    return bool(re.match(r"^DX-\d{2}$", _s(v)))

def _is_valid_coord(lat: Any, lon: Any) -> bool:
    if lat is None or lon is None:
        return False
    try:
        lat_f = float(lat)
        lon_f = float(lon)
    except Exception:
        return False
    if not math.isfinite(lat_f) or not math.isfinite(lon_f):
        return False
    if abs(lat_f) > 90 or abs(lon_f) > 180:
        return False
    if lat_f == 0.0 and lon_f == 0.0:
        return False
    return True

def _first_present(rec: Dict[str, Any], keys: List[str]) -> Any:
    for key in keys:
        if key in rec and not _is_missing_value(rec.get(key)):
            return rec.get(key)
    return None

def _pick_float(rec: Dict[str, Any], keys: List[str]) -> Optional[float]:
    for key in keys:
        if key not in rec:
            continue
        val = _safe_float(rec.get(key))
        if val is not None:
            return float(val)
    return None

def _normalize_sheet_token(name: str) -> str:
    return re.sub(r"[\s_\-]+", "", _s(name)).casefold()

def _choose_sheet_name_from_workbook(
    wb: Any,
    requested: str,
    required_headers: Optional[List[str]] = None,
    prefer_keywords: Optional[List[str]] = None,
) -> str:
    names = list(getattr(wb, "sheetnames", []) or [])
    if not names:
        raise ValueError("Workbook has no sheets.")
    if requested in names:
        return requested

    requested_norm = _normalize_sheet_token(requested)
    for name in names:
        if _normalize_sheet_token(name) == requested_norm:
            return name

    required_headers = required_headers or []
    prefer_keywords = prefer_keywords or []
    best_name = names[0]
    best_score = -1

    for name in names:
        score = 0
        name_norm = _normalize_sheet_token(name)
        if requested_norm and requested_norm in name_norm:
            score += 20
        for kw in prefer_keywords:
            if kw and _normalize_sheet_token(kw) in name_norm:
                score += 8

        ws = wb[name]
        seen: set[str] = set()
        max_row = min(int(getattr(ws, "max_row", 0) or 0), 40)
        max_col = min(int(getattr(ws, "max_column", 0) or 0), 120)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                s = _s(ws.cell(r, c).value)
                if s:
                    seen.add(s)
        score += 12 * sum(1 for header in required_headers if header in seen)
        if score > best_score:
            best_name = name
            best_score = score

    return best_name

LAT_VALUE_KEYS = [
    "lat",
    "latitude",
    "Latitude",
    "LATITUDE",
    "LAT",
    "center_latitude",
    "LATITUDEFRONT",
    "緯度",
]

LON_VALUE_KEYS = [
    "lon",
    "lng",
    "longitude",
    "Longitude",
    "LONGITUDE",
    "LON",
    "center_longitude",
    "LONGITUDEFRONT",
    "経度",
]

ROAD_VALUE_KEYS = [
    "道路",
    "road",
    "road_name",
    "road_label",
    "street",
    "route_name",
    "route_group_main_v2",
    "route_group_main",
    "route_group",
]
LOC_VALUE_KEYS = [
    "ロケーション",
    "location",
    "location_label",
    "place",
    "area",
    "city",
    "town",
    "village",
    "suburb",
    "neighbourhood",
]

def _is_missing_label_text(v: Any) -> bool:
    s = _s(v)
    if not s:
        return True
    return s.casefold() in {"-", "unknown", "none", "nan", "null", "不明", "道路名なし", "locationなし"}

GENERIC_LOCATION_VALUES = {"Highway", "Local", "Prefectural", "National"}

ROAD_TYPE_LABELS = {
    "motorway": "高速道路",
    "motorway_link": "高速道路接続部",
    "trunk": "自動車専用道路",
    "trunk_link": "自動車専用道路接続部",
    "primary": "一般道",
    "secondary": "一般道",
    "tertiary": "一般道",
    "residential": "一般道",
    "unclassified": "一般道",
    "service": "サービス道路",
    "living_street": "生活道路",
    "road": "道路",
}

def _is_generic_location_value(v: Any) -> bool:
    return _s(v) in GENERIC_LOCATION_VALUES

def normalize_row_coordinates(rec: Dict[str, Any]) -> Dict[str, Any]:
    lat = _pick_float(rec, LAT_VALUE_KEYS)
    lon = _pick_float(rec, LON_VALUE_KEYS)
    if _is_valid_coord(lat, lon):
        rec["lat"] = float(lat)
        rec["lon"] = float(lon)
        rec["has_latlon"] = True
    else:
        rec["lat"] = None
        rec["lon"] = None
        rec["has_latlon"] = False
    return rec

# --------------------------- OSM label cache ---------------------------

def load_osm_label_cache(path: Path) -> Dict[str, List[Any]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def load_osm_reverse_cache(path: Path) -> Dict[str, Dict[str, Any]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def load_json_object_cache(path: Path) -> Dict[str, Dict[str, Any]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else {}

def save_osm_reverse_cache(path: Path, cache: Dict[str, Dict[str, Any]]) -> None:
    tmp = path.with_suffix(".tmp.json")
    tmp.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    tmp.replace(path)

def save_json_object_cache(path: Path, cache: Dict[str, Dict[str, Any]]) -> None:
    tmp = path.with_suffix(".tmp.json")
    tmp.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    tmp.replace(path)

def _cache_keys(lat: float, lon: float) -> List[str]:
    # The provided osm_label_cache.json uses a slightly odd rounding (often lat:4, lon:3).
    # Try several candidates to improve hit-rate.
    keys = []
    for lat_d, lon_d in [(5,5), (5,4), (4,4), (4,3), (3,3)]:
        keys.append(f"{round(lat, lat_d)},{round(lon, lon_d)}")
    # also try fixed formatting
    keys.append(f"{lat:.4f},{lon:.3f}")
    keys.append(f"{lat:.4f},{lon:.4f}")
    keys.append(f"{lat:.5f},{lon:.5f}")
    # unique preserve order
    out=[]
    seen=set()
    for k in keys:
        if k not in seen:
            seen.add(k)
            out.append(k)
    return out

def _road_label_from_tags(tags: Dict[str, Any]) -> Optional[str]:
    if not isinstance(tags, dict):
        return None
    for key in ["name:ja", "name", "official_name", "loc_name", "alt_name", "ref"]:
        val = _s(tags.get(key))
        if not _is_missing_label_text(val):
            return val
    road_type = _s(tags.get("highway"))
    if road_type:
        return ROAD_TYPE_LABELS.get(road_type, road_type)
    return None

def _location_label_from_tags(tags: Dict[str, Any]) -> Optional[str]:
    if not isinstance(tags, dict):
        return None
    for key in ["addr:suburb", "addr:quarter", "addr:city", "addr:town", "addr:village", "destination", "junction:ref"]:
        val = _s(tags.get(key))
        if val and not _is_generic_location_value(val):
            return val
    return None

def _project_local_xy_m(lat: float, lon: float, ref_lat: float, ref_lon: float) -> Tuple[float, float]:
    scale = math.cos(math.radians(ref_lat))
    x = (lon - ref_lon) * 111320.0 * scale
    y = (lat - ref_lat) * 110540.0
    return x, y

def _point_segment_distance_m(
    lat: float,
    lon: float,
    a_lat: float,
    a_lon: float,
    b_lat: float,
    b_lon: float,
) -> float:
    px, py = _project_local_xy_m(lat, lon, lat, lon)
    ax, ay = _project_local_xy_m(a_lat, a_lon, lat, lon)
    bx, by = _project_local_xy_m(b_lat, b_lon, lat, lon)
    dx = bx - ax
    dy = by - ay
    if dx == 0.0 and dy == 0.0:
        return math.hypot(px - ax, py - ay)
    t = ((px - ax) * dx + (py - ay) * dy) / ((dx * dx) + (dy * dy))
    t = max(0.0, min(1.0, t))
    qx = ax + t * dx
    qy = ay + t * dy
    return math.hypot(px - qx, py - qy)

def _way_distance_m(lat: float, lon: float, geometry: List[Dict[str, Any]]) -> Optional[float]:
    coords: List[Tuple[float, float]] = []
    for item in geometry or []:
        la = _safe_float(item.get("lat")) if isinstance(item, dict) else None
        lo = _safe_float(item.get("lon")) if isinstance(item, dict) else None
        if _is_valid_coord(la, lo):
            coords.append((float(la), float(lo)))
    if not coords:
        return None
    if len(coords) == 1:
        return _point_segment_distance_m(lat, lon, coords[0][0], coords[0][1], coords[0][0], coords[0][1])
    best = None
    for i in range(len(coords) - 1):
        d = _point_segment_distance_m(lat, lon, coords[i][0], coords[i][1], coords[i + 1][0], coords[i + 1][1])
        if best is None or d < best:
            best = d
    return best

def overpass_lookup_labels(lat: float, lon: float, overpass_cache: Dict[str, Dict[str, Any]]) -> Tuple[Optional[str], Optional[str], str]:
    for k in _cache_keys(lat, lon):
        item = overpass_cache.get(k)
        if not isinstance(item, dict):
            continue
        road = item.get("road") or item.get("road_label")
        loc = item.get("location") or item.get("location_label")
        if road or loc:
            return _s(road) or None, _s(loc) or None, "osm_overpass_cache"
    return None, None, "none"

def osm_lookup_labels(
    lat: float,
    lon: float,
    osm_cache: Dict[str, List[Any]],
    rev_cache: Dict[str, Dict[str, Any]],
    overpass_cache: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Tuple[Optional[str], Optional[str], str]:
    """
    Returns (road, location, source_tag).
    osm_label_cache format example: key -> ["Highway","motorway","新東名高速道路", None]
    reverse cache format: key -> {road, location, ...}
    overpass cache format: key -> {road, location, highway, distance_m, ...}
    """
    road = None
    loc = None
    source_tags: List[str] = []
    for k in _cache_keys(lat, lon):
        if k in rev_cache:
            rc = rev_cache[k]
            road = road or rc.get("road") or rc.get("road_label")
            loc = loc or rc.get("location") or rc.get("location_label")
            if road or loc:
                source_tags.append("osm_reverse_cache")
        if overpass_cache and k in overpass_cache:
            oc = overpass_cache[k]
            if not road:
                road = oc.get("road") or oc.get("road_label")
            if not loc:
                loc = oc.get("location") or oc.get("location_label")
            if road or loc:
                source_tags.append("osm_overpass_cache")
        if k in osm_cache:
            v = osm_cache[k]
            if isinstance(v, list):
                kind = _s(v[0] if len(v) > 0 else None)
                road_type = _s(v[1] if len(v) > 1 else None)
                road_name = _s(v[2] if len(v) > 2 else None)
                place = _s(v[3] if len(v) > 3 else None)
                if not road:
                    if not _is_missing_label_text(road_name):
                        road = road_name
                    elif road_type:
                        road = ROAD_TYPE_LABELS.get(road_type, road_type)
                    elif kind:
                        road = ROAD_TYPE_LABELS.get(kind.casefold(), kind)
                if not loc and place and not _is_generic_location_value(place):
                    loc = place
                if road or loc:
                    source_tags.append("osm_label_cache")
        if road and loc:
            break
    return (road or None), (loc or None), "+".join(dict.fromkeys(source_tags)) or "none"

def overpass_query_nearest_road(
    lat: float,
    lon: float,
    radius_m: float,
    timeout_s: float = 20.0,
    user_agent: str = "KAIT-nearmiss-map/1.0",
) -> Optional[Dict[str, Any]]:
    try:
        import requests
    except Exception:
        return None

    endpoints = [
        "https://overpass-api.de/api/interpreter",
        "https://overpass.kumi.systems/api/interpreter",
    ]
    radii = []
    for r in [radius_m, max(radius_m * 2.0, 80.0)]:
        if r > 0 and r not in radii:
            radii.append(r)

    headers = {"User-Agent": user_agent, "Content-Type": "text/plain; charset=utf-8"}
    for radius in radii:
        query = (
            f"[out:json][timeout:{max(10, int(timeout_s))}];"
            f'way(around:{float(radius):.1f},{lat:.7f},{lon:.7f})["highway"];'
            "out body geom qt;"
        )
        for endpoint in endpoints:
            try:
                res = requests.post(endpoint, data=query.encode("utf-8"), headers=headers, timeout=timeout_s)
                if res.status_code != 200:
                    continue
                js = res.json()
            except Exception:
                continue
            best = None
            best_score = None
            for el in js.get("elements", []) if isinstance(js, dict) else []:
                tags = el.get("tags") or {}
                if not isinstance(tags, dict):
                    continue
                dist_m = _way_distance_m(lat, lon, el.get("geometry") or [])
                if dist_m is None:
                    continue
                road = _road_label_from_tags(tags)
                loc = _location_label_from_tags(tags)
                highway = _s(tags.get("highway"))
                score = (
                    float(dist_m),
                    0 if road and road not in ROAD_TYPE_LABELS.values() else 1,
                    0 if _s(tags.get("name")) else 1,
                    highway,
                )
                if best is None or score < best_score:
                    best = {
                        "road": road,
                        "location": loc,
                        "highway": highway,
                        "name": _s(tags.get("name")) or _s(tags.get("name:ja")),
                        "ref": _s(tags.get("ref")),
                        "distance_m": float(dist_m),
                        "radius_m": float(radius),
                        "endpoint": endpoint,
                        "fetched_at": datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
                    }
                    best_score = score
            if best is not None:
                return best
    return None

def osm_reverse_geocode(lat: float, lon: float, user_agent: str, timeout_s: float = 8.0) -> Optional[Dict[str, Any]]:
    """
    Nominatim reverse. Optional (requires internet).
    """
    try:
        import requests
        url = "https://nominatim.openstreetmap.org/reverse"
        params = {
            "format": "jsonv2",
            "lat": f"{lat}",
            "lon": f"{lon}",
            "zoom": 18,
            "addressdetails": 1,
        }
        headers = {"User-Agent": user_agent}
        r = requests.get(url, params=params, headers=headers, timeout=timeout_s)
        if r.status_code != 200:
            return None
        return r.json()
    except Exception:
        return None

def extract_road_loc_from_nominatim(js: Dict[str, Any]) -> Tuple[Optional[str], Optional[str]]:
    addr = js.get("address") or {}
    road = addr.get("road") or addr.get("motorway") or addr.get("trunk") or addr.get("residential")
    # location: prefer amenity/hamlet/suburb/city
    loc = addr.get("neighbourhood") or addr.get("suburb") or addr.get("hamlet") or addr.get("village") or addr.get("town") or addr.get("city")
    # If still none, try display_name
    if not road:
        dn = js.get("display_name")
        if isinstance(dn, str) and dn:
            road = dn.split(",")[0].strip()
    return road, loc

def enrich_row_location_labels(
    rec: Dict[str, Any],
    osm_cache: Dict[str, List[Any]],
    rev_cache: Dict[str, Dict[str, Any]],
    allow_reverse_fetch: bool,
    max_requests: int,
    requests_used: int,
    sleep_s: float,
    overpass_cache: Optional[Dict[str, Dict[str, Any]]] = None,
    overpass_radius_m: float = 40.0,
    overpass_timeout_s: float = 20.0,
) -> int:
    if not _is_valid_coord(rec.get("lat"), rec.get("lon")):
        road0 = _first_present(rec, ROAD_VALUE_KEYS)
        loc0 = _first_present(rec, LOC_VALUE_KEYS)
        rec["道路"] = "" if _is_missing_label_text(road0) else _s(road0)
        rec["ロケーション"] = "" if _is_missing_label_text(loc0) else _s(loc0)
        rec["道路"] = rec["道路"] or "不明"
        rec["ロケーション"] = rec["ロケーション"] or "不明"
        return requests_used

    lat = float(rec["lat"])
    lon = float(rec["lon"])
    road = _s(_first_present(rec, ROAD_VALUE_KEYS))
    loc = _s(_first_present(rec, LOC_VALUE_KEYS))

    if _is_missing_label_text(road):
        road = ""
    if _is_missing_label_text(loc):
        loc = ""

    if not road or not loc:
        r0, l0, _src = osm_lookup_labels(lat, lon, osm_cache, rev_cache, overpass_cache)
        if r0 and not _is_missing_label_text(r0) and not road:
            road = _s(r0)
        if l0 and not _is_missing_label_text(l0) and not loc:
            loc = _s(l0)

    if allow_reverse_fetch and requests_used < max_requests and not road:
        overpass_hit = overpass_query_nearest_road(
            lat,
            lon,
            radius_m=overpass_radius_m,
            timeout_s=overpass_timeout_s,
            user_agent="KAIT-nearmiss-map/1.0",
        )
        if overpass_hit:
            key = _cache_keys(lat, lon)[0]
            if overpass_cache is not None:
                overpass_cache[key] = overpass_hit
            requests_used += 1
            if sleep_s > 0:
                time.sleep(sleep_s)
            if overpass_hit.get("road") and not _is_missing_label_text(overpass_hit.get("road")):
                road = _s(overpass_hit.get("road"))
            if overpass_hit.get("location") and not _is_missing_label_text(overpass_hit.get("location")) and not loc:
                loc = _s(overpass_hit.get("location"))

    if allow_reverse_fetch and requests_used < max_requests and (not road or not loc):
        js = osm_reverse_geocode(lat, lon, user_agent="KAIT-nearmiss-map/1.0")
        if js:
            r1, l1 = extract_road_loc_from_nominatim(js)
            key = _cache_keys(lat, lon)[0]
            rev_cache[key] = {
                "road": r1,
                "location": l1,
                "display_name": js.get("display_name"),
                "fetched_at": datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
            }
            requests_used += 1
            if sleep_s > 0:
                time.sleep(sleep_s)
            if r1 and not _is_missing_label_text(r1) and not road:
                road = _s(r1)
            if l1 and not _is_missing_label_text(l1) and not loc:
                loc = _s(l1)

    rec["道路"] = road or "不明"
    rec["ロケーション"] = loc or "不明"
    return requests_used


def _extract_nominatim_location_label_ja(js: Dict[str, Any]) -> Tuple[Optional[str], Optional[str]]:
    """
    Improved Nominatim address parser with Japanese-aware location extraction.

    Location priority (most granular first):
      city_district → quarter → neighbourhood → suburb → hamlet
    Combined with city context (city / town / village) to produce "市 地区" style label.

    Road priority:
      road → pedestrian → footway → cycleway → motorway → trunk → primary →
      secondary → tertiary → residential → service → unclassified

    Falls back to amenity or junction name as location if no area found.
    Returns (road, location) — either may be None.
    """
    addr = js.get("address") or {}

    # Road / street name
    road: Optional[str] = (
        addr.get("road")
        or addr.get("pedestrian")
        or addr.get("footway")
        or addr.get("cycleway")
        or addr.get("motorway")
        or addr.get("trunk")
        or addr.get("primary")
        or addr.get("secondary")
        or addr.get("tertiary")
        or addr.get("residential")
        or addr.get("service")
        or addr.get("unclassified")
    )
    # Last-resort road: first token of display_name
    if not road:
        dn = js.get("display_name")
        if isinstance(dn, str) and dn:
            road = dn.split(",")[0].strip() or None

    # Granular area (district / neighbourhood level)
    area: Optional[str] = (
        addr.get("city_district")
        or addr.get("quarter")
        or addr.get("neighbourhood")
        or addr.get("suburb")
        or addr.get("hamlet")
    )

    # City context
    city: Optional[str] = (
        addr.get("city")
        or addr.get("town")
        or addr.get("village")
        or addr.get("county")
    )

    # Build location label
    if area and city and area != city:
        loc: Optional[str] = f"{city} {area}"
    elif area:
        loc = area
    elif city:
        loc = city
    else:
        # Amenity / junction as last resort
        loc = addr.get("amenity") or addr.get("junction") or None

    return road, loc


def enrich_location_with_osm(
    rows: List[Dict[str, Any]],
    osm_cache: Dict[str, List[Any]],
    rev_cache: Dict[str, Dict[str, Any]],
    overpass_cache: Optional[Dict[str, Dict[str, Any]]],
    enable_fetch: bool,
    max_requests: int,
    sleep_s: float = 1.0,
) -> Dict[str, int]:
    """
    Targeted OSM enrichment pass: resolves '不明' location / road labels using
    cached data first, then optional Nominatim reverse-geocoding.

    Only processes rows where:
      - has_latlon is True (valid non-(0,0) coordinates present)
      - ロケーション is missing/unknown  OR  道路 is missing/unknown

    Cache strategy:
      1. Check osm_cache (label cache) via osm_lookup_labels()
      2. Check rev_cache (reverse geocode cache) keyed by _cache_keys()
      3. If enable_fetch and requests_used < max_requests: call Nominatim
         with _extract_nominatim_location_label_ja() for better Japanese labels

    Args:
        rows:             merged row dicts (mutated in-place)
        osm_cache:        osm_label_cache contents
        rev_cache:        osm_reverse_cache contents (mutated with new entries)
        overpass_cache:   overpass cache (read-only in this function)
        enable_fetch:     if True, call Nominatim for cache misses
        max_requests:     maximum Nominatim requests allowed in this pass
        sleep_s:          seconds to sleep between Nominatim requests

    Returns:
        stats dict with keys:
          location_resolved_from_cache, location_resolved_from_fetch,
          location_still_unknown, road_resolved_from_cache,
          road_resolved_from_fetch, fetch_requests_used, fetch_errors
    """
    stats: Dict[str, int] = {
        "location_resolved_from_cache": 0,
        "location_resolved_from_fetch": 0,
        "location_still_unknown": 0,
        "road_resolved_from_cache": 0,
        "road_resolved_from_fetch": 0,
        "fetch_requests_used": 0,
        "fetch_errors": 0,
    }

    for rec in rows:
        # Only process rows with valid coordinates
        if not rec.get("has_latlon"):
            continue

        lat_raw = rec.get("lat")
        lon_raw = rec.get("lon")
        if not _is_valid_coord(lat_raw, lon_raw):
            continue

        lat = float(lat_raw)  # type: ignore[arg-type]
        lon = float(lon_raw)  # type: ignore[arg-type]

        # Current values
        cur_loc = _s(rec.get("ロケーション", ""))
        cur_road = _s(rec.get("道路", ""))
        need_loc = _is_missing_label_text(cur_loc) or not cur_loc
        need_road = _is_missing_label_text(cur_road) or not cur_road

        if not need_loc and not need_road:
            continue  # Already fully labelled

        # ---- Step 1: cache lookup ----
        if need_loc or need_road:
            r_cache, l_cache, _src = osm_lookup_labels(lat, lon, osm_cache, rev_cache, overpass_cache)
            if r_cache and not _is_missing_label_text(r_cache) and need_road:
                rec["道路"] = _s(r_cache)
                cur_road = rec["道路"]
                need_road = False
                stats["road_resolved_from_cache"] += 1
            if l_cache and not _is_missing_label_text(l_cache) and need_loc:
                rec["ロケーション"] = _s(l_cache)
                cur_loc = rec["ロケーション"]
                need_loc = False
                stats["location_resolved_from_cache"] += 1

        # ---- Step 2: also check rev_cache directly with new ja extractor ----
        # (osm_lookup_labels may have used the old extract_road_loc_from_nominatim;
        #  re-check using the improved parser if the entry is present)
        if (need_loc or need_road):
            for key in _cache_keys(lat, lon):
                entry = rev_cache.get(key)
                if entry and isinstance(entry, dict) and "display_name" in entry:
                    # Re-extract with improved parser using stored nominatim-style dict
                    fake_js: Dict[str, Any] = {
                        "address": entry.get("address") or {},
                        "display_name": entry.get("display_name") or "",
                    }
                    r_ja, l_ja = _extract_nominatim_location_label_ja(fake_js)
                    if r_ja and not _is_missing_label_text(r_ja) and need_road:
                        rec["道路"] = _s(r_ja)
                        need_road = False
                        stats["road_resolved_from_cache"] += 1
                    if l_ja and not _is_missing_label_text(l_ja) and need_loc:
                        rec["ロケーション"] = _s(l_ja)
                        need_loc = False
                        stats["location_resolved_from_cache"] += 1
                    break  # found a rev_cache entry; stop trying keys

        # ---- Step 3: Nominatim fetch ----
        if (need_loc or need_road) and enable_fetch and stats["fetch_requests_used"] < max_requests:
            try:
                js = osm_reverse_geocode(lat, lon, user_agent="KAIT-nearmiss-map/1.0")
                if js:
                    r_ja, l_ja = _extract_nominatim_location_label_ja(js)
                    # Store full address in rev_cache for future runs
                    key0 = _cache_keys(lat, lon)[0]
                    rev_cache[key0] = {
                        "road": r_ja,
                        "location": l_ja,
                        "address": js.get("address") or {},
                        "display_name": js.get("display_name"),
                        "fetched_at": datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
                    }
                    stats["fetch_requests_used"] += 1
                    if sleep_s > 0:
                        time.sleep(sleep_s)
                    if r_ja and not _is_missing_label_text(r_ja) and need_road:
                        rec["道路"] = _s(r_ja)
                        need_road = False
                        stats["road_resolved_from_fetch"] += 1
                    if l_ja and not _is_missing_label_text(l_ja) and need_loc:
                        rec["ロケーション"] = _s(l_ja)
                        need_loc = False
                        stats["location_resolved_from_fetch"] += 1
                else:
                    stats["fetch_errors"] += 1
            except Exception:
                stats["fetch_errors"] += 1

        # Ensure fields are always set (fallback to 不明 if still missing)
        if need_loc:
            rec["ロケーション"] = rec.get("ロケーション") or "不明"
            stats["location_still_unknown"] += 1
        if need_road:
            rec["道路"] = rec.get("道路") or "不明"

    return stats


# --------------------------- JP progress sheet parser (keeps ALL rows) ---------------------------

def parse_jp_progress(excel: Path, sheet: str = "集計表") -> Tuple[List[Dict[str, Any]], Dict[str, str], str]:
    """
    Parse 集計表 with:
      header row: HDD-No, フォルダ, Scene No, 抽出ファイル名, ニアミス検出位置（秒）, ニアミスレベル, DX-01..DX-18
      subheader row: DX names (合流, カットイン, ...)
      risk subheader row: (大, 中, 小) under ニアミスレベル

    IMPORTANT: keep rows even if 抽出ファイル名 is empty (counts must reflect ~13k rows).
    """
    from openpyxl import load_workbook
    wb = load_workbook(excel, data_only=True)
    sheet_name = _choose_sheet_name_from_workbook(
        wb,
        sheet,
        required_headers=["HDD-No", "フォルダ", "Scene No", "抽出ファイル名", "ニアミス検出位置（秒）", "ニアミスレベル"],
        prefer_keywords=["集計", "progress", "summary"],
    )
    ws = wb[sheet_name]

    header_row = None
    for r in range(1, 401):
        vals = [ws.cell(r, c).value for c in range(1, 80)]
        if any(v == "フォルダ" for v in vals) and any(str(v).strip() == "HDD-No" for v in vals if v is not None):
            header_row = r
            break
    if header_row is None:
        raise ValueError("集計表のヘッダ行が見つかりません（HDD-No / フォルダ）。")

    col_idx: Dict[str, int] = {}
    for c in range(1, 80):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            col_idx[s] = c

    required = ["HDD-No", "フォルダ", "Scene No", "抽出ファイル名", "ニアミス検出位置（秒）", "ニアミスレベル"]
    for k in required:
        if k not in col_idx:
            raise ValueError(f"必要な列がありません: {k}")

    risk_big_c = col_idx["ニアミスレベル"]
    risk_mid_c = risk_big_c + 1
    risk_sml_c = risk_big_c + 2

    # DX columns + label map
    dx_cols: List[Tuple[str, int]] = []
    dx_name_map: Dict[str, str] = {}
    for i in range(1, 19):
        k = f"DX-{i:02d}"
        c = col_idx.get(k)
        if c is not None:
            dx_cols.append((k, c))
            nm = ws.cell(header_row + 1, c).value
            if nm is not None and str(nm).strip():
                dx_name_map[k] = str(nm).strip()

    events: List[Dict[str, Any]] = []
    cur_hdd = None
    cur_folder = None

    for r in range(header_row + 2, ws.max_row + 1):
        hdd = ws.cell(r, col_idx["HDD-No"]).value
        folder = ws.cell(r, col_idx["フォルダ"]).value
        scene = ws.cell(r, col_idx["Scene No"]).value
        vfile = ws.cell(r, col_idx["抽出ファイル名"]).value
        sec = ws.cell(r, col_idx["ニアミス検出位置（秒）"]).value

        # skip fully empty
        if hdd is None and folder is None and scene is None and vfile is None:
            continue

        if hdd is not None and str(hdd).strip() != "":
            cur_hdd = hdd
        if folder is not None and str(folder).strip() != "":
            cur_folder = folder

        folder_s_raw = str(cur_folder).strip() if cur_folder is not None else ""
        folder_s = normalize_carweek_token(folder_s_raw) or folder_s_raw
        # keep row even if folder missing? We need car_week for matching, but keep counts.
        # We'll store empty and set has_latlon false.
        scene_s = str(scene).strip() if scene is not None else ""
        vfile_s = str(vfile).strip() if vfile is not None else ""
        sec_f = _safe_float(sec)

        # risk
        risk_big_val = ws.cell(r, risk_big_c).value
        risk_mid_val = ws.cell(r, risk_mid_c).value
        risk_sml_val = ws.cell(r, risk_sml_c).value
        big = _is_mark(risk_big_val)
        mid = _is_mark(risk_mid_val)
        sml = _is_mark(risk_sml_val)
        if big:
            risk = "大"; score = 40
        elif mid:
            risk = "中"; score = 30
        elif sml:
            risk = "小"; score = 20
        else:
            risk = "不明"; score = 0

        # DX parsing rule: first ● is dx_main; ○ are dx_subs
        dx_main: Optional[str] = None
        dx_subs: List[str] = []
        for k, c in dx_cols:
            cell = ws.cell(r, c).value
            if _has_symbol(cell, "●"):
                if dx_main is None:
                    dx_main = k
            elif _has_symbol(cell, "〇"):
                dx_subs.append(k)

        scene_i = _safe_int(scene_s)
        car_id = folder_s.split("_")[0] if folder_s and "_" in folder_s else (folder_s if folder_s else "")

        events.append({
            "_src": "jp_progress",
            "hdd": str(cur_hdd).strip() if cur_hdd is not None else "",
            "car_week": folder_s,
            "car_id": car_id,
            "No": scene_i if scene_i is not None else scene_s,
            "scene_raw": scene_s,
            "video": vfile_s,
            "sec": sec_f,
            "ニアミスレベル": risk,
            "risk_level": risk,
            "score": score,
            "dx_main": dx_main,
            "dx_subs": dx_subs,
        })
    return events, dx_name_map, sheet_name


def parse_jp_progress_clean(excel: Path, sheet: str = "集計表") -> Tuple[List[Dict[str, Any]], Dict[str, str], str, Dict[str, int]]:
    """
    Clean JP parser used by the map generator.

    Keep rows that still have meaningful payload even if Scene No / video are
    blank, because those rows should contribute to the workbook total.
    """
    from openpyxl import load_workbook

    folder_header = "フォルダ"
    video_header = "抽出ファイル名"
    sec_header = "ニアミス検出位置（秒）"
    risk_header = "ニアミスレベル"

    wb = load_workbook(excel, data_only=True)
    sheet_name = _choose_sheet_name_from_workbook(
        wb,
        sheet,
        required_headers=["HDD-No", folder_header, "Scene No", video_header, sec_header, risk_header],
        prefer_keywords=["集計", "progress", "summary"],
    )
    ws = wb[sheet_name]

    header_row = None
    for r in range(1, 401):
        vals = [ws.cell(r, c).value for c in range(1, 80)]
        if any(v == folder_header for v in vals) and any(str(v).strip() == "HDD-No" for v in vals if v is not None):
            header_row = r
            break
    if header_row is None:
        raise ValueError("集計表のヘッダ行が見つかりません（HDD-No / フォルダ）。")

    col_idx: Dict[str, int] = {}
    for c in range(1, 80):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            col_idx[s] = c

    required = ["HDD-No", folder_header, "Scene No", video_header, sec_header, risk_header]
    for k in required:
        if k not in col_idx:
            raise ValueError(f"必要な列がありません: {k}")

    risk_big_c = col_idx[risk_header]
    risk_mid_c = risk_big_c + 1
    risk_sml_c = risk_big_c + 2

    dx_cols: List[Tuple[str, int]] = []
    dx_name_map: Dict[str, str] = {}
    for i in range(1, 19):
        k = f"DX-{i:02d}"
        c = col_idx.get(k)
        if c is not None:
            dx_cols.append((k, c))
            nm = ws.cell(header_row + 1, c).value
            if nm is not None and str(nm).strip():
                dx_name_map[k] = str(nm).strip()

    events: List[Dict[str, Any]] = []
    row_stats: Dict[str, int] = {
        "sheet_row_span_rows": max(0, ws.max_row - (header_row + 1)),
        "sheet_nonblank_rows": 0,
        "skipped_blank_rows": 0,
        "skipped_empty_context_rows": 0,
        "skipped_non_event_rows": 0,
    }
    cur_hdd = None
    cur_folder = None
    scan_col_end = max(32, ws.max_column)

    for r in range(header_row + 2, ws.max_row + 1):
        row_has_any_value = False
        for c in range(2, scan_col_end + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if str(v).strip() != "":
                row_has_any_value = True
                break
        if row_has_any_value:
            row_stats["sheet_nonblank_rows"] += 1
        else:
            row_stats["skipped_blank_rows"] += 1
            continue

        hdd = ws.cell(r, col_idx["HDD-No"]).value
        folder = ws.cell(r, col_idx[folder_header]).value
        scene = ws.cell(r, col_idx["Scene No"]).value
        vfile = ws.cell(r, col_idx[video_header]).value
        sec = ws.cell(r, col_idx[sec_header]).value

        if hdd is None and folder is None and scene is None and vfile is None:
            row_stats["skipped_empty_context_rows"] += 1
            continue

        if hdd is not None and str(hdd).strip() != "":
            cur_hdd = hdd
        if folder is not None and str(folder).strip() != "":
            cur_folder = folder

        folder_s_raw = str(cur_folder).strip() if cur_folder is not None else ""
        folder_s = normalize_carweek_token(folder_s_raw) or folder_s_raw
        scene_s = str(scene).strip() if scene is not None else ""
        vfile_s = str(vfile).strip() if vfile is not None else ""
        sec_f = _safe_float(sec)

        risk_big_val = ws.cell(r, risk_big_c).value
        risk_mid_val = ws.cell(r, risk_mid_c).value
        risk_sml_val = ws.cell(r, risk_sml_c).value
        big = _is_mark(risk_big_val)
        mid = _is_mark(risk_mid_val)
        sml = _is_mark(risk_sml_val)
        if big:
            risk = "大"
            score = 40
        elif mid:
            risk = "中"
            score = 30
        elif sml:
            risk = "小"
            score = 20
        else:
            risk = "不明"
            score = 0

        dx_main: Optional[str] = None
        dx_subs: List[str] = []
        has_dx_signal = False
        for k, c in dx_cols:
            cell = ws.cell(r, c).value
            if _has_symbol(cell, "●"):
                if dx_main is None:
                    dx_main = k
                has_dx_signal = True
            elif _has_symbol(cell, "〇"):
                dx_subs.append(k)
                has_dx_signal = True

        has_case_signal = bool(
            scene_s
            or vfile_s
            or sec_f is not None
            or big
            or mid
            or sml
            or has_dx_signal
        )
        if not has_case_signal:
            row_stats["skipped_non_event_rows"] += 1
            continue

        scene_i = _safe_int(scene_s)
        car_id = folder_s.split("_")[0] if folder_s and "_" in folder_s else (folder_s if folder_s else "")

        events.append({
            "_src": "jp_progress",
            "hdd": str(cur_hdd).strip() if cur_hdd is not None else "",
            "car_week": folder_s,
            "car_id": car_id,
            "No": scene_i if scene_i is not None else scene_s,
            "scene_raw": scene_s,
            "video": vfile_s,
            "sec": sec_f,
            "ニアミスレベル": risk,
            "risk_level": risk,
            "score": score,
            "dx_main": dx_main,
            "dx_subs": dx_subs,
        })

    row_stats["kept_rows"] = len(events)
    return events, dx_name_map, sheet_name, row_stats

# --------------------------- log matching (reuse v67 behavior) ---------------------------

TS_RE = re.compile(r"(\d{4})\.(\d{2})\.(\d{2})\s+(\d{2})\.(\d{2})\.(\d{2})\.(\d{3})")
CARWEEK_RE = re.compile(r"(?P<car>\d{2})_(?P<start>\d{6})-(?P<end>\d{6})")
HDD_RE = re.compile(r"(?:^|[\\/])HDD(?P<car>\d{2})(?:[\\/]|$)", re.IGNORECASE)
TRIP_LOG_HEADER_RE = re.compile(r"^\s*day\s*,\s*time\s*,", re.IGNORECASE)
DATE_FMT_HINTS = [
    "%Y/%m/%d %H:%M:%S.%f",
    "%Y/%m/%d %H:%M:%S",
]
STRUCTURED_TIME_COLUMNS = [
    "timestamp",
    "time",
    "ts",
    "datetime",
    "DateTime",
    "date_time",
    "recorded_at",
    "日時",
]
STRUCTURED_LAT_COLUMNS = [
    "latitude",
    "lat",
    "LAT",
    "LATITUDE",
    "center_latitude",
    "LATITUDEFRONT",
    "緯度",
]
STRUCTURED_LON_COLUMNS = [
    "longitude",
    "lon",
    "lng",
    "LON",
    "LONGITUDE",
    "center_longitude",
    "LONGITUDEFRONT",
    "経度",
]
STRUCTURED_SPEED_COLUMNS = [
    "speed",
    "Speed",
    "SPEED",
    "vel",
    "velocity",
    "速度",
]

def parse_timestamp_from_video_name(video_name: str) -> Optional[datetime.datetime]:
    m = TS_RE.search(video_name)
    if not m:
        return None
    y,mo,d,hh,mm,ss,ms = map(int, m.groups())
    try:
        return datetime.datetime(y,mo,d,hh,mm,ss,ms*1000)
    except Exception:
        return None

def _parse_timestamp_text(value: Any) -> Optional[datetime.datetime]:
    s = _s(value)
    if not s:
        return None
    for fmt in DATE_FMT_HINTS:
        try:
            return datetime.datetime.strptime(s, fmt)
        except Exception:
            continue
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.isna(ts):
            return None
        if getattr(ts, "tzinfo", None) is not None:
            ts = ts.tz_localize(None)
        return ts.to_pydatetime()
    except Exception:
        return None

def parse_timestamp_from_log_path(path: str) -> Optional[datetime.datetime]:
    return parse_timestamp_from_video_name(Path(path).name) or parse_timestamp_from_video_name(path)

def _looks_like_yymmdd(value: str) -> bool:
    if not value or len(value) != 6 or not value.isdigit():
        return False
    try:
        datetime.datetime.strptime(value, "%y%m%d")
        return True
    except Exception:
        return False

def _repair_carweek_end(start_ymd: str, tail_digits: str) -> Optional[str]:
    if not start_ymd:
        return None
    digits = re.sub(r"\D", "", tail_digits or "")
    if not digits:
        return None
    candidates: List[str] = []
    if len(digits) >= 6:
        candidates.append(digits[:6])
    else:
        prefix = start_ymd[:max(0, 6 - len(digits))]
        candidates.append((prefix + digits)[-6:])
        if len(digits) <= 4:
            candidates.append(start_ymd[:2] + digits.rjust(4, start_ymd[2:2 + max(0, 4 - len(digits))]))
        if len(digits) == 5:
            candidates.append(start_ymd[:1] + digits)
    for cand in candidates:
        if _looks_like_yymmdd(cand):
            return cand
    return candidates[0] if candidates else None

def extract_car_id_from_token(raw: Any) -> str:
    s = _s(raw).replace("\\", "/")
    m = CARWEEK_RE.search(s)
    if m:
        return m.group("car")
    m = re.search(r"(^|/)(?P<car>\d{1,2})_", s)
    if m:
        return m.group("car").zfill(2)
    m = HDD_RE.search(s)
    if m:
        return m.group("car")
    return ""

def normalize_carweek_token(raw: Any) -> str:
    s = _s(raw).replace("\\", "/")
    m = CARWEEK_RE.search(s)
    if m:
        return f"{m.group('car')}_{m.group('start')}-{m.group('end')}"

    car_id = extract_car_id_from_token(s)
    digits = re.findall(r"\d+", s)
    if digits and car_id and digits[0].zfill(2) == car_id:
        digits = digits[1:]
    if not digits:
        return ""

    start_ymd = next((chunk[:6] for chunk in digits if len(chunk) >= 6 and _looks_like_yymmdd(chunk[:6])), None)
    if not start_ymd:
        combined = "".join(digits)
        if len(combined) >= 6 and _looks_like_yymmdd(combined[:6]):
            start_ymd = combined[:6]
    if not start_ymd:
        return ""

    tail_chunks: List[str] = []
    consumed_start = False
    for chunk in digits:
        if not consumed_start and len(chunk) >= 6 and chunk[:6] == start_ymd:
            consumed_start = True
            if len(chunk) > 6:
                tail_chunks.append(chunk[6:])
            continue
        if consumed_start:
            tail_chunks.append(chunk)
    end_ymd = _repair_carweek_end(start_ymd, "".join(tail_chunks))
    if car_id and start_ymd and end_ymd:
        return f"{car_id}_{start_ymd}-{end_ymd}"
    return ""

def _build_log_entry(path_text: str) -> Dict[str, Any]:
    path_str = str(Path(path_text))
    carweek = ""
    for part in Path(path_str).parts:
        if isinstance(part, str) and CARWEEK_RE.fullmatch(part):
            carweek = part
    norm_carweek = normalize_carweek_token(carweek or path_str)
    ts = parse_timestamp_from_log_path(path_str)
    return {
        "path": path_str,
        "carweek": carweek or norm_carweek,
        "norm_carweek": norm_carweek or carweek,
        "car_id": extract_car_id_from_token(carweek or path_str),
        "ts": ts.isoformat() if ts else None,
    }

def _prepare_log_index(payload_data: Any) -> Dict[str, Any]:
    if isinstance(payload_data, dict) and "by_carweek" in payload_data:
        by_carweek_raw = payload_data.get("by_carweek") or {}
        entries_raw = payload_data.get("entries") or []
    elif isinstance(payload_data, dict):
        by_carweek_raw = payload_data
        entries_raw = []
    else:
        by_carweek_raw = {}
        entries_raw = []

    path_to_entry: Dict[str, Dict[str, Any]] = {}
    for raw in entries_raw:
        if not isinstance(raw, dict):
            continue
        path_str = _s(raw.get("path"))
        if not path_str:
            continue
        ts = _parse_timestamp_text(raw.get("ts")) or parse_timestamp_from_log_path(path_str)
        path_to_entry[path_str] = {
            "path": path_str,
            "carweek": _s(raw.get("carweek")),
            "norm_carweek": _s(raw.get("norm_carweek")) or normalize_carweek_token(raw.get("carweek") or path_str),
            "car_id": _s(raw.get("car_id")) or extract_car_id_from_token(raw.get("carweek") or path_str),
            "ts": ts,
            "ts_key": ts.timestamp() if ts is not None else None,
        }

    by_carweek: Dict[str, List[str]] = {}
    for key, files in by_carweek_raw.items():
        if not isinstance(files, list):
            continue
        clean_key = _s(key)
        by_carweek.setdefault(clean_key, [])
        for path_text in files:
            path_str = str(Path(path_text))
            if path_str not in by_carweek[clean_key]:
                by_carweek[clean_key].append(path_str)
            if path_str not in path_to_entry:
                raw_entry = _build_log_entry(path_str)
                ts = _parse_timestamp_text(raw_entry.get("ts"))
                path_to_entry[path_str] = {
                    "path": path_str,
                    "carweek": raw_entry.get("carweek") or clean_key,
                    "norm_carweek": raw_entry.get("norm_carweek") or normalize_carweek_token(clean_key),
                    "car_id": raw_entry.get("car_id") or extract_car_id_from_token(clean_key),
                    "ts": ts,
                    "ts_key": ts.timestamp() if ts is not None else None,
                }

    by_norm_carweek: Dict[str, List[Dict[str, Any]]] = {}
    by_car: Dict[str, List[Dict[str, Any]]] = {}
    all_entries: List[Dict[str, Any]] = []
    by_carweek_entries: Dict[str, List[Dict[str, Any]]] = {}

    def _entry_sort_key(entry: Dict[str, Any]) -> Tuple[float, str]:
        return (entry.get("ts_key") if entry.get("ts_key") is not None else float("inf"), entry["path"])

    for entry in path_to_entry.values():
        all_entries.append(entry)
        cw = _s(entry.get("carweek"))
        norm_cw = _s(entry.get("norm_carweek"))
        car_id = _s(entry.get("car_id"))
        if cw:
            by_carweek_entries.setdefault(cw, []).append(entry)
            by_carweek.setdefault(cw, [])
            if entry["path"] not in by_carweek[cw]:
                by_carweek[cw].append(entry["path"])
        if norm_cw:
            by_norm_carweek.setdefault(norm_cw, []).append(entry)
        if car_id:
            by_car.setdefault(car_id, []).append(entry)

    all_entries.sort(key=_entry_sort_key)
    for group in list(by_carweek_entries.values()) + list(by_norm_carweek.values()) + list(by_car.values()):
        group.sort(key=_entry_sort_key)

    return {
        "by_carweek": by_carweek,
        "by_carweek_entries": by_carweek_entries,
        "by_norm_carweek": by_norm_carweek,
        "by_car": by_car,
        "all_entries": all_entries,
    }

def _read_logs_manifest(logs_root: Path, manifest_path: Path) -> List[str]:
    text = manifest_path.read_text(encoding="utf-8")
    files: List[str] = []
    seen = set()
    for line in text.splitlines():
        raw = line.strip().strip('"').strip("'")
        if not raw or raw.startswith("#"):
            continue
        p = Path(raw)
        if not p.is_absolute():
            p = logs_root / p
        norm = str(p)
        if norm not in seen:
            seen.add(norm)
            files.append(norm)
    return files

def build_carweek_file_index(
    logs_root: Path,
    logs_glob: str,
    cache_path: Path,
    logs_manifest: Optional[Path] = None,
) -> Dict[str, Any]:
    import glob

    manifest_meta = None
    if logs_manifest and logs_manifest.exists():
        manifest_meta = {
            "path": str(logs_manifest),
            "mtime_ns": int(logs_manifest.stat().st_mtime_ns),
        }

    if cache_path.exists():
        try:
            data = json.loads(cache_path.read_text(encoding="utf-8"))
            meta = data.get("meta") if isinstance(data, dict) else None
            same_request = (
                isinstance(meta, dict)
                and meta.get("logs_root") == str(logs_root)
                and meta.get("logs_glob") == logs_glob
                and meta.get("logs_manifest") == manifest_meta
            )
            if same_request and isinstance(data, dict) and "data" in data:
                return _prepare_log_index(data["data"])
        except Exception:
            pass

    if manifest_meta:
        files = _read_logs_manifest(logs_root, logs_manifest)
    else:
        patt = os.path.join(str(logs_root), logs_glob)
        files = glob.glob(patt, recursive=True)

    by: Dict[str, List[str]] = {}
    entry_payload: List[Dict[str, Any]] = []
    for fp in sorted(set(files)):
        entry = _build_log_entry(fp)
        cw = _s(entry.get("carweek")) or _s(entry.get("norm_carweek"))
        if cw:
            by.setdefault(cw, []).append(entry["path"])
        entry_payload.append(entry)
    payload = {
        "meta": {
            "logs_root": str(logs_root),
            "logs_glob": logs_glob,
            "logs_manifest": manifest_meta,
            "files": len(files),
            "carweeks": len(by),
        },
        "data": {
            "by_carweek": by,
            "entries": entry_payload,
        },
    }
    cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return _prepare_log_index(payload["data"])

def _rank_entries_near_target(
    entries: List[Dict[str, Any]],
    target_ts: datetime.datetime,
    limit: int,
    max_delta_s: Optional[float] = None,
) -> List[str]:
    target_key = target_ts.timestamp()
    ranked: List[Tuple[float, str]] = []
    fallback: List[str] = []
    for entry in entries:
        path_str = _s(entry.get("path"))
        ts_key = entry.get("ts_key")
        if not path_str:
            continue
        if ts_key is None:
            fallback.append(path_str)
            continue
        delta_s = abs(float(ts_key) - target_key)
        if max_delta_s is not None and delta_s > max_delta_s:
            continue
        ranked.append((delta_s, path_str))
    ranked.sort(key=lambda item: (item[0], item[1]))
    out = [path for _, path in ranked[:limit]]
    for path_str in fallback:
        if len(out) >= limit:
            break
        if path_str not in out:
            out.append(path_str)
    return out

def candidate_log_files_for_event(
    log_index: Dict[str, Any],
    car_week: str,
    car_id: str,
    target_ts: datetime.datetime,
    limit: int = 24,
) -> List[str]:
    out: List[str] = []
    seen: set[str] = set()

    def add_paths(paths: List[str]) -> None:
        for path_str in paths:
            if not path_str or path_str in seen:
                continue
            seen.add(path_str)
            out.append(path_str)
            if len(out) >= limit:
                return

    cw = _s(car_week)
    norm_cw = normalize_carweek_token(cw)
    cid = _s(car_id) or extract_car_id_from_token(cw)

    exact_entries = log_index.get("by_carweek_entries", {}).get(cw, [])
    if exact_entries:
        add_paths(_rank_entries_near_target(exact_entries, target_ts, limit=12))
    if norm_cw and norm_cw != cw:
        add_paths(_rank_entries_near_target(log_index.get("by_norm_carweek", {}).get(norm_cw, []), target_ts, limit=12))
    if cid:
        add_paths(_rank_entries_near_target(log_index.get("by_car", {}).get(cid, []), target_ts, limit=16, max_delta_s=45 * 86400.0))
    if len(out) < limit:
        add_paths(_rank_entries_near_target(log_index.get("all_entries", []), target_ts, limit=max(6, limit - len(out)), max_delta_s=3 * 86400.0))
    if not out:
        add_paths(log_index.get("by_carweek", {}).get(cw, [])[:limit])
    return out[:limit]

def _open_text_any(path_text: str):
    for enc in ("utf-8-sig", "cp932", "utf-8"):
        try:
            return open(path_text, "r", encoding=enc, newline="")
        except Exception:
            continue
    return open(path_text, "r", encoding="utf-8", errors="replace", newline="")

@lru_cache(maxsize=192)
def _load_log_rows(path_text: str) -> Tuple[Tuple[float, str, float, float, Optional[float]], ...]:
    path_str = str(Path(path_text))
    rows: List[Tuple[float, str, float, float, Optional[float]]] = []

    try:
        with _open_text_any(path_str) as f:
            sample = f.read(512)
    except Exception:
        sample = ""

    looks_like_trip_log = path_str.lower().endswith(".txt") or bool(TRIP_LOG_HEADER_RE.search(sample))
    if looks_like_trip_log:
        try:
            with _open_text_any(path_str) as f:
                reader = csv.reader(f)
                for rec in reader:
                    if not rec:
                        continue
                    head = ",".join(rec[:2]).strip()
                    if TRIP_LOG_HEADER_RE.match(head):
                        continue
                    if len(rec) < 8:
                        continue
                    ts = _parse_timestamp_text(f"{rec[0].strip()} {rec[1].strip()}")
                    lat = _safe_float(rec[5])
                    lon = _safe_float(rec[6])
                    if ts is None or not _is_valid_coord(lat, lon):
                        continue
                    speed = _safe_float(rec[7])
                    rows.append((ts.timestamp(), ts.isoformat(), float(lat), float(lon), speed))
        except Exception:
            rows = []

    if not rows:
        try:
            df = pd.read_csv(path_str, low_memory=False)
            time_col = next((cand for cand in STRUCTURED_TIME_COLUMNS if cand in df.columns), None)
            lat_col = next((cand for cand in STRUCTURED_LAT_COLUMNS if cand in df.columns), None)
            lon_col = next((cand for cand in STRUCTURED_LON_COLUMNS if cand in df.columns), None)
            spd_col = next((cand for cand in STRUCTURED_SPEED_COLUMNS if cand in df.columns), None)
            if time_col and lat_col and lon_col:
                ts_series = pd.to_datetime(df[time_col], errors="coerce", utc=False)
                try:
                    ts_series = ts_series.dt.tz_localize(None)
                except Exception:
                    pass
                for idx, ts in ts_series.items():
                    if pd.isna(ts):
                        continue
                    lat = _safe_float(df.at[idx, lat_col])
                    lon = _safe_float(df.at[idx, lon_col])
                    if not _is_valid_coord(lat, lon):
                        continue
                    speed = _safe_float(df.at[idx, spd_col]) if spd_col else None
                    dt = ts.to_pydatetime() if hasattr(ts, "to_pydatetime") else ts
                    rows.append((dt.timestamp(), dt.isoformat(), float(lat), float(lon), speed))
        except Exception:
            rows = []

    rows.sort(key=lambda item: item[0])
    return tuple(rows)

def _find_in_logfile(path_text: str, target_ts: datetime.datetime, tolerance_s: float) -> Optional[Dict[str, Any]]:
    rows = _load_log_rows(path_text)
    if not rows:
        return None
    target_key = target_ts.timestamp()
    keys = [row[0] for row in rows]
    idx = bisect.bisect_left(keys, target_key)
    best: Optional[Tuple[float, Tuple[float, str, float, float, Optional[float]]]] = None
    for cand_idx in (idx - 1, idx, idx + 1):
        if cand_idx < 0 or cand_idx >= len(rows):
            continue
        row = rows[cand_idx]
        delta_s = abs(row[0] - target_key)
        if delta_s > tolerance_s:
            continue
        if best is None or delta_s < best[0]:
            best = (delta_s, row)
    if best is None:
        return None
    delta_s, row = best
    return {
        "lat": row[2],
        "lon": row[3],
        "speed": row[4],
        "delta_s": float(delta_s),
        "ts": row[1],
    }

def find_latlon_in_logfiles(carweek_files: List[str], target_ts: datetime.datetime, tolerance_s: float) -> Optional[Dict[str, Any]]:
    """
    Search through candidate log files for closest timestamp within tolerance.
    Supports both structured CSV logs and trip-style text logs with repeated headers.
    """
    best: Optional[Dict[str, Any]] = None
    best_dt: Optional[float] = None
    for fp in carweek_files:
        cand = _find_in_logfile(fp, target_ts, tolerance_s)
        if not cand:
            continue
        dt = _safe_float(cand.get("delta_s"))
        if dt is None:
            continue
        if best is None or best_dt is None or dt < best_dt:
            best = cand
            best_dt = dt
    return best

def _parse_latlon_cache_key(key: Any) -> Tuple[str, str, Optional[float]]:
    s = _s(key)
    first = s.find("|")
    last = s.rfind("|")
    if first < 0 or last <= first:
        return "", "", None
    return s[:first], s[first + 1:last], _safe_float(s[last + 1:])

def _video_sec_cache_key(video: str, sec: Optional[float]) -> str:
    sec_s = f"{float(sec):.1f}" if sec is not None else ""
    return f"{_s(video)}|{sec_s}"

def build_video_sec_latlon_cache_index(latlon_cache: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for raw_key, raw_val in latlon_cache.items():
        if not isinstance(raw_val, dict):
            continue
        _, video, sec = _parse_latlon_cache_key(raw_key)
        if not video or sec is None:
            continue
        lat = _safe_float(raw_val.get("lat"))
        lon = _safe_float(raw_val.get("lon"))
        if not _is_valid_coord(lat, lon):
            continue
        key = _video_sec_cache_key(video, sec)
        prev = out.get(key)
        prev_dt = _safe_float(prev.get("delta_s")) if isinstance(prev, dict) else None
        cur_dt = _safe_float(raw_val.get("delta_s"))
        if prev is None or (cur_dt is not None and (prev_dt is None or cur_dt < prev_dt)):
            out[key] = raw_val
    return out

def merge_legacy_labels_into_event(ev: Dict[str, Any], legacy_hit: Optional[Dict[str, Any]]) -> None:
    if not isinstance(legacy_hit, dict):
        return
    for key in ["道路", "ロケーション", "平均速度", "最低速度", "最高速度", "snap_dist_m", "格納先", "分類", "分類.1", "分類名", "dx_code", "dx_name"]:
        if key not in legacy_hit or _is_missing_value(legacy_hit.get(key)):
            continue
        cur = ev.get(key)
        if key in {"道路", "ロケーション", "分類名"}:
            if _is_missing_label_text(cur):
                ev[key] = legacy_hit[key]
        elif _is_missing_value(cur):
            ev[key] = legacy_hit[key]

# --------------------------- Legacy 1018 loader ---------------------------

def load_legacy_1018(excel: Path, sheet: str = "AllPoints_1018") -> Tuple[pd.DataFrame, str]:
    from openpyxl import load_workbook

    wb = load_workbook(excel, read_only=True, data_only=True)
    sheet_name = _choose_sheet_name_from_workbook(
        wb,
        sheet,
        required_headers=["No", "car_week", "center_latitude", "center_longitude", "LATITUDEFRONT", "LONGITUDEFRONT", "緯度", "経度"],
        prefer_keywords=["allpoints", "1018", "legacy"],
    )
    df = pd.read_excel(excel, sheet_name=sheet_name)
    return df, sheet_name

# --------------------------- exposure loader ---------------------------

def load_exposure(exposure_dir: Path) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Returns:
      exposure_km_by_label (route_group_main_v2 or route_group_main)
      totals_overall (json)
      totals_by_car (list)
      totals_by_carweek (list)
    """
    exposure_km_by_label: Dict[str, Dict[str, Any]] = {}
    totals_overall: Dict[str, Any] = {}
    totals_by_car: List[Dict[str, Any]] = []
    totals_by_carweek: List[Dict[str, Any]] = []

    try:
        p = exposure_dir / "totals_overall.json"
        if p.exists():
            totals_overall = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        totals_overall = {}
    try:
        p = exposure_dir / "totals_by_car.csv"
        if p.exists():
            totals_by_car = pd.read_csv(p).to_dict(orient="records")
    except Exception:
        totals_by_car = []
    try:
        p = exposure_dir / "totals_by_carweek.csv"
        if p.exists():
            totals_by_carweek = pd.read_csv(p).to_dict(orient="records")
    except Exception:
        totals_by_carweek = []

    totals_overall = _resolve_overall_metric_summary(
        totals_overall,
        [
            ("totals_by_carweek.csv", totals_by_carweek),
            ("totals_by_car.csv", totals_by_car),
        ],
    )

    # label exposure if available
    # Prefer pre-aggregated road groups, then label totals, then carweek-label totals.
    for fname in ["exposure_by_roadgroup.csv", "totals_by_label.csv", "totals_by_carweek_label.csv"]:
        p = exposure_dir / fname
        if p.exists():
            try:
                d = pd.read_csv(p)
                label_col = None
                km_col = None
                dur_col = None
                steps_col = None
                expected_col = None
                imputed_col = None
                frozen_col = None
                gap_col = None
                for c in d.columns:
                    cl = c.lower()
                    if cl in ["label", "label_name", "road_group", "route_group", "route_group_main", "route_group_main_v2"]:
                        label_col = c
                    if cl in ["dist_geo_clamped_km", "distance_km", "dist_hv_clamped_km", "dist_speed_km", "km", "total_km"]:
                        km_col = c
                    if cl in ["duration_h", "time_h", "hours"]:
                        dur_col = c
                    if cl == "steps":
                        steps_col = c
                    if cl == "expected_distance_km":
                        expected_col = c
                    if cl == "imputed_km":
                        imputed_col = c
                    if cl == "frozen_km":
                        frozen_col = c
                    if cl == "time_gap_h":
                        gap_col = c
                if label_col and km_col:
                    for _, row in d.iterrows():
                        lab = _s(row[label_col])
                        km = _safe_float(row[km_col]) or 0.0
                        if lab:
                            cur = exposure_km_by_label.get(lab, {})
                            expected_km = _safe_float(row[expected_col]) if expected_col else None
                            if expected_km is None:
                                expected_km = float(km)
                            rec = {
                                "distance_km": float(km),
                                "expected_distance_km": float(expected_km),
                                "imputed_km": _safe_float(row[imputed_col]) if imputed_col else _safe_float(cur.get("imputed_km")),
                                "frozen_km": _safe_float(row[frozen_col]) if frozen_col else _safe_float(cur.get("frozen_km")),
                                "time_gap_h": _safe_float(row[gap_col]) if gap_col else _safe_float(cur.get("time_gap_h")),
                                "duration_h": _safe_float(row[dur_col]) if dur_col else _safe_float(cur.get("duration_h")),
                                "steps": _safe_int(row[steps_col]) if steps_col else _safe_int(cur.get("steps")),
                                "source": fname,
                            }
                            rec_rank = float(rec.get("expected_distance_km") or rec.get("distance_km") or 0.0)
                            cur_rank = float(cur.get("expected_distance_km") or cur.get("distance_km") or 0.0)
                            if lab not in exposure_km_by_label or rec_rank > cur_rank:
                                exposure_km_by_label[lab] = rec
            except Exception:
                pass

    return exposure_km_by_label, totals_overall, totals_by_car, totals_by_carweek


def build_legacy_route_group_mapping(df: pd.DataFrame) -> Dict[str, str]:
    group_cols = [c for c in ["route_group_main_v2", "route_group_main", "route_group", "route_name"] if c in df.columns]
    if not group_cols:
        return {}
    if "route_name" in df.columns:
        key_col = "route_name"
    elif "route_ref" in df.columns:
        key_col = "route_ref"
    else:
        return {}
    target_col = "route_group_main_v2" if "route_group_main_v2" in df.columns else group_cols[0]
    tmp = df[[key_col, target_col]].copy()
    tmp[key_col] = tmp[key_col].astype(str).str.strip()
    tmp[target_col] = tmp[target_col].astype(str).str.strip()
    tmp = tmp[(tmp[key_col] != "") & (tmp[key_col].str.lower() != "nan")]
    tmp = tmp[(tmp[target_col] != "") & (tmp[target_col].str.lower() != "nan")]
    tmp = tmp[tmp[target_col].str.lower() != "unknown"]
    if tmp.empty:
        return {}
    gb = tmp.groupby([key_col, target_col]).size().reset_index(name="n")
    gb = gb.sort_values(["n"], ascending=False)
    mapping: Dict[str, str] = {}
    for key, sub in gb.groupby(key_col, sort=False):
        mapping[_s(key)] = _s(sub.iloc[0][target_col])
    return mapping


def _metric_distance_km(rec: Dict[str, Any]) -> Optional[float]:
    for k in ["dist_geo_clamped_km", "distance_km", "dist_hv_clamped_km", "dist_speed_km", "km", "total_km"]:
        if k in rec:
            v = _safe_float(rec.get(k))
            if v is not None and not math.isnan(v):
                return float(v)
    return None


def _metric_duration_h(rec: Dict[str, Any]) -> Optional[float]:
    for k in ["duration_h", "time_h", "hours"]:
        if k in rec:
            v = _safe_float(rec.get(k))
            if v is not None and not math.isnan(v):
                return float(v)
    return None


def _metric_steps(rec: Dict[str, Any]) -> Optional[int]:
    for k in ["steps", "count"]:
        if k in rec:
            v = _safe_int(rec.get(k))
            if v is not None:
                return int(v)
    return None


def _metric_float(rec: Dict[str, Any], keys: List[str]) -> Optional[float]:
    for k in keys:
        if k in rec:
            v = _safe_float(rec.get(k))
            if v is not None and not math.isnan(v):
                return float(v)
    return None


def _metric_int(rec: Dict[str, Any], keys: List[str]) -> Optional[int]:
    for k in keys:
        if k in rec:
            v = _safe_int(rec.get(k))
            if v is not None:
                return int(v)
    return None


def _empty_scope_total() -> Dict[str, Any]:
    return {
        "distance_km": 0.0,
        "duration_h": 0.0,
        "steps": 0,
        "dist_speed_km": 0.0,
        "dist_geo_clamped_km": 0.0,
        "imputed_km": 0.0,
        "steps_imputed": 0,
        "frozen_km": 0.0,
        "time_gap_h": 0.0,
        "time_frozen_h": 0.0,
        "time_active_h": 0.0,
        "time_stopped_h": 0.0,
        "time_unavailable_h": 0.0,
        "time_unavailable_pct": 0.0,
        "imputed_step_pct": 0.0,
        "observed_distance_pct": 0.0,
        "estimated_missing_distance_pct": 0.0,
        "expected_distance_km": 0.0,
    }


def _accumulate_scope_total(cur: Dict[str, Any], rec: Dict[str, Any]) -> None:
    dist = _metric_distance_km(rec)
    dur = _metric_duration_h(rec)
    steps = _metric_steps(rec)
    geo = _metric_float(rec, ["dist_geo_clamped_km", "distance_km", "dist_hv_clamped_km", "dist_speed_km", "km", "total_km"])
    sensor = _metric_float(rec, ["dist_speed_km"])
    imputed = _metric_float(rec, ["imputed_km"])
    frozen = _metric_float(rec, ["frozen_km"])
    time_gap = _metric_float(rec, ["time_gap_h"])
    time_frozen = _metric_float(rec, ["time_frozen_h"])
    time_active = _metric_float(rec, ["time_active_h"])
    time_stopped = _metric_float(rec, ["time_stopped_h"])
    steps_imputed = _metric_int(rec, ["steps_imputed"])

    if dist is not None:
        cur["distance_km"] += float(dist)
    if dur is not None:
        cur["duration_h"] += float(dur)
    if steps is not None:
        cur["steps"] += int(steps)
    if sensor is not None:
        cur["dist_speed_km"] += float(sensor)
    if geo is not None:
        cur["dist_geo_clamped_km"] += float(geo)
    if imputed is not None:
        cur["imputed_km"] += float(imputed)
    if steps_imputed is not None:
        cur["steps_imputed"] += int(steps_imputed)
    if frozen is not None:
        cur["frozen_km"] += float(frozen)
    if time_gap is not None:
        cur["time_gap_h"] += float(time_gap)
    if time_frozen is not None:
        cur["time_frozen_h"] += float(time_frozen)
    if time_active is not None:
        cur["time_active_h"] += float(time_active)
    if time_stopped is not None:
        cur["time_stopped_h"] += float(time_stopped)


def _finalize_scope_total(cur: Dict[str, Any]) -> Dict[str, Any]:
    geo = float(cur.get("dist_geo_clamped_km") or cur.get("distance_km") or 0.0)
    imputed = float(cur.get("imputed_km") or 0.0)
    steps = int(cur.get("steps") or 0)
    steps_imputed = int(cur.get("steps_imputed") or 0)
    duration_h = float(cur.get("duration_h") or 0.0)
    time_frozen_h = float(cur.get("time_frozen_h") or 0.0)
    time_gap_h = float(cur.get("time_gap_h") or 0.0)
    time_unavailable_h = time_frozen_h + time_gap_h
    expected_distance_km = geo + imputed
    cur["time_unavailable_h"] = time_unavailable_h
    cur["time_unavailable_pct"] = round((time_unavailable_h / duration_h * 100.0) if duration_h else 0.0, 3)
    cur["expected_distance_km"] = expected_distance_km
    cur["observed_distance_pct"] = round((geo / expected_distance_km * 100.0) if expected_distance_km else 0.0, 3)
    cur["estimated_missing_distance_pct"] = round((imputed / expected_distance_km * 100.0) if expected_distance_km else 0.0, 3)
    cur["imputed_step_pct"] = round((steps_imputed / steps * 100.0) if steps else 0.0, 3)
    return cur


def _summarize_metric_records(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    summary = {"distance_km": 0.0, "duration_h": 0.0, "steps": 0}
    saw_metric = False
    for rec in records or []:
        if not isinstance(rec, dict):
            continue
        dist = _metric_distance_km(rec)
        dur = _metric_duration_h(rec)
        steps = _metric_steps(rec)
        if dist is not None:
            summary["distance_km"] += float(dist)
            saw_metric = True
        if dur is not None:
            summary["duration_h"] += float(dur)
            saw_metric = True
        if steps is not None:
            summary["steps"] += int(steps)
            saw_metric = True
    if not saw_metric:
        return {}
    return summary


def _has_meaningful_metric_summary(rec: Optional[Dict[str, Any]]) -> bool:
    if not isinstance(rec, dict):
        return False
    dist = _metric_distance_km(rec) or 0.0
    dur = _metric_duration_h(rec) or 0.0
    steps = _metric_steps(rec) or 0
    return abs(float(dist)) > 1e-9 or abs(float(dur)) > 1e-9 or int(steps) > 0


def _resolve_overall_metric_summary(
    primary: Optional[Dict[str, Any]],
    fallback_sources: List[Tuple[str, List[Dict[str, Any]]]],
) -> Dict[str, Any]:
    base = dict(primary or {})
    if _has_meaningful_metric_summary(base):
        return base

    for source_name, records in fallback_sources:
        agg = _summarize_metric_records(records)
        if not _has_meaningful_metric_summary(agg):
            continue
        resolved = dict(base)
        resolved["distance_km"] = float(agg.get("distance_km") or 0.0)
        resolved["dist_geo_clamped_km"] = float(agg.get("distance_km") or 0.0)
        resolved["duration_h"] = float(agg.get("duration_h") or 0.0)
        resolved["steps"] = int(agg.get("steps") or 0)
        resolved["_fallback_source"] = source_name
        return resolved

    return base


def _scope_key_from_parts(parts: List[str]) -> str:
    return "__".join(parts)


def load_route_metrics(exposure_dir: Path, legacy_df: pd.DataFrame, top_n: int = 12) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    unknown_highway = "\u9ad8\u901f\u9053\u8def\uff08\u9053\u8def\u540d\u306a\u3057\uff09"
    unknown_general = "\u4e00\u822c\u9053\uff08\u9053\u8def\u540d\u306a\u3057\uff09"
    unknown_group = "Unknown"
    name_to_group = build_legacy_route_group_mapping(legacy_df)

    def _load_records(fname: str) -> List[Dict[str, Any]]:
        p = exposure_dir / fname
        if not p.exists():
            return []
        try:
            return pd.read_csv(p).to_dict(orient="records")
        except Exception:
            return []

    def _map_label_to_group(label: str, rec: Dict[str, Any]) -> str:
        label = _s(label)
        if not label:
            return unknown_group
        if label in name_to_group:
            return name_to_group[label]
        group_hint = _s(rec.get("group"))
        if group_hint.lower() == "highway":
            return unknown_highway
        if group_hint:
            return unknown_general
        return unknown_group

    def _aggregate_metric_rows(
        records: List[Dict[str, Any]],
        scope_fields: List[str],
        name_field: str,
        name_transform,
    ) -> Any:
        def _finalize_row_payload(row: Dict[str, Any]) -> Dict[str, Any]:
            members = row.get("_members", set())
            base = {k: v for k, v in row.items() if k != "_members"}
            out = _finalize_scope_total(base)
            out["segment_count"] = len(members) if members else 1
            return out

        grouped: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for rec in records:
            name_raw = _s(rec.get(name_field))
            if not name_raw or name_raw.lower() == "nan":
                continue
            name = _s(name_transform(name_raw, rec))
            if not name or name.lower() == "nan":
                continue
            scope_parts = [_s(rec.get(field)) for field in scope_fields]
            if scope_fields and any(not p for p in scope_parts):
                continue
            dist = _metric_distance_km(rec)
            if dist is None:
                continue
            scope_key = _scope_key_from_parts(scope_parts) if scope_fields else "__all__"
            gk = (scope_key, name)
            cur = grouped.get(gk)
            if cur is None:
                cur = _empty_scope_total()
                cur["name"] = name
                cur["_members"] = set()
                grouped[gk] = cur
            _accumulate_scope_total(cur, rec)
            cur["_members"].add(name_raw)

        if not scope_fields:
            rows = [_finalize_row_payload(v) for (_, _), v in grouped.items()]
            rows.sort(key=lambda row: (-float(row.get("expected_distance_km") or row.get("distance_km") or 0.0), row.get("name") or ""))
            return rows[:top_n]

        scoped: Dict[str, List[Dict[str, Any]]] = {}
        for (scope_key, _), row in grouped.items():
            scoped.setdefault(scope_key, []).append(_finalize_row_payload(row))
        for scope_key, rows in scoped.items():
            rows.sort(key=lambda row: (-float(row.get("expected_distance_km") or row.get("distance_km") or 0.0), row.get("name") or ""))
            scoped[scope_key] = rows[:top_n]
        return scoped

    def _build_scope_totals(records: List[Dict[str, Any]], scope_fields: List[str]) -> Dict[str, Dict[str, Any]]:
        out: Dict[str, Dict[str, Any]] = {}
        for rec in records:
            parts = [_s(rec.get(field)) for field in scope_fields]
            if any(not p for p in parts):
                continue
            scope_key = _scope_key_from_parts(parts)
            cur = out.setdefault(scope_key, _empty_scope_total())
            _accumulate_scope_total(cur, rec)
        for scope_key, cur in out.items():
            out[scope_key] = _finalize_scope_total(cur)
        return out

    totals_by_label = _load_records("totals_by_label.csv")
    totals_by_car_label = _load_records("totals_by_car_label.csv")
    totals_by_carweek_label = _load_records("totals_by_carweek_label.csv")
    totals_by_carweek_car_label = _load_records("totals_by_carweek_car_label.csv")
    totals_by_car = _load_records("totals_by_car.csv")
    totals_by_carweek = _load_records("totals_by_carweek.csv")
    totals_by_carweek_car = _load_records("totals_by_carweek_car.csv")
    totals_overall = {}
    try:
        p_overall = exposure_dir / "totals_overall.json"
        if p_overall.exists():
            totals_overall = json.loads(p_overall.read_text(encoding="utf-8"))
    except Exception:
        totals_overall = {}

    totals_overall = _resolve_overall_metric_summary(
        totals_overall,
        [
            ("totals_by_carweek.csv", totals_by_carweek),
            ("totals_by_car.csv", totals_by_car),
            ("totals_by_carweek_car.csv", totals_by_carweek_car),
        ],
    )

    route_metrics = {
        "labels": {
            "overall": _aggregate_metric_rows(totals_by_label, [], "label_name", lambda label, _rec: label),
            "by_car": _aggregate_metric_rows(totals_by_car_label, ["car_id"], "label_name", lambda label, _rec: label),
            "by_carweek": _aggregate_metric_rows(totals_by_carweek_label, ["car_week"], "label_name", lambda label, _rec: label),
            "by_carweek_car": _aggregate_metric_rows(totals_by_carweek_car_label, ["car_week", "car_id"], "label_name", lambda label, _rec: label),
        },
        "roadgroups": {
            "overall": _aggregate_metric_rows(totals_by_label, [], "label_name", _map_label_to_group),
            "by_car": _aggregate_metric_rows(totals_by_car_label, ["car_id"], "label_name", _map_label_to_group),
            "by_carweek": _aggregate_metric_rows(totals_by_carweek_label, ["car_week"], "label_name", _map_label_to_group),
            "by_carweek_car": _aggregate_metric_rows(totals_by_carweek_car_label, ["car_week", "car_id"], "label_name", _map_label_to_group),
        },
    }

    overall_totals = _empty_scope_total()
    _accumulate_scope_total(overall_totals, totals_overall)
    scope_totals = {
        "overall": _finalize_scope_total(overall_totals),
        "by_car": _build_scope_totals(totals_by_car, ["car_id"]),
        "by_carweek": _build_scope_totals(totals_by_carweek, ["car_week"]),
        "by_carweek_car": _build_scope_totals(totals_by_carweek_car, ["car_week", "car_id"]),
    }
    return route_metrics, scope_totals

# --------------------------- HTML template (rich UI baseline) ---------------------------

HTML_TEMPLATE = r"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Near-miss Hotspots (v67, JP + legacy1018)</title>

  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
  <!-- markercluster removed: use L.layerGroup + L.divIcon only (avoids canvas/cluster rendering failures on some envs) -->

  <style>
    html, body { height: 100%; margin: 0; font-family: system-ui, -apple-system, "Segoe UI", Roboto, "Noto Sans JP", "Hiragino Kaku Gothic ProN", "Meiryo", sans-serif; }
    #map { height: 100%; width: 100%; background:#ffffff; }
    .leaflet-container { background:#ffffff; }
    /* Tile seam fix: GPU compositing + overlap trick prevents sub-pixel white lines between tiles */
    .leaflet-tile-pane { transform: translate3d(0,0,0); will-change: transform; }
    .leaflet-tile {
      outline: 1px solid transparent;
      -webkit-backface-visibility: hidden;
      backface-visibility: hidden;
      transform: translate3d(0,0,0);
    }

    .topbar {
      position: fixed; top: 16px; left: 16px; z-index: 2400;
      display: flex; gap: 10px; align-items: center;
    }
    .chip {
      background: rgba(255,255,255,0.92);
      border: 1px solid rgba(0,0,0,0.08);
      border-radius: 999px;
      padding: 10px 14px;
      box-shadow: 0 10px 30px rgba(0,0,0,0.10);
      cursor: pointer; user-select: none;
      font-weight: 800;
      display: flex; align-items: center; gap: 10px;
    }
    .chip .dot { width: 9px; height: 9px; border-radius: 99px; background:#60a5fa; }
    .chip.active .dot { background:#22c55e; }
    .chip:active { transform: translateY(1px); }

    /* Right-side panels: bounded (no out-of-bound) */
    .panel{
      position: fixed;
      top: 72px; right: 16px; bottom: 16px;
      width: 380px; max-width: calc(100vw - 32px);
      background: rgba(255,255,255,0.94);
      border: 1px solid rgba(0,0,0,0.08);
      border-radius: 18px;
      box-shadow: 0 18px 45px rgba(0,0,0,0.18);
      overflow: hidden;
      display: none;
      z-index: 1800;
      flex-direction: column;
    }
    .panel.active{ display:flex; }
    .panel header{
      padding: 14px 16px 10px 16px;
      display:flex; align-items:flex-start; justify-content:space-between; gap:12px;
      border-bottom: 1px solid rgba(0,0,0,0.06);
      flex: 0 0 auto;
    }
    .panel{ font-size: 13px; }
    .panel header h2 { margin:0; font-size: 16px; line-height:1.2; }
    .panel header .sub { margin-top:4px; color:#6b7280; font-size: 11px; }
    .closebtn{
      border: none; background: rgba(0,0,0,0.04);
      width: 34px; height: 34px; border-radius: 12px;
      cursor:pointer; font-size: 20px; line-height: 34px;
    }
    .panel .body{
      padding: 12px 14px 14px 14px;
      overflow: auto;
      flex: 1 1 auto;
      min-height: 0;
    }

    .card{
      background: rgba(255,255,255,0.9);
      border: 1px solid rgba(0,0,0,0.06);
      border-radius: 12px;
      padding: 12px 12px;
      margin-bottom: 12px;
    }
    .row { display:flex; gap:10px; }
    .row > * { flex: 1; }
    label { display:block; font-size: 11.5px; color:#374151; margin-bottom: 6px; font-weight: 700; }
    select, input[type="text"], input[type="number"]{
      width: 100%;
      padding: 10px 10px;
      border-radius: 12px;
      border: 1px solid rgba(0,0,0,0.12);
      background: rgba(255,255,255,0.95);
      outline: none;
      box-sizing: border-box;
    }
    input[type="range"]{ width:100%; }
    .hint{ color:#6b7280; font-size: 12px; line-height: 1.35; }
    .btn{
      display:inline-flex; align-items:center; justify-content:center;
      padding: 10px 12px; border-radius: 12px;
      border: 1px solid rgba(0,0,0,0.10);
      background: rgba(0,0,0,0.04);
      cursor:pointer; font-weight: 800;
    }
    
    .btn.small{ padding: 6px 10px; border-radius: 10px; font-size: 12px; font-weight: 800; }
.btn.primary{ background:#111827; color:white; border-color:#111827; }
    .btn:active{ transform: translateY(1px); }

    .dx-list{
      height: 220px; overflow:auto;
      border-radius: 12px; border: 1px solid rgba(0,0,0,0.10);
      background: rgba(255,255,255,0.95);
      padding: 10px 10px;
    }
    .dx-item{
      display:flex; align-items:center; gap:10px;
      padding: 8px 6px; border-radius: 10px;
    }
    .dx-item:hover{ background: rgba(0,0,0,0.03); }
    .dx-item input{ width: 18px; height: 18px; }

    .legend{
      position:absolute; left:16px; bottom:16px; z-index: 1100;
      width: 360px; max-width: calc(100vw - 32px);
      background: rgba(255,255,255,0.92);
      border: 1px solid rgba(0,0,0,0.08);
      border-radius: 18px;
      box-shadow: 0 16px 40px rgba(0,0,0,0.18);
      padding: 12px 14px;
    }
    .legend h3 { margin:0 0 8px 0; font-size: 14px; }
    .bar{
      height: 14px; border-radius: 999px; border: 1px solid rgba(0,0,0,0.12);
      background: linear-gradient(90deg, #22c55e, #facc15, #fb923c, #ef4444);
    }
    .bar.blue{
      background: linear-gradient(90deg, #93c5fd, #a7f3d0, #fef08a, #fb7185);
    }
    .legend .small{ font-size: 12px; color:#6b7280; margin-top: 8px; }

    .pill{
      display:inline-flex; align-items:center; gap:6px;
      border:1px solid rgba(0,0,0,0.12);
      border-radius: 999px;
      padding: 6px 10px;
      font-size: 12px; color:#111827;
      background: rgba(255,255,255,0.9);
      cursor:pointer; user-select:none;
    }
    .pill.active{ background:#111827; color:white; border-color:#111827; }

    .leaflet-heatmap-layer, canvas.leaflet-heatmap-layer { pointer-events: none !important; }

    .popup-table { width: 100%; border-collapse: collapse; }
    .popup-table td { padding: 2px 4px; vertical-align: top; font-size: 12px; }
    .popup-table td.k { color:#6b7280; width: 88px; white-space: nowrap; }
    .popup-title { font-weight: 900; margin: 0 0 6px 0; }
    .boxlink { font-weight:900; }

    
    .rules{ margin: 8px 0 12px 0; }
    .rules summary{ cursor:pointer; font-weight:800; font-size:12.5px; color:#111827; }
    .rulesBody{ padding: 8px 2px 0 2px; color:#374151; font-size:12px; }
    .rulesBody ul{ margin: 8px 0 0 18px; padding:0; }
    .rulesBody li{ margin: 6px 0; }
/* Analysis list rows */
    .barRow{
      display:grid;
      grid-template-columns: 1.6fr 1fr 80px;
      align-items:center;
      gap:10px;
      padding: 6px 8px;
      border: 1px solid rgba(0,0,0,0.06);
      border-radius: 12px;
      margin-bottom: 8px;
      cursor: pointer;
      background: rgba(255,255,255,0.92);
    }
    .barRow:hover{ background: rgba(0,0,0,0.03); }
    .barName{ font-weight:800; font-size:13px; line-height:1.15; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; overflow:hidden; word-break:break-word; }
    .barSub{ display:-webkit-box; -webkit-line-clamp:1; -webkit-box-orient:vertical; overflow:hidden; }
    .miniBar{
      height: 8px; border-radius: 999px;
      background: rgba(0,0,0,0.06);
      overflow:hidden;
    }
    .miniBar > div{
      height: 100%;
      background: #22c55e;
      width: 0%;
    }
    .muted{ color:#6b7280; font-size: 12px; }
    .btn:disabled{
      opacity: 0.45;
      cursor: not-allowed;
      filter: saturate(0.55);
    }
    .leaflet-marker-icon.nearmiss-point-icon,
    .leaflet-marker-icon.nearmiss-agg-icon,
    .leaflet-marker-icon.nearmiss-debug-icon{
      background: transparent;
      border: none;
    }
    .nm-point-dot{
      width: 10px;
      height: 10px;
      border-radius: 999px;
      border: 1.5px solid rgba(255,255,255,0.96);
      box-shadow: 0 0 0 1px rgba(17,24,39,0.22), 0 4px 10px rgba(17,24,39,0.10);
      box-sizing: border-box;
      opacity: 0.94;
    }
    .nm-agg-dot{
      display:flex;
      align-items:center;
      justify-content:center;
      border-radius:999px;
      border: 1.5px solid rgba(255,255,255,0.98);
      box-shadow: 0 0 0 1px rgba(17,24,39,0.18), 0 8px 18px rgba(17,24,39,0.12);
      color:#111827;
      font-weight:900;
      line-height:1;
      box-sizing:border-box;
      background:#d1d5db;
      padding: 0 3px;
      white-space:nowrap;
      opacity: 0.93;
    }
    .nm-debug-dot{
      display:flex;
      align-items:center;
      justify-content:center;
      width: 18px;
      height: 18px;
      border-radius:999px;
      background:#2563eb;
      color:#fff;
      font-size:11px;
      font-weight:900;
      border: 2px solid rgba(255,255,255,0.96);
      box-shadow: 0 0 0 1px rgba(17,24,39,0.25), 0 10px 22px rgba(37,99,235,0.25);
      box-sizing:border-box;
    }

  </style>
</head>
<body>
<div id="map"></div>

<div class="topbar">
  <div class="chip" id="chipFilter"><span class="dot"></span>フィルタ</div>
  <div class="chip" id="chipSummary"><span class="dot"></span>Summary</div>
  <div class="chip" id="chipAnalysis"><span class="dot"></span>Analysis</div>
  <div class="chip" id="chipRoutes"><span class="dot"></span>Routes</div>
</div>

<!-- Filter panel -->
<div id="panelFilter" class="panel">
  <header>
    <div>
      <h2>Near-miss Hotspots</h2>
      <div class="sub" id="metaLine">Excelから再現</div>
    </div>
    <button class="closebtn" id="closeFilter" title="close">×</button>
  </header>
  <div class="body">
    <div class="card">
      <div class="hint" id="metaHint"></div>
    </div>

    <div class="card">
      <div class="row">
        <div>
          <label>道路（集約）</label>
          <select id="roadSelect"></select>
        </div>
        <div>
          <label>ロケーション</label>
          <select id="locSelect"></select>
        </div>
      </div>
      <div class="row" style="margin-top:10px;">
        <div>
          <label>Score 最小</label>
          <input id="scoreMin" type="number" value="0" min="0" max="45" step="1"/>
        </div>
        <div>
          <label>Score 最大</label>
          <input id="scoreMax" type="number" value="45" min="0" max="45" step="1"/>
        </div>
      </div>

      <details style="margin-top:10px;">
        <summary style="cursor:pointer; font-weight:800;">Speed filter（km/h）</summary>
        <div class="hint" style="margin-top:6px;">※ ポップアップの 最高/最低/平均 の数値をクリックすると、ここに範囲が自動入力される（±5km/h）。ヒストグラムのバークリックも同様。</div>

        <div class="row" style="margin-top:10px;">
          <div>
            <label>平均速度 min</label>
            <input id="speedAvgMin" type="number" placeholder="-" step="0.1"/>
          </div>
          <div>
            <label>平均速度 max</label>
            <input id="speedAvgMax" type="number" placeholder="-" step="0.1"/>
          </div>
        </div>

        <div class="row" style="margin-top:10px;">
          <div>
            <label>最高速度 min</label>
            <input id="speedMaxMin" type="number" placeholder="-" step="0.1"/>
          </div>
          <div>
            <label>最高速度 max</label>
            <input id="speedMaxMax" type="number" placeholder="-" step="0.1"/>
          </div>
        </div>

        <div class="row" style="margin-top:10px;">
          <div>
            <label>最低速度 min</label>
            <input id="speedMinMin" type="number" placeholder="-" step="0.1"/>
          </div>
          <div>
            <label>最低速度 max</label>
            <input id="speedMinMax" type="number" placeholder="-" step="0.1"/>
          </div>
        </div>

        <div class="row" style="margin-top:10px;">
          <div style="flex:1;">
            <button class="btn" id="clearSpeedFilters">Clear speed filters</button>
          </div>
        </div>
      </details>

      <div style="margin-top:10px;">
        <label>DX（複数選択）</label>
        <input id="dxSearch" type="text" placeholder="search dx..." style="margin-bottom:8px;"/>
        <div style="display:flex; gap:10px; margin-bottom:8px;">
          <button class="btn" id="dxAll">Select all</button>
          <button class="btn" id="dxNone">Select none</button>
        </div>
        <div class="dx-list" id="dxList"></div>
        <div class="hint" style="margin-top:8px;">
          DXクリックで単一DXに切替（Analysisパネルからも同じ操作が可能）。
        </div>
      </div>

      <div style="margin-top:12px;">
        <label>Layers</label>
        <div style="display:flex; gap:10px; flex-wrap:wrap;">
          <span class="pill active" id="togglePoints">Points</span>
          <span class="pill active" id="toggleHeatAvg">Heat AVG</span>
          <span class="pill" id="toggleHeatFreq">Heat FREQ</span>
        </div>
      </div>

      <div style="margin-top:12px;">
        <label>Heat opacity (<span id="opVal">35</span>%)</label>
        <input id="heatOpacity" type="range" min="5" max="85" value="35"/>
      </div>

      <div style="margin-top:12px;">
        <label>Grouping (auto by zoom)</label>
        <div class="hint" style="margin-top:6px;">
          <span id="groupingMetersText">-</span>
        </div>
        <div class="hint" style="margin-top:6px;">
          zoom ≥ 15: individual points<br/>
          zoom 14: 100m · 13: 200m · 12: 500m · 11: 1000m · ≤10: 2000m
        </div>
      </div>

      <div style="margin-top:12px; display:flex; gap:10px;">
        <button class="btn primary" id="applyNow">Apply</button>
        <button class="btn" id="resetAll">Reset</button>
      </div>
    </div>

    <div class="card">
      <div class="hint" id="statusText">表示中: -</div>
    </div>
  </div>
</div>

<!-- Summary panel (kept simple) -->
<div id="panelSummary" class="panel">
  <header>
    <div>
      <h2>Summary / Insights</h2>
      <div class="sub">主要指標: avg_score / n / 上位道路・ロケーション</div>
    </div>
    <button class="closebtn" id="closeSummary" title="close">×</button>
  </header>
  <div class="body">
    <div class="card">
      <div class="hint">※ Summaryは「説明用の短い要点」。詳細はAnalysisで確認。</div>
    </div>
    <div class="card">
      <div style="font-weight:900; margin-bottom:8px;">Most dangerous roads（metric = avg × log(n+1)）</div>
      <div id="sumTopRoads"></div>
    </div>
    <div class="card">
      <div style="font-weight:900; margin-bottom:8px;">Most dangerous locations（metric = avg × log(n+1)）</div>
      <div id="sumTopLocs"></div>
    </div>
  </div>
</div>

<!-- Analysis panel (v59-style, updates instantly) -->
<div id="panelAnalysis" class="panel">
  <header>
    <div>
      <h2>Interactive Analysis</h2>
      <div class="sub">フィルタ結果に連動して、分布/ランキング/欠損を即時に更新</div>
    </div>
    <div class="hdrBtns"><button class="btn small" id="analysisReset" title="Reset filters">Reset</button><button class="closebtn" id="closeAnalysis" title="close">×</button></div>
  </header>
  <div class="body">
    <div class="card">
      <div class="hint" id="analysisMeta"></div>
      <div style="display:flex; gap:12px; align-items:center; flex-wrap:wrap; margin-top:10px;">
        <label style="display:flex; gap:8px; align-items:center; margin:0; font-weight:900;">
          <span style="font-size:12px; color:#374151;">表示上位</span>
          <input id="analysisTopN" type="number" min="5" max="50" value="14"
                 style="width:72px; padding:8px 10px; border-radius:12px; border:1px solid rgba(0,0,0,0.12);" />
        </label>
        <label style="display:flex; gap:8px; align-items:center; margin:0; font-weight:900;">
          <input id="hideUnknownRoads" type="checkbox" checked />
          <span style="font-size:12px; color:#374151;">道路「Unknown」を除外（ランキングのみ）</span>
        </label>
      </div>
      <div class="hint" style="margin-top:8px;">
        ※ 緯度経度があっても、OSM側に道路名（name/ref）が無い区間（特にランプ/リンク）では「道路名なし」として集計する。
      </div>
    </div>

    <div class="card">
      <div style="font-weight:900; margin-bottom:10px;">DX breakdown（クリックでDX単一フィルタ）</div>
      <div id="dxBreakdown"></div>
      <div class="hint">DX breakdown for the current filter. Click a DX to isolate it on the map and refresh the rest of the analysis instantly.</div>
    </div>

    <div class="card">
      <div style="font-weight:900; margin-bottom:10px;">Top road within each DX</div>
      <div id="dxRoadLead"></div>
      <div class="hint">Mapped incidents only. Bars compare the top road share inside each DX bucket, so large DX classes do not automatically dominate.</div>
    </div>

    <div class="card">
      <div style="font-weight:900; margin-bottom:10px;">Over-indexed DX per road</div>
      <div id="dxRoadLift"></div>
      <div class="hint">Shows where a DX appears more often than its overall baseline. Rankings use smoothed lift plus minimum-support thresholds to reduce small-sample bias.</div>
    </div>

    <div class="card">
      <div style="font-weight:900; margin-bottom:10px;">Dominant DX by avg-speed band</div>
      <div id="speedBandDxLead"></div>
      <div class="hint">DXクリック → フィルタのDXをその1つに切替し、地図も更新。</div>
    </div>

    <div class="card">
      <div style="font-weight:900; margin-bottom:10px;">Top roads（by count）</div>
      <div id="roadCount"></div>
      <div class="hint">道路クリック → roadフィルタに設定。集計対象は shown points のみ。</div>
    </div>

    <div class="card" id="roadRateCard" style="display:none;">
      <div style="font-weight:900; margin-bottom:10px;">Top roads（per 1km）</div>
      <div id="roadRate"></div>
      <div class="hint">露出距離(exposure)がある道路だけ表示。集計対象は shown points のみ。</div>
    </div>

    <div class="card">
      <div style="font-weight:900; margin-bottom:10px;">Speed histogram（avg speed）</div>
      <div style="display:flex; gap:10px; flex-wrap:wrap;">
        <div style="flex:1; min-width:220px;">
          <div class="hint" style="margin-bottom:6px;">平均速度</div>
          <div id="speedHistAvg"></div>
        </div>
        <div style="flex:1; min-width:220px;">
          <div class="hint" style="margin-bottom:6px;">最高速度</div>
          <div id="speedHistMax"></div>
        </div>
        <div style="flex:1; min-width:220px;">
          <div class="hint" style="margin-bottom:6px;">最低速度</div>
          <div id="speedHistMin"></div>
        </div>
      </div>
      <div class="hint">速度は「平均速度」列。欠損は除外。</div>
    </div>

    <div class="card">
      <div style="font-weight:900; margin-bottom:10px;">Missing lat/lon（地図に出ない）</div>
      <div id="missingLatLon"></div>
    </div>

    <div class="card" id="cellDetailsCard" style="display:none;">
      <div style="display:flex; justify-content:space-between; align-items:center; gap:10px;">
        <div style="font-weight:900;">Grouped cell details</div>
        <button class="btn small" onclick="clearCellDetails(); return false;">Clear</button>
      </div>
      <div id="cellDetailsMeta" class="hint" style="margin-top:6px;"></div>
      <div id="cellDetailsList" style="margin-top:10px; max-height: 280px; overflow:auto; border-top:1px solid rgba(0,0,0,0.08); padding-top:10px;"></div>
      <div class="hint" style="margin-top:8px;">※ グループ化表示（zoom≤14）でも、セル内の事象一覧をここに表示できます。</div>
    </div>

  </div>
</div>

<!-- Routes panel (v67) -->
<div id="panelRoutes" class="panel">
  <header>
    <div>
      <h2>Routes</h2>
      <div class="sub">car_week route exposure and path preview (routes_by_*.geojson)</div>
    </div>
    <button class="closebtn" id="closeRoutes" title="close">×</button>
  </header>
  <div class="body">
    <div class="card">
      <div class="row">
        <div>
          <label>car_week</label>
          <select id="carweekRouteSelect"></select>
        </div>
      </div>
      <div class="card" style="margin-top:10px;">
        <div style="font-weight:900; margin-bottom:6px;">Distance ran (km)</div>
        <div id="kmTotals" class="hint">-</div>
      </div>
<div style="margin-top:10px;">
        <label>route_group</label>
        <select id="routeGroupSelect"></select>
        <div id="routeGroupKm" class="hint">-</div>
      </div>
</div>
      <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap;">
        <button class="btn primary" id="btnShowCarweek">Show car_week route</button>
        <button class="btn" id="btnClearRoutes">Clear</button>
      </div>
      <div class="card" style="margin-top:12px;">
        <div style="font-weight:900; margin-bottom:6px;">Selected route summary</div>
        <div id="routeScopeSummary" class="hint">Select car_week to inspect exposure by route segment.</div>
      </div>
      <div class="card" style="margin-top:12px;">
        <div style="font-weight:900; margin-bottom:6px;">Top road segments</div>
        <div id="routeSegmentList" class="hint">-</div>
      </div>
      <div class="card" style="margin-top:12px;">
        <div style="font-weight:900; margin-bottom:6px;">Top road groups</div>
        <div id="routeRoadGroupList" class="hint">-</div>
      </div>
      <div class="card" style="margin-top:12px;">
        <div style="font-weight:900; margin-bottom:6px;">Near-miss rate by road segment (cases / observed km)</div>
        <div id="routeSegmentRateList" class="hint" style="max-height:360px;overflow-y:auto;">-</div>
      </div>
      <div class="card" style="margin-top:12px;">
        <div style="font-weight:900; margin-bottom:6px;">Near-miss rate by road group (cases / observed km)</div>
        <div id="routeRoadGroupRateList" class="hint" style="max-height:360px;overflow-y:auto;">-</div>
      </div>
      <div class="hint" style="margin-top:10px;"><div id="routesGroupingText" class="hint" style="margin-bottom:6px;"></div>
        ※ Rate = cases ÷ observed km (GPS-measured distance). Expected km shown in muted text where available (scheduled distance). Small denominators inflate rates — check the obs km value alongside the rate.<br/>
        ※ ルートは downsample 済み（--routes-max-points）。ポイントをクリックすると「Show route snippet」で近傍区間だけハイライトできる。
      </div>
    </div>

    <div class="card">
      <details class="rules">
        <summary>Help: Grouping + zoom rules</summary>
        <div class="hint" style="margin-top:8px;">
          <ul style="margin:8px 0 0 18px; padding:0;">
            <li><b>Grouping is REAL computation:</b> At zoom ≤14, points are aggregated into a <b>square grid in meters</b> (EPSG:3857 Web Mercator). Each visible marker represents <b>all incidents inside one grid cell</b> (count + avg score).</li>
            <li><b>Grouping condition (exact):</b> Convert each point (lat,lon) → (x,y) meters, then compute <code>gx=floor(x/cellMeters)</code>, <code>gy=floor(y/cellMeters)</code>. Same (gx,gy) ⇒ same group/cell.</li>
            <li><b>Cell size by zoom (meters):</b>
              <div style="margin-top:6px; font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, 'Liberation Mono', 'Courier New', monospace; font-size: 12px; line-height: 1.5;">
                z>=15: individual points (no grouping)<br/>
                z14: 120m &nbsp;|&nbsp; z13: 260m &nbsp;|&nbsp; z12: 700m &nbsp;|&nbsp; z11: 1600m &nbsp;|&nbsp; z10: 3500m<br/>
                z9: 8000m &nbsp;|&nbsp; z8: 15000m &nbsp;|&nbsp; z<=7: 30000m
              </div>
              <div style="margin-top:6px;">Zoom change triggers recompute automatically (<b>zoomend → apply</b>), so grouping changes with zoom.</div>
            </li>
            <li><b>Heat recompute uses the same cells:</b> after filters, AVG-Heat uses <b>avg score per cell</b>, FREQ-Heat uses <b>count per cell</b>. (So heat is also “grouped”, not cosmetic.)</li>
            <li><b>Routes are independent:</b> Show/Clear does not affect heat/points. Clear removes route + snippet.</li>
            <li><b>Show route snippet (popup):</b> draws the selected incident's <b>scene route</b> when available, otherwise falls back to the related <b>car_week route</b>. Dedicated car-level routes are hidden here because they are currently too noisy and unreliable.</li>
</ul>
        </div>
      </details>
    </div>

  </div>
</div>


<div class="legend" id="legendBox">
  <div style="display:flex; align-items:center; justify-content:space-between;">
    <h3>Legend</h3>
    <span class="pill active" id="legendMode">AVG</span>
  </div>

  <div style="margin-top:8px;">
    <div style="font-size:12px; color:#6b7280; margin-bottom:6px;">Points (score)</div>
    <div class="bar"></div>
  </div>

  <div style="margin-top:10px;">
    <div style="font-size:12px; color:#6b7280; margin-bottom:6px;">Heat AVG (grid avg score)</div>
    <div class="bar"></div>
  </div>

  <div style="margin-top:10px;">
    <div style="font-size:12px; color:#6b7280; margin-bottom:6px;">Heat FREQ (grid density)</div>
    <div class="bar blue"></div>
  </div>

  <div class="small" id="quantText">P50=? / P75=? / P90=?</div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<!-- markercluster script removed: use L.layerGroup only -->
<script src="https://unpkg.com/leaflet.heat@0.2.0/dist/leaflet-heat.js"></script>

<script>
const DATA = __DATA_JSON__;

const ROUTES_CARWEEK = __ROUTES_CARWEEK_JSON__;
const ROUTES_CAR = __ROUTES_CAR_JSON__;
const ROUTES_SCENE = __ROUTES_SCENE_JSON__;

const EXPOSURE = DATA.exposure || {};
const ROAD_GROUP_LOOKUP = DATA.road_group_lookup || {};
const TOTALS_OVERALL = DATA.totals_overall || null;
const TOTALS_BY_CAR = DATA.totals_by_car || {};
const TOTALS_BY_CARWEEK = DATA.totals_by_carweek || {};
const TOTALS_BY_CARWEEK_CAR = DATA.totals_by_carweek_car || {};
const ROUTE_METRICS = DATA.route_metrics || {labels:{}, roadgroups:{}};
const ROUTE_SCOPE_TOTALS = DATA.route_scope_totals || {overall:{}, by_car:{}, by_carweek:{}, by_carweek_car:{}};
const DX_OPTIONS = Array.isArray(DATA.dx_options) ? DATA.dx_options : [];
const DX_LABEL = new Map(DX_OPTIONS.map(o => [o["分類"], o.label]));
const EARTH_R = 6378137.0;
const DEBUG_HASH = String(location.hash || "");

function numOrNull(v){
  const n = Number(v);
  return Number.isFinite(n) ? n : null;
}

function isValidLatLon(lat, lon){
  if (!Number.isFinite(lat) || !Number.isFinite(lon)) return false;
  if (Math.abs(lat) > 90 || Math.abs(lon) > 180) return false;
  if (lat === 0 && lon === 0) return false;
  return true;
}

function dxLabelFor(row){
  const code = String(row?.["分類"] || row?.dx_code || "Unknown").trim() || "Unknown";
  const label = String(row?.["分類名"] || row?.dx_name || DX_LABEL.get(code) || "").trim();
  if (label) return label;
  return code === "Unknown" ? "不明" : code;
}

function normalizeRowForMap(row, index){
  const out = row || {};
  out.__i = Number.isFinite(Number(out.__i)) ? Number(out.__i) : index;
  out.lat = numOrNull(out.lat ?? out.latitude ?? out.LATITUDE ?? out.center_latitude ?? out.LATITUDEFRONT ?? out["緯度"]);
  out.lon = numOrNull(out.lon ?? out.lng ?? out.longitude ?? out.LONGITUDE ?? out.center_longitude ?? out.LONGITUDEFRONT ?? out["経度"]);
  out.has_latlon = isValidLatLon(out.lat, out.lon);
  if (!out.has_latlon){
    out.lat = null;
    out.lon = null;
  }

  out.score = numOrNull(out.score);
  if (!Number.isFinite(out.score)){
    const lvl = String(out["ニアミスレベル"] || out.risk_level || "").trim();
    out.score = (lvl === "大") ? 40 : (lvl === "中") ? 30 : (lvl === "小") ? 20 : 0;
  }

  out["平均速度"] = numOrNull(out["平均速度"]);
  out["最低速度"] = numOrNull(out["最低速度"]);
  out["最高速度"] = numOrNull(out["最高速度"]);
  out["snap_dist_m"] = numOrNull(out["snap_dist_m"]);

  out["分類"] = String(out["分類"] ?? out.dx_code ?? "Unknown").trim() || "Unknown";
  out["分類名"] = dxLabelFor(out);
  out["分類_1"] = String(out["分類_1"] ?? out["分類.1"] ?? "").trim();
  out["道路"] = String(out["道路"] || out.road || "不明").trim() || "不明";
  out["ロケーション"] = String(out["ロケーション"] || out.location || "不明").trim() || "不明";
  out.dx_code = out["分類"];
  out.dx_name = out["分類名"];
  out.road = out["道路"];
  out.location = out["ロケーション"];
  out.speed_avg = out["平均速度"];
  out.speed_min = out["最低速度"];
  out.speed_max = out["最高速度"];
  return out;
}

function exposureKmValue(ex){
  const expected = Number(ex?.expected_distance_km);
  if (Number.isFinite(expected) && expected > 0) return expected;
  const dist = Number(ex?.distance_km);
  if (Number.isFinite(dist) && dist > 0) return dist;
  const km = Number(ex?.km);
  return (Number.isFinite(km) && km > 0) ? km : NaN;
}

const ALL_ROWS = (Array.isArray(DATA.rows) ? DATA.rows : []).map((row, index) => normalizeRowForMap(row, index));

let map;
let pointsLayer;
let aggLayer;
let heatAvgLayer;
let heatFreqLayer;
let currentCellFilter = null;
let activeRouteFilter = null;
let lastWindowError = "";
let lastStatusPayload = null;

function clamp(v, lo, hi){ return Math.max(lo, Math.min(hi, v)); }

function scoreToColor(score){
  if (score >= 45) return "#ef4444";
  if (score >= 40) return "#fb923c";
  if (score >= 30) return "#facc15";
  if (score >= 20) return "#22c55e";
  return "#9ca3af";
}

function fmt(v, digits=2){
  if (v === null || v === undefined || Number.isNaN(v)) return "-";
  return Number(v).toFixed(digits);
}

function layerCount(layer){
  try{
    return (layer && typeof layer.getLayers === "function") ? layer.getLayers().length : 0;
  }catch(_err){
    return 0;
  }
}

function renderStatusText(payload){
  if (payload) lastStatusPayload = payload;
  const el = document.getElementById("statusText");
  if (!el) return;
  const p = lastStatusPayload || {};
  const baseText = p.baseText || "表示中: -";
  const shownPointsLen = Number.isFinite(p.shownPointsLen) ? p.shownPointsLen : 0;
  const pointsLayerCount = Number.isFinite(p.pointsLayerCount) ? p.pointsLayerCount : layerCount(pointsLayer);
  const aggLayerCount = Number.isFinite(p.aggLayerCount) ? p.aggLayerCount : layerCount(aggLayer);
  const errText = lastWindowError || "none";
  el.textContent = `${baseText} | shownPoints.length=${shownPointsLen} | pointsLayer.getLayers().length=${pointsLayerCount} | aggLayer.getLayers().length=${aggLayerCount} | window.onerror=${errText}`;
}

window.onerror = function(message, source, lineno, colno, error){
  const core = error && error.message ? error.message : String(message || "unknown error");
  lastWindowError = `${core}${lineno ? ` @${lineno}:${colno || 0}` : ""}`;
  renderStatusText(lastStatusPayload);
  return false;
};

window.addEventListener("unhandledrejection", (event) => {
  const reason = event?.reason;
  lastWindowError = reason && reason.message ? reason.message : String(reason || "unhandled rejection");
  renderStatusText(lastStatusPayload);
});

function createMarkerGroup(_kind){
  // Always use plain LayerGroup + L.divIcon markers.
  // MarkerCluster is not used: it relies on canvas rendering which can silently
  // produce "no nodes" in some browser/environment combinations.
  return L.layerGroup();
}

function createSafeHeatLayer(options){
  if (typeof L.heatLayer === "function"){
    return L.heatLayer([], options);
  }
  const noop = L.layerGroup();
  noop.setLatLngs = function(){ return noop; };
  noop.setOptions = function(){ return noop; };
  noop.redraw = function(){ return noop; };
  return noop;
}

function makePointMarker(p){
  return L.marker([p.lat, p.lon], {
    pane: "pointsPane",
    icon: L.divIcon({
      className: "nearmiss-point-icon",
      html: `<div class="nm-point-dot" style="background:${scoreToColor(p.score)};"></div>`,
      iconSize: [10, 10],
      iconAnchor: [5, 5],
      popupAnchor: [0, -8],
    }),
  });
}

function makeAggMarker(c, maxCount){
  const ratio = (maxCount > 0) ? (c.count / maxCount) : 0;
  const size = Math.round(18 + 20 * Math.sqrt(ratio));
  const fontSize = Math.max(10, Math.min(14, Math.round(size * 0.38)));
  return L.marker([c.center[0], c.center[1]], {
    pane: "aggPane",
    icon: L.divIcon({
      className: "nearmiss-agg-icon",
      html: `<div class="nm-agg-dot" style="width:${size}px;height:${size}px;background:${scoreToColor(c.avgScore)};font-size:${fontSize}px;"><span>${c.count}</span></div>`,
      iconSize: [size, size],
      iconAnchor: [size / 2, size / 2],
      popupAnchor: [0, -Math.round(size / 2)],
    }),
  });
}

function buildPopup(p){
  const box = p["BOXリンク"] ? `<a class="boxlink" href="${p["BOXリンク"]}" target="_blank" rel="noopener">BOX</a>` : "";
  const title = `No: ${p.No ?? "-"}　${p["分類名"] || p["分類"]}　${p["分類_1"]}　score:${p.score}`;
  return `
    <div style="min-width: 260px; max-width: 320px;"><div class="popup-body">
      <div class="popup-title">${title} ${box}</div>
      <table class="popup-table">
        <tr><td class="k">ニアミス</td><td>${p["ニアミスレベル"] || "-"}</td></tr>
        <tr><td class="k">道路</td><td>${p["道路"]}</td></tr>
        <tr><td class="k">ロケ</td><td>${p["ロケーション"]}</td></tr>
        <tr><td class="k">最高</td><td><a href="#" onclick="speedClick('max', ${Number(p['最高速度'])}); return false;">${fmt(p["最高速度"],1)} km/h</a></td></tr>
        <tr><td class="k">最低</td><td><a href="#" onclick="speedClick('min', ${Number(p['最低速度'])}); return false;">${fmt(p["最低速度"],1)} km/h</a></td></tr>
        <tr><td class="k">平均</td><td><a href="#" onclick="speedClick('avg', ${Number(p['平均速度'])}); return false;">${fmt(p["平均速度"],1)} km/h</a></td></tr>
        <tr><td class="k">car_week</td><td>${p["car_week"] || "-"}</td></tr>
        <tr><td class=\"k\">car_id</td><td>${p.car_id || "-"}</td></tr>
        <tr><td class=\"k\">route</td><td><button class=\"btn small\" onclick=\"showRouteSnippet(${p.__i}); return false;\">Show route snippet</button></td></tr>
        <tr><td class="k">格納先</td><td style="word-break:break-all;">${p["格納先"] || "-"}</td></tr>
        <tr><td class="k">snap(m)</td><td>${fmt(p["snap_dist_m"],3)}</td></tr>
        <tr><td class="k">lat/lon</td><td>${fmt(p.lat,6)}, ${fmt(p.lon,6)}</td></tr>
      </table>
      </div>
    </div>
  `;
}

function initMap(){
  map = L.map("map", { zoomControl: false, renderer: L.svg() });

  const base1 = L.tileLayer("https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png", { attribution: '&copy; OpenStreetMap &copy; CARTO' });
  const base2 = L.tileLayer("https://tile.openstreetmap.org/{z}/{x}/{y}.png", { attribution: '&copy; OpenStreetMap contributors' });
  base1.addTo(map);

  map.createPane("aggPane");
  map.getPane("aggPane").style.zIndex = "640";
  map.createPane("pointsPane");
  map.getPane("pointsPane").style.zIndex = "650";
  map.createPane("debugPane");
  map.getPane("debugPane").style.zIndex = "660";

  window.__debugMarker__ = null;
  if (DEBUG_HASH.includes("debug")){
    const _dm = L.marker([35.681236, 139.767125], {
      pane: "debugPane",
      icon: L.divIcon({
        className: "nearmiss-debug-icon",
        html: `<div class="nm-debug-dot">D</div>`,
        iconSize: [18, 18],
        iconAnchor: [9, 9],
        popupAnchor: [0, -10],
      }),
    }).addTo(map);
    _dm.bindPopup("DEBUG marker: Tokyo | shownPoints=? (will update on applyAll)").openPopup();
    window.__debugMarker__ = _dm;
  }

  // Put zoom bottom-right to avoid top-left chips.
  L.control.zoom({ position: "bottomright" }).addTo(map);
  L.control.layers({ "Carto (Light)": base1, "OSM": base2 }, null, { position: "topright" }).addTo(map);

  pointsLayer = createMarkerGroup("points");
  aggLayer = createMarkerGroup("agg");

  heatAvgLayer = createSafeHeatLayer({
    radius: 20, blur: 24, maxZoom: 17, minOpacity: 0.10,
    gradient: { 0.0:"#22c55e", 0.45:"#facc15", 0.7:"#fb923c", 1.0:"#ef4444" }
  });

  heatFreqLayer = createSafeHeatLayer({
    radius: 26, blur: 30, maxZoom: 17, minOpacity: 0.10,
    gradient: { 0.0:"#93c5fd", 0.35:"#a7f3d0", 0.65:"#fef08a", 1.0:"#fb7185" }
  });

  pointsLayer.addTo(map);
  heatAvgLayer.addTo(map);

  const bounds = L.latLngBounds([[__MIN_LAT__, __MIN_LON__], [__MAX_LAT__, __MAX_LON__]]);
  if (bounds.isValid()){
    map.fitBounds(bounds.pad(0.10));
  } else {
    map.setView([35.681236, 139.767125], 6);
  }
  // Recompute grouping/heat when zoom changes (grouping is meaningful by zoom)
  map.on("zoomend", ()=>{ applyAll(); });
  setTimeout(()=>{ try{ map.invalidateSize(); }catch(_err){} }, 0);
  renderStatusText({ baseText: "表示中: 初期化完了", shownPointsLen: 0, pointsLayerCount: 0, aggLayerCount: 0 });
}

function uniqueSorted(arr){ return Array.from(new Set(arr)).sort((a,b)=> a.localeCompare(b, "ja")); }

function buildSelectOptions(selectEl, values, includeAll=true, allLabel="全て"){
  const opts = [];
  if (includeAll) opts.push(`<option value="">${allLabel}</option>`);
  for (const v of values){
    opts.push(`<option value="${String(v).replaceAll('"','&quot;')}">${v}</option>`);
  }
  selectEl.innerHTML = opts.join("");
}

function normalizeDxOption(o){
  if (typeof o === "string") return {"分類": o, label: o};
  if (!o || typeof o !== "object") return {"分類": "Unknown", label: "不明"};
  const code = o["分類"] ?? o.code ?? o.dx ?? o.value ?? o.id ?? o.key;
  const label = o.label ?? o.name ?? o["名称"] ?? o["name"] ?? code;
  return {"分類": (code ?? "Unknown"), label: (label ?? "不明")};
}

function buildDxList(dxOptions){
  const dxList = document.getElementById("dxList");
  const norm = (dxOptions || []).map(normalizeDxOption).filter(o => !!o["分類"]);
  dxList.innerHTML = norm.map(o => {
    const code = o["分類"]; 
    const label = o.label;
    return `
      <div class="dx-item">
        <input type="checkbox" value="${code}" checked />
        <div style="min-width:0;">
          <div style="font-weight:900; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${code}</div>
          <div class="muted" style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${label}</div>
        </div>
      </div>
    `;
  }).join("");

  // clicking row text toggles checkbox
  dxList.querySelectorAll(".dx-item").forEach(item => {
    item.addEventListener("click", (ev) => {
      if (ev.target.tagName.toLowerCase() === "input") return;
      const cb = item.querySelector("input[type=checkbox]");
      cb.checked = !cb.checked;
    });
  });
}

function setOnlyDx(dxCode){
  document.querySelectorAll("#dxList input[type=checkbox]").forEach(cb => {
    cb.checked = (cb.value === dxCode);
  });
}


function readNumOrNullById(id){
  const el = document.getElementById(id);
  if (!el) return null;
  const t = String(el.value || "").trim();
  if (t === "") return null;
  const v = Number(t);
  return Number.isFinite(v) ? v : null;
}

function getFilters(){
  const road = document.getElementById("roadSelect").value || "";
  const loc = document.getElementById("locSelect").value || "";
  const smin = Number(document.getElementById("scoreMin").value || 0);
  const smax = Number(document.getElementById("scoreMax").value || 45);
  const dxSet = new Set();
  document.querySelectorAll("#dxList input[type=checkbox]").forEach(cb => { if (cb.checked) dxSet.add(cb.value); });
  const heatOpacity = Number(document.getElementById("heatOpacity").value || 35);

  const zoom = (typeof map !== "undefined" && map) ? map.getZoom() : 14;
  const groupMeters = groupingMetersForZoom(zoom);

  // UI: show current grouping
  const gt = document.getElementById("groupingMetersText");
  const rt = document.getElementById("routesGroupingText");
  const msg = (groupMeters <= 0)
    ? `zoom ${zoom}: individual points (no grouping)`
    : `zoom ${zoom}: ${groupMeters}m grid grouping`;
  if (gt) gt.textContent = msg;
  if (rt) rt.textContent = "Aggregation: " + msg;

  const avgMin = readNumOrNullById('speedAvgMin');
  const avgMax = readNumOrNullById('speedAvgMax');
  const maxMin = readNumOrNullById('speedMaxMin');
  const maxMax = readNumOrNullById('speedMaxMax');
  const minMin = readNumOrNullById('speedMinMin');
  const minMax = readNumOrNullById('speedMinMax');

  return { road, loc, smin, smax, dxSet, heatOpacity, zoom, groupMeters, avgMin, avgMax, maxMin, maxMax, minMin, minMax };
}

function setHeatOpacity(percent){
  const minOpacity = clamp((Number(percent) || 35) / 100, 0.05, 0.85);
  [heatAvgLayer, heatFreqLayer].forEach(layer => {
    if (!layer || typeof layer.setOptions !== "function") return;
    try{
      layer.setOptions({ minOpacity });
      if (typeof layer.redraw === "function") layer.redraw();
    }catch(_err){}
  });
}

function getDrawablePoints(rows){
  return (rows || []).filter(p => !!p?.has_latlon && isValidLatLon(Number(p.lat), Number(p.lon)));
}


function filterRows(all, f){
  const out = [];
  for (const p of all){
    if (activeRouteFilter?.car_week){
      const rowCarWeek = String(p?.car_week || "").trim();
      if (rowCarWeek !== activeRouteFilter.car_week) continue;
    }
    if (f.road && p["道路"] !== f.road) continue;
    if (f.loc && p["ロケーション"] !== f.loc) continue;
    if (!f.dxSet.has(p["分類"])) continue;
    if (p.score < f.smin || p.score > f.smax) continue;
    // speed filters (km/h). Missing values are excluded when filter is active.
    const vAvg = Number(p["平均速度"]);
    const vMax = Number(p["最高速度"]);
    const vMin = Number(p["最低速度"]);
    if (f.avgMin !== null && f.avgMin !== undefined){
      if (!Number.isFinite(vAvg) || vAvg < f.avgMin) continue;
    }
    if (f.avgMax !== null && f.avgMax !== undefined){
      if (!Number.isFinite(vAvg) || vAvg > f.avgMax) continue;
    }
    if (f.maxMin !== null && f.maxMin !== undefined){
      if (!Number.isFinite(vMax) || vMax < f.maxMin) continue;
    }
    if (f.maxMax !== null && f.maxMax !== undefined){
      if (!Number.isFinite(vMax) || vMax > f.maxMax) continue;
    }
    if (f.minMin !== null && f.minMin !== undefined){
      if (!Number.isFinite(vMin) || vMin < f.minMin) continue;
    }
    if (f.minMax !== null && f.minMax !== undefined){
      if (!Number.isFinite(vMin) || vMin > f.minMax) continue;
    }

    if (currentCellFilter && p.has_latlon){
      if (p.lat < currentCellFilter.minLat || p.lat >= currentCellFilter.maxLat) continue;
      if (p.lon < currentCellFilter.minLon || p.lon >= currentCellFilter.maxLon) continue;
    }
    out.push(p);
  }
  return out;
}

function groupingMetersForZoom(z){
  // Broader low-zoom cells keep the nationwide view readable.
  if (z >= 15) return 0;       // 0 = no grouping (individual points)
  if (z >= 14) return 120;     // 120m grid
  if (z >= 13) return 260;     // 260m grid
  if (z >= 12) return 700;     // 700m grid
  if (z >= 11) return 1600;    // 1.6km grid
  if (z >= 10) return 3500;    // 3.5km grid
  if (z >= 9) return 8000;     // 8km grid
  if (z >= 8) return 15000;    // 15km grid
  return 30000;                // 30km grid
}

function latLonToMeters(lat, lon){
  // EPSG:3857 (Web Mercator), meters
  const R = 6378137.0;
  const x = R * (lon * Math.PI / 180.0);
  const y = R * Math.log(Math.tan(Math.PI/4 + (lat * Math.PI / 180.0)/2));
  return [x, y];
}

function metersToLatLon(x, y){
  // WebMercator inverse (EPSG:3857 meters -> WGS84 degrees)
  const lon = (x / EARTH_R) * (180.0 / Math.PI);
  const lat = (2.0 * Math.atan(Math.exp(y / EARTH_R)) - Math.PI / 2.0) * (180.0 / Math.PI);
  return [lat, lon];
}

function computeGridAggMeters(points, cellMeters){
  if (!cellMeters || cellMeters <= 0){
    // No grouping: return per-point "cells" so heat can still normalize
    const out = [];
    for (const p of points){
      if (!isValidLatLon(Number(p.lat), Number(p.lon))) continue;
      out.push({
        key: String(p.__i),
        gx: null, gy: null,
        count: 1,
        sumScore: p.score,
        avgScore: p.score,
        center: [p.lat, p.lon],
        items: [p],
        bounds: null
      });
    }
    return out;
  }
  const cells = new Map();
  for (const p of points){
    if (!isValidLatLon(Number(p.lat), Number(p.lon))) continue;
    const xy = latLonToMeters(p.lat, p.lon);
    const gx = Math.floor(xy[0] / cellMeters);
    const gy = Math.floor(xy[1] / cellMeters);
    const key = gx + "," + gy;
    let c = cells.get(key);
    if (!c){
      c = {key, gx, gy, count: 0, sumScore: 0, sumLat: 0, sumLon: 0, items: []};
      cells.set(key, c);
    }
    c.count += 1;
    c.sumScore += p.score;
    c.sumLat += p.lat;
    c.sumLon += p.lon;
    c.items.push(p);
  }
  const arr = [];
  for (const c of cells.values()){
    const center = [c.sumLat / c.count, c.sumLon / c.count];

    // bounds for this cell in lat/lon (approx)
    const x0 = c.gx * cellMeters, x1 = (c.gx + 1) * cellMeters;
    const y0 = c.gy * cellMeters, y1 = (c.gy + 1) * cellMeters;
    const ll0 = metersToLatLon(x0, y0);
    const ll1 = metersToLatLon(x1, y1);
    const minLat = Math.min(ll0[0], ll1[0]);
    const maxLat = Math.max(ll0[0], ll1[0]);
    const minLon = Math.min(ll0[1], ll1[1]);
    const maxLon = Math.max(ll0[1], ll1[1]);

    arr.push({
      key: c.key,
      gx: c.gx, gy: c.gy,
      count: c.count,
      sumScore: c.sumScore,
      avgScore: c.sumScore / c.count,
      center,
      items: c.items,
      bounds: {minLat, maxLat, minLon, maxLon}
    });
  }
  return arr;
}
function quantiles(values, ps=[0.5,0.75,0.9]){
  const v = values.filter(x=>Number.isFinite(x)).sort((a,b)=>a-b);
  if (v.length === 0) return ps.map(_=>NaN);
  const q = (p) => {
    const idx = (v.length - 1) * p;
    const lo = Math.floor(idx);
    const hi = Math.ceil(idx);
    if (lo === hi) return v[lo];
    const w = idx - lo;
    return v[lo]*(1-w) + v[hi]*w;
  };
  return ps.map(q);
}

function updateLegendFromCells(cells){
  const avgs = cells.map(c=>c.avgScore);
  const [p50,p75,p90] = quantiles(avgs);
  document.getElementById("quantText").textContent =
    `AVG quantiles: P50=${fmt(p50,1)} / P75=${fmt(p75,1)} / P90=${fmt(p90,1)} (score)`;
}

function updatePointsLayer(points){
  pointsLayer.clearLayers();
  window.__POINT_MARKERS_BY_I__ = {};
  for (const p of points){
    if (!isValidLatLon(Number(p.lat), Number(p.lon))) continue;
    const m = makePointMarker(p);
    m.bindPopup(buildPopup(p), { maxWidth: 420 });
    if (p && (p.__i !== undefined)) window.__POINT_MARKERS_BY_I__[p.__i] = m;
    pointsLayer.addLayer(m);
  }
}

function updateHeatLayers(points, cellMeters){
  const cells = computeGridAggMeters(points, cellMeters);
  if (cells.length === 0){
    if (heatAvgLayer && typeof heatAvgLayer.setLatLngs === "function") heatAvgLayer.setLatLngs([]);
    if (heatFreqLayer && typeof heatFreqLayer.setLatLngs === "function") heatFreqLayer.setLatLngs([]);
    document.getElementById("quantText").textContent = "AVG quantiles: P50=? / P75=? / P90=?";
    window.__LAST_CELLS__ = [];
    window.__LAST_CELL_METERS__ = cellMeters || 0;
    return [];
  }
  const maxCount = Math.max(...cells.map(c=>c.count));
  const maxAvg = Math.max(...cells.map(c=>c.avgScore));
  const heatAvg = [];
  const heatFreq = [];
  for (const c of cells){
    const wAvg = (maxAvg > 0) ? clamp(c.avgScore / maxAvg, 0, 1) : 0;
    heatAvg.push([c.center[0], c.center[1], wAvg]);
    const wFreq = (maxCount > 0) ? Math.pow(c.count / maxCount, 0.55) : 0;
    heatFreq.push([c.center[0], c.center[1], wFreq]);
  }
  if (heatAvgLayer && typeof heatAvgLayer.setLatLngs === "function") heatAvgLayer.setLatLngs(heatAvg);
  if (heatFreqLayer && typeof heatFreqLayer.setLatLngs === "function") heatFreqLayer.setLatLngs(heatFreq);
  updateLegendFromCells(cells);
  window.__LAST_CELLS__ = cells;
  window.__LAST_CELL_METERS__ = cellMeters || 0;
  // Build an index so grouped-cell markers can show underlying point details
  window.__CELL_INDEX__ = {};
  for (const c of cells){
    if (c && c.key) window.__CELL_INDEX__[c.key] = c;
  }
  return cells;
}

function updateAggLayer(cells, cellMeters){
  if (!aggLayer) aggLayer = L.layerGroup();
  aggLayer.clearLayers();
  if (!cells || cells.length === 0) return;
  const maxCount = Math.max(...cells.map(c=>c.count));
  for (const c of cells){
    if (!isValidLatLon(Number(c.center?.[0]), Number(c.center?.[1]))) continue;
    const m = makeAggMarker(c, maxCount);
    m.bindPopup(
      `<b>Grouped cell</b><br/>count=${c.count}<br/>avg_score=${fmt(c.avgScore,1)}<br/>cell=${cellMeters}m
       <div style="margin-top:10px; display:flex; gap:8px; flex-wrap:wrap;">
         <button class="btn small" data-cell-action="details" data-cell-key="${c.key}">Details</button>
         <button class="btn small" data-cell-action="zoom" data-cell-key="${c.key}">Zoom</button>
        </div>`,
      { maxWidth: 420 }
    );
    m.on('popupopen', (ev)=>{
      // Get the popup root element — fall back to getPopup() in case ev.popup is missing
      const popup = ev?.popup || m.getPopup();
      const root = popup && typeof popup.getElement === 'function' ? popup.getElement() : null;
      if (!root) return;

      // Clone-and-replace pattern: guarantees a clean listener on every popup open,
      // avoiding stale Leaflet DomEvent bindings and double-bind issues.
      ['details', 'zoom'].forEach((action)=>{
        const old = root.querySelector(`[data-cell-action="${action}"]`);
        if (!old) return;
        const fresh = old.cloneNode(true);
        old.parentNode.replaceChild(fresh, old);
        fresh.addEventListener('click', (e)=>{
          e.stopPropagation();
          e.preventDefault();
          m.closePopup();
          if (action === 'details'){
            try{ showCellDetails(c.key); }catch(_e){}
          } else {
            try{ zoomToCell(c.key); }catch(_e){}
          }
        });
      });
    });
    // NOTE: No m.on('click', showCellDetails) here — that would open the Analysis panel
    // immediately on every marker click, covering the map and making the popup unreachable.
    // The Details button inside the popup (bound above via popupopen) is the correct trigger.
    aggLayer.addLayer(m);
  }
}




function clearCellDetails(){
  const card = document.getElementById('cellDetailsCard');
  if (card) card.style.display = 'none';
  const meta = document.getElementById('cellDetailsMeta'); if (meta) meta.textContent = '';
  const list = document.getElementById('cellDetailsList'); if (list) list.innerHTML = '';
}

function zoomToCell(key){
  try{
    const c = window.__CELL_INDEX__ ? window.__CELL_INDEX__[key] : null;
    if (!c || !c.bounds) return;
    const b = c.bounds;
    const bounds = L.latLngBounds([[b.minLat, b.minLon],[b.maxLat, b.maxLon]]);
    map.fitBounds(bounds.pad(0.15));
  }catch(e){}
}

function showCellDetails(key){
  const c = window.__CELL_INDEX__ ? window.__CELL_INDEX__[key] : null;
  if (!c){
    alert('Cell details not available (try Apply or zoom).');
    return;
  }
  openPanel('panelAnalysis');
  const card = document.getElementById('cellDetailsCard');
  const meta = document.getElementById('cellDetailsMeta');
  const list = document.getElementById('cellDetailsList');
  if (!card || !meta || !list) return;

  const n = (c.items || []).length;
  meta.textContent = `cell=${key} / n=${n} / avg_score=${fmt(c.avgScore,1)} / grid=${window.__LAST_CELL_METERS__||0}m`;

  const items = (c.items || []).slice().sort((a,b)=>{
    const ds = (b.score||0) - (a.score||0);
    if (ds !== 0) return ds;
    return String(a.No||'').localeCompare(String(b.No||''), 'ja');
  });

  const maxShow = 250;
  const shown = items.slice(0, maxShow);

  list.innerHTML = shown.map(p=>{
    const dx = p['分類名'] || p['分類'] || 'Unknown';
    const lvl = p['ニアミスレベル'] || '-';
    const road = p['道路'] || '-';
    const loc  = p['ロケーション'] || '-';
    const cw   = p['car_week'] || '-';
    const cid  = p.car_id || '-';
    const no   = (p.No !== undefined && p.No !== null) ? p.No : '-';
    const sc   = fmt(p.score,1);
    const ii   = p.__i;
    return `<div class="barRow" style="cursor:pointer;" data-idx="${ii}">
      <div style="min-width:0;">
        <div style="font-weight:900; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">No:${no}　${dx}　${lvl}　score:${sc}</div>
        <div class="muted" style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${road} / ${loc}　|　${cw} / ${cid}</div>
      </div>
      <div style="text-align:right; font-weight:900;">↗</div>
    </div>`;
  }).join('') + (n>maxShow ? `<div class="hint">... and ${n-maxShow} more (zoom in to see individual popups)</div>` : '');

  card.style.display = 'block';
  const panelBody = card.closest('.body');
  if (panelBody && typeof panelBody.scrollTo === 'function'){
    setTimeout(()=>{
      try{
        const targetTop = Math.max(0, card.offsetTop - 10);
        panelBody.scrollTo({ top: targetTop, behavior: 'smooth' });
      }catch(_err){}
    }, 0);
  } else {
    setTimeout(()=>{
      try{ card.scrollIntoView({ behavior: 'smooth', block: 'start' }); }catch(_err){}
    }, 0);
  }

  list.querySelectorAll('.barRow').forEach(el=>{
    el.addEventListener('click', ()=>{
      const ii = Number(el.getAttribute('data-idx'));
      const p = ALL_ROWS.find(x => x.__i === ii);
      if (p && p.has_latlon){
        map.setView([p.lat, p.lon], Math.max(map.getZoom(), 16), {animate:true});
        try{
          const mk = window.__POINT_MARKERS_BY_I__ ? window.__POINT_MARKERS_BY_I__[ii] : null;
          setTimeout(()=>{ const mk2 = window.__POINT_MARKERS_BY_I__ ? window.__POINT_MARKERS_BY_I__[ii] : null; if (mk2) mk2.openPopup(); }, 350);
        }catch(e){}
      }
    });
  });
}

function updateStatus(filters, matchedTotal, shownPoints, cellCount){
  const shownCount = Array.isArray(shownPoints) ? shownPoints.length : Number(shownPoints || 0);
  const missing = Math.max(0, matchedTotal - shownCount);
  const grouping = (filters.groupMeters <= 0) ? "individual" : `${filters.groupMeters}m`;
  renderStatusText({
    baseText: `表示中: points=${shownCount} / matched=${matchedTotal} / missing_latlon=${missing} / gridCells=${cellCount} / grouping=${grouping} (zoom=${filters.zoom})`,
    shownPointsLen: shownCount,
    pointsLayerCount: layerCount(pointsLayer),
    aggLayerCount: layerCount(aggLayer),
  });
  // Update #debug marker popup so shownPoints count is visible on-map
  if (window.__debugMarker__){
    try{
      window.__debugMarker__.bindPopup(
        `DEBUG: Tokyo<br/>shownPoints=${shownCount}<br/>matched=${matchedTotal}<br/>missing_latlon=${missing}<br/>gridCells=${cellCount}`
      );
    }catch(_){}
  }
}


function zoomToPoints(points){
  const validPoints = getDrawablePoints(points);
  if (!validPoints || validPoints.length === 0) return;
  const latlngs = validPoints.map(p => [p.lat, p.lon]);
  const bounds = L.latLngBounds(latlngs);
  if (bounds.isValid()){
    map.fitBounds(bounds.pad(0.12));
  }
}

function toggleLayer(pillId, layer){
  const pill = document.getElementById(pillId);
  pill.addEventListener("click", () => {
    pill.classList.toggle("active");
    const on = pill.classList.contains("active");
    if (layer){
      if (on) layer.addTo(map); else map.removeLayer(layer);
    }
    if (pillId === "togglePoints"){
      if (on) pointsLayer.addTo(map); else map.removeLayer(pointsLayer);
    }
    applyAll();
  });
}

/* ===== Shared road-name helpers (global scope — used by buildAnalysis AND initRoutesUI) ===== */

// IMPORTANT: This function MUST be at global scope. It was previously defined
// inside initRoutesUI(), which caused a ReferenceError in buildAnalysis().
function isMissingRoadName(name){
  const s = String(name || '').trim();
  return !s || s === 'Unknown' || s === '不明';
}

/* ===== Analysis building (v59 style) ===== */

const SPEED_CLICK_HALF_RANGE = 5.0;
const ANALYSIS_MIN_DX_ROAD_CASES = 5;
const ANALYSIS_MIN_SPEED_BAND_CASES = 12;
const ANALYSIS_SPEED_BAND_KMH = 20;
const ANALYSIS_MIN_OVERINDEX_ROAD_CASES = 12;
const ANALYSIS_MIN_OVERINDEX_PAIR_CASES = 4;
const ANALYSIS_OVERINDEX_PRIOR = 12;

function setSpeedMinMax(kind, a, b){
  const setPair = (minId, maxId)=>{
    const elMin = document.getElementById(minId);
    const elMax = document.getElementById(maxId);
    if (elMin) elMin.value = (a === null || a === undefined) ? "" : String(a);
    if (elMax) elMax.value = (b === null || b === undefined) ? "" : String(b);
  };
  if (kind === "avg") setPair("speedAvgMin","speedAvgMax");
  if (kind === "max") setPair("speedMaxMin","speedMaxMax");
  if (kind === "min") setPair("speedMinMin","speedMinMax");
}

function speedClick(kind, v){
  if (!Number.isFinite(v)) return;
  const a = Math.max(0, v - SPEED_CLICK_HALF_RANGE);
  const b = v + SPEED_CLICK_HALF_RANGE;
  setSpeedMinMax(kind, a.toFixed(1), b.toFixed(1));
  currentCellFilter = null;
  applyAll();
  openPanel("panelFilter");
}

function buildBarList(containerId, rows, clickFn, labelFmt){
  const el = document.getElementById(containerId);
  if (!el) return;
  if (!rows || rows.length === 0){
    el.innerHTML = `<div class="hint">該当なし</div>`;
    return;
  }
  const maxV = Math.max(...rows.map(r=>r.n));
  el.innerHTML = rows.map(r=>{
    const w = (maxV>0) ? Math.round((r.n/maxV)*100) : 0;
    const right = labelFmt ? labelFmt(r) : `<b>${r.n}</b>`;
    return `
      <div class="barRow" data-key="${encodeURIComponent(r.key)}" title="click">
        <div style="min-width:0;">
          <div class="barName" title="${r.label}">${r.label}</div>
          <div class="muted">${r.sub || ""}</div>
        </div>
        <div class="miniBar"><div style="width:${w}%;"></div></div>
        <div style="text-align:right;">${right}</div>
      </div>
    `;
  }).join("");

  el.querySelectorAll(".barRow").forEach(row=>{
    row.addEventListener("click", ()=>{
      const key = decodeURIComponent(row.getAttribute("data-key"));
      clickFn(key);
    });
  });
}

function buildAnalysis(filteredRows, shownPoints, f){
  const total = filteredRows.length;
  const shown = shownPoints.length;
  const missing = Math.max(0, total - shown);
  const roadRowsSource = shownPoints;
  const roadTotal = roadRowsSource.length;

  const TOP_N = Number(document.getElementById("analysisTopN")?.value || 14);
  const hideUnknownRoads = (document.getElementById("hideUnknownRoads")?.checked ?? true);

  document.getElementById("analysisMeta").textContent =
    `filters: Road=${f.road||"全て"}, Loc=${f.loc||"全て"}, DX=${f.dxSet.size}, score=[${f.smin}..${f.smax}] ／ matched=${total}, shown=${shown}, missing_latlon=${missing}`;

  const dxBreakdownHint = document.getElementById("dxBreakdown")?.parentElement?.querySelector(".hint");
  if (dxBreakdownHint) dxBreakdownHint.textContent = "DX breakdown for the current filter. Click a DX to isolate it on the map and refresh the rest of the analysis instantly.";
  const speedBandHint = document.getElementById("speedBandDxLead")?.parentElement?.querySelector(".hint");
  if (speedBandHint) speedBandHint.textContent = "Bands with fewer than 12 incidents are hidden. Bars compare the leading DX share inside each speed band, with light smoothing to reduce small-sample bias.";

  // DX breakdown (counts)
  const dxCounts = new Map();
  for (const p of filteredRows){
    const code = p["分類"] || "Unknown";
    dxCounts.set(code, (dxCounts.get(code)||0)+1);
  }
  const dxRows = Array.from(dxCounts.entries())
    .map(([code,n])=>{
      const label = (DX_LABEL.get(code) || code);
      const pct = total>0 ? (n*100/total) : 0;
      return {key: code, label: `${code}`, sub: label, n, pct};
    })
    .sort((a,b)=> b.n - a.n)
    .slice(0, TOP_N);

  buildBarList("dxBreakdown", dxRows,
    (dxCode)=>{
      setOnlyDx(dxCode);
      currentCellFilter = null;
      applyAll();
      // keep panel open, and zoom to filtered points
      const f2 = getFilters();
      const rows2 = filterRows(ALL_ROWS, f2);
      const pts2 = getDrawablePoints(rows2);
      zoomToPoints(pts2);
    },
    (r)=> `<b>${r.n}</b> <span class="muted">(${r.pct.toFixed(1)}%)</span>`
  );

  // Top road within each DX (mapped rows only, compare within-DX share)
  try{
  const dxRoadBuckets = new Map();
  for (const p of roadRowsSource){
    const dxCode = String(p.dx_code || "Unknown").trim() || "Unknown";
    const road = String(p.road || "").trim();
    if (isMissingRoadName(road)) continue;
    let bucket = dxRoadBuckets.get(dxCode);
    if (!bucket){
      bucket = {dxCode, total: 0, roads: new Map(), scoreSum: 0};
      dxRoadBuckets.set(dxCode, bucket);
    }
    bucket.total += 1;
    bucket.scoreSum += Number(p.score) || 0;
    bucket.roads.set(road, (bucket.roads.get(road) || 0) + 1);
  }
  const dxRoadRowsAll = Array.from(dxRoadBuckets.values()).map((bucket)=>{
    let topRoad = "";
    let topCount = 0;
    for (const [road, n] of bucket.roads.entries()){
      if (n > topCount || (n === topCount && String(road).localeCompare(String(topRoad), "ja") < 0)){
        topRoad = road;
        topCount = n;
      }
    }
    const distinctRoads = Math.max(1, bucket.roads.size);
    const rawShare = bucket.total > 0 ? (topCount / bucket.total) : 0;
    const smoothShare = bucket.total > 0 ? ((topCount + 1) / (bucket.total + distinctRoads)) : 0;
    return {
      key: JSON.stringify({dx: bucket.dxCode, road: topRoad}),
      label: bucket.dxCode,
      sub: `${dxLabelFor({dx_code: bucket.dxCode})} -> ${topRoad}`,
      n: smoothShare,
      rawShare,
      topCount,
      dxTotal: bucket.total,
      lowSample: bucket.total < ANALYSIS_MIN_DX_ROAD_CASES,
    };
  }).filter((row)=> row.topCount > 0);
  const dxRoadRowsPreferred = dxRoadRowsAll.filter((row)=> !row.lowSample);
  const dxRoadRows = (dxRoadRowsPreferred.length > 0 ? dxRoadRowsPreferred : dxRoadRowsAll)
    .sort((a,b)=> (b.dxTotal - a.dxTotal) || (b.rawShare - a.rawShare) || String(a.label).localeCompare(String(b.label), "ja"))
    .slice(0, Math.max(6, Math.min(TOP_N, dxRoadRowsAll.length || TOP_N)));
  buildBarList("dxRoadLead", dxRoadRows,
    (payload)=>{
      try{
        const parsed = JSON.parse(payload || "{}");
        if (!parsed.dx || !parsed.road) return;
        setOnlyDx(String(parsed.dx));
        const sel = document.getElementById("roadSelect");
        if (sel) sel.value = String(parsed.road);
        currentCellFilter = null;
        applyAll();
        const f2 = getFilters();
        const rows2 = filterRows(ALL_ROWS, f2);
        const pts2 = getDrawablePoints(rows2);
        zoomToPoints(pts2);
      }catch(_err){}
    },
    (r)=> `<b>${(r.rawShare*100).toFixed(1)}%</b> <span class="muted">(${r.topCount}/${r.dxTotal})</span>${r.lowSample ? ' <span class="muted">low n</span>' : ''}`
  );
  }catch(_e){ console.error("[buildAnalysis:dxRoadLead]", _e); }

  // DX over-index by road (smoothed lift vs overall DX baseline)
  try{
  const namedRoadRows = roadRowsSource.filter((p)=>{
    const road = String(p.road || "").trim();
    const dxCode = String(p.dx_code || "Unknown").trim() || "Unknown";
    return !isMissingRoadName(road) && dxCode !== "Unknown";
  });
  const overallDxNamed = new Map();
  const roadDxBuckets = new Map();
  for (const p of namedRoadRows){
    const road = String(p.road || "").trim();
    const dxCode = String(p.dx_code || "Unknown").trim() || "Unknown";
    overallDxNamed.set(dxCode, (overallDxNamed.get(dxCode) || 0) + 1);
    let bucket = roadDxBuckets.get(road);
    if (!bucket){
      bucket = {road, total: 0, dxCounts: new Map()};
      roadDxBuckets.set(road, bucket);
    }
    bucket.total += 1;
    bucket.dxCounts.set(dxCode, (bucket.dxCounts.get(dxCode) || 0) + 1);
  }
  const totalNamedRoadRows = namedRoadRows.length;
  const dxRoadLiftRows = [];
  for (const bucket of roadDxBuckets.values()){
    if (bucket.total < ANALYSIS_MIN_OVERINDEX_ROAD_CASES) continue;
    for (const [dxCode, pairCount] of bucket.dxCounts.entries()){
      if (pairCount < ANALYSIS_MIN_OVERINDEX_PAIR_CASES) continue;
      const overallDxCount = Number(overallDxNamed.get(dxCode) || 0);
      if (overallDxCount <= 0 || totalNamedRoadRows <= 0) continue;
      const baselineShare = overallDxCount / totalNamedRoadRows;
      if (!(baselineShare > 0)) continue;
      const observedShare = pairCount / bucket.total;
      const smoothedShare = (pairCount + (ANALYSIS_OVERINDEX_PRIOR * baselineShare)) / (bucket.total + ANALYSIS_OVERINDEX_PRIOR);
      const lift = smoothedShare / baselineShare;
      if (!(lift > 1.0)) continue;
      const excessPp = (observedShare - baselineShare) * 100.0;
      dxRoadLiftRows.push({
        key: JSON.stringify({dx: dxCode, road: bucket.road}),
        label: `${dxCode} -> ${bucket.road}`,
        sub: `road share ${(observedShare*100).toFixed(1)}% vs base ${(baselineShare*100).toFixed(1)}%`,
        n: lift,
        lift,
        pairCount,
        roadTotal: bucket.total,
        baselineShare,
        observedShare,
        excessPp,
      });
    }
  }
  dxRoadLiftRows.sort((a,b)=>
    (b.lift - a.lift) ||
    (b.pairCount - a.pairCount) ||
    (b.roadTotal - a.roadTotal) ||
    String(a.label).localeCompare(String(b.label), "ja")
  );

  // Group by road: one row per road showing its most over-indexed DX.
  // Label = road name; n = lift of dominant DX.
  const _byRoad = new Map();
  for (const row of dxRoadLiftRows){
    let _parsed; try{ _parsed = JSON.parse(row.key || "{}"); }catch(_e){ continue; }
    const _road = String(_parsed.road || "").trim();
    if (!_road) continue;
    if (!_byRoad.has(_road) || row.lift > _byRoad.get(_road).lift){
      _byRoad.set(_road, {...row, _road, _dx: String(_parsed.dx || "")});
    }
  }
  const dxRoadGroupedRows = Array.from(_byRoad.values())
    .sort((a,b)=>
      (b.lift - a.lift) ||
      (b.roadTotal - a.roadTotal) ||
      String(a._road).localeCompare(String(b._road), "ja")
    )
    .slice(0, TOP_N)
    .map((row)=>({
      ...row,
      key: JSON.stringify({dx: row._dx, road: row._road}),
      label: row._road,
      sub: `${row._dx} ${dxLabelFor({dx_code: row._dx})}: ${((row.observedShare||0)*100).toFixed(1)}% on road vs ${((row.baselineShare||0)*100).toFixed(1)}% overall (+${((row.excessPp)||0).toFixed(1)}pp)`,
    }));

  buildBarList("dxRoadLift", dxRoadGroupedRows,
    (payload)=>{
      try{
        const parsed = JSON.parse(payload || "{}");
        if (!parsed.dx || !parsed.road) return;
        setOnlyDx(String(parsed.dx));
        const sel = document.getElementById("roadSelect");
        if (sel) sel.value = String(parsed.road);
        currentCellFilter = null;
        applyAll();
        const f2 = getFilters();
        const rows2 = filterRows(ALL_ROWS, f2);
        const pts2 = getDrawablePoints(rows2);
        zoomToPoints(pts2);
      }catch(_err){}
    },
    (r)=> `<b>${r.lift.toFixed(2)}x</b> <span class="muted">(${r.pairCount}/${r.roadTotal} cases)</span>`
  );
  }catch(_e){ console.error("[buildAnalysis:dxRoadLift]", _e); }

  // Top roads by count
  try{
  const rc = new Map();
  for (const p of roadRowsSource){
    const road = p["道路"] || "Unknown";
    rc.set(road, (rc.get(road)||0)+1);
  }
  const roadRows = Array.from(rc.entries())
    .filter(([road,_])=> !(hideUnknownRoads && road === "Unknown"))
    .map(([road,n])=>({key: road, label: road, sub:"", n}))
    .sort((a,b)=> b.n - a.n)
    .slice(0, TOP_N);

  buildBarList("roadCount", roadRows,
    (road)=>{
      const sel = document.getElementById("roadSelect");
      if (sel) sel.value = (road === "全て") ? "" : road;
      currentCellFilter = null;
      applyAll();
      const f2 = getFilters();
      const rows2 = filterRows(ALL_ROWS, f2);
      const pts2 = getDrawablePoints(rows2);
      zoomToPoints(pts2);
    },
    (r)=> `<b>${r.n}</b> <span class="muted">(${(roadTotal? r.n*100/roadTotal:0).toFixed(1)}%)</span>`
  );
  }catch(_e){ console.error("[buildAnalysis:roadCount]", _e); }


  // Top roads per 1km (requires exposure_by_roadgroup.csv)
  try{
  const expKeys = Object.keys(EXPOSURE || {});
  const rateCard = document.getElementById("roadRateCard");
  if (expKeys.length === 0){
    if (rateCard) rateCard.style.display = "none";
    const rr = document.getElementById("roadRate");
    if (rr) rr.innerHTML = '<div class="muted">exposure 未読込（距離/時間の分母が無いため、率は表示しません）</div>';
  } else {
    if (rateCard) rateCard.style.display = "";
    const rateRows = [];
    for (const [road, n] of rc.entries()){
      if (hideUnknownRoads && road === "Unknown") continue;
      const ex = EXPOSURE[road];
      if (!ex) continue;
      const km = exposureKmValue(ex);
      if (!Number.isFinite(km) || km <= 0) continue;
      const perKm = n / km; // cases per 1km
      const observedKm = Number(ex?.distance_km);
      const observedText = (Number.isFinite(observedKm) && Math.abs(observedKm - km) > 0.01)
        ? ` / obs ${observedKm.toFixed(1)}km`
        : '';
      rateRows.push({key: road, label: road, sub: `exp ${km.toFixed(1)}km${observedText}`, n: perKm, count: n, km, observedKm});
    }
    rateRows.sort((a,b)=> b.n - a.n);
    const topRate = rateRows.slice(0, TOP_N);
    buildBarList("roadRate", topRate,
      (road)=>{
        const sel = document.getElementById("roadSelect");
        if (sel) sel.value = (road === "全て") ? "" : road;
        currentCellFilter = null;
        applyAll();
      },
      (r)=> `<b>${r.n.toFixed(2)}</b> <span class="muted">cases/km</span> <span class="muted">(count=${r.count}, exp=${r.km.toFixed(1)}km)</span>`
    );
  }
  }catch(_e){ console.error("[buildAnalysis:roadRate]", _e); }


  // Speed histograms (avg / max / min) - clickable bins set speed filters
  function renderSpeedHist(containerId, fieldName, kind){
    const vals = [];
    for (const p of filteredRows){
      const v = Number(p[fieldName]);
      if (Number.isFinite(v)) vals.push(v);
    }
    if (vals.length === 0){
      const el = document.getElementById(containerId);
      if (el) el.innerHTML = `<div class="hint">該当なし</div>`;
      return;
    }
    const maxV = Math.max(...vals);
    const minV = Math.min(...vals);
    const lo = Math.floor(Math.max(0, minV/10)*10);
    const hi = Math.ceil(maxV/10)*10;
    const step = 10;
    const edges = [];
    for (let x = lo; x <= hi + 1e-9; x += step) edges.push(x);
    if (edges.length < 2) { edges.push(edges[0] + step); }
    const bc = new Array(edges.length-1).fill(0);
    for (const v of vals){
      let bi = Math.floor((v - edges[0]) / step);
      if (bi < 0) bi = 0;
      if (bi >= bc.length) bi = bc.length - 1;
      bc[bi] += 1;
    }
    const sRows = bc.map((n,i)=>{
      const a = edges[i], b = edges[i+1];
      const la = Math.round(a), lb = Math.round(b);
      return {key: `${i}`, label: `${la}–${lb}`, sub:"km/h", n, a, b};
    }).filter(r=>r.n>0);

    buildBarList(containerId, sRows,
      (k)=>{
        const i = Number(k);
        if (!Number.isFinite(i)) return;
        const row = sRows.find(r=>Number(r.key)===i) || null;
        // row might be missing if filtered; instead use edges
        const a = edges[i], b = edges[i+1];
        if (!Number.isFinite(a) || !Number.isFinite(b)) return;
        setSpeedMinMax(kind, a, b);
        currentCellFilter = null;
        applyAll();
        openPanel("panelFilter");
      },
      (r)=> `<b>${r.n}</b>`
    );
  }

  try{ renderSpeedHist("speedHistAvg", "平均速度", "avg"); }catch(_e){ console.error("[buildAnalysis:speedHistAvg]", _e); }
  try{ renderSpeedHist("speedHistMax", "最高速度", "max"); }catch(_e){ console.error("[buildAnalysis:speedHistMax]", _e); }
  try{ renderSpeedHist("speedHistMin", "最低速度", "min"); }catch(_e){ console.error("[buildAnalysis:speedHistMin]", _e); }


  // Over-indexed DX by avg-speed band.
  // Instead of showing the most COMMON DX per band (which always shows カットイン because it
  // dominates overall), show the DX that is most OVER-REPRESENTED in that band relative to
  // its overall frequency — i.e., the highest lift = bandShare / baselineShare.
  // This highlights which incident type is unusually elevated at each speed range.
  try{
  // Step 1: overall DX baseline across all rows that have speed data
  const _sbOverallDx = new Map();
  let _sbOverallTotal = 0;
  for (const p of filteredRows){
    if (!Number.isFinite(Number(p.speed_avg))) continue;
    const dxCode = String(p.dx_code || "Unknown").trim() || "Unknown";
    _sbOverallDx.set(dxCode, (_sbOverallDx.get(dxCode) || 0) + 1);
    _sbOverallTotal++;
  }

  // Step 2: bucket by speed band
  const speedBandBuckets = new Map();
  for (const p of filteredRows){
    const avgSpeed = Number(p.speed_avg);
    if (!Number.isFinite(avgSpeed)) continue;
    const dxCode = String(p.dx_code || "Unknown").trim() || "Unknown";
    const bandLo = Math.floor(Math.max(0, avgSpeed) / ANALYSIS_SPEED_BAND_KMH) * ANALYSIS_SPEED_BAND_KMH;
    const bandHi = bandLo + ANALYSIS_SPEED_BAND_KMH;
    const bandKey = `${bandLo}:${bandHi}`;
    let bucket = speedBandBuckets.get(bandKey);
    if (!bucket){
      bucket = {bandLo, bandHi, total: 0, dxCounts: new Map()};
      speedBandBuckets.set(bandKey, bucket);
    }
    bucket.total += 1;
    bucket.dxCounts.set(dxCode, (bucket.dxCounts.get(dxCode) || 0) + 1);
  }

  // Step 3: for each band, find DX with highest smoothed lift vs overall baseline
  const speedBandRows = Array.from(speedBandBuckets.values())
    .filter((bucket)=> bucket.total >= ANALYSIS_MIN_SPEED_BAND_CASES)
    .map((bucket)=>{
      let topDx = "Unknown";
      let topLift = -Infinity;
      let topCount = 0;
      let topBaseline = 0;
      for (const [dxCode, n] of bucket.dxCounts.entries()){
        if (dxCode === "Unknown") continue;
        const overallN = _sbOverallDx.get(dxCode) || 0;
        if (overallN <= 0 || _sbOverallTotal <= 0) continue;
        const baseline = overallN / _sbOverallTotal;
        // Laplace-smoothed lift: pull bandShare toward baseline to penalise tiny counts
        const smoothedBandShare = (n + ANALYSIS_OVERINDEX_PRIOR * baseline) / (bucket.total + ANALYSIS_OVERINDEX_PRIOR);
        const lift = smoothedBandShare / baseline;
        if (lift > topLift || (lift === topLift && String(dxCode).localeCompare(String(topDx), "ja") < 0)){
          topDx = dxCode;
          topLift = lift;
          topCount = n;
          topBaseline = baseline;
        }
      }
      // Fall back to raw count winner if no valid lift found
      if (!Number.isFinite(topLift) || topLift <= 0){
        for (const [dxCode, n] of bucket.dxCounts.entries()){
          if (n > topCount || (n === topCount && String(dxCode).localeCompare(String(topDx), "ja") < 0)){
            topDx = dxCode; topCount = n;
          }
        }
        topLift = 1.0;
        topBaseline = _sbOverallTotal > 0 ? ((_sbOverallDx.get(topDx) || 0) / _sbOverallTotal) : 0;
      }
      const bandShare = bucket.total > 0 ? topCount / bucket.total : 0;
      const excessPp = (bandShare - topBaseline) * 100;
      return {
        key: JSON.stringify({dx: topDx, a: bucket.bandLo, b: bucket.bandHi}),
        label: `${bucket.bandLo}-${bucket.bandHi} km/h`,
        sub: `${topDx} ${dxLabelFor({dx_code: topDx})}`,
        n: Math.max(0, topLift),
        lift: topLift,
        rawShare: bandShare,
        baselineShare: topBaseline,
        excessPp,
        topCount,
        bandTotal: bucket.total,
        bandLo: bucket.bandLo,
      };
    })
    .sort((a,b)=> (a.bandLo - b.bandLo) || (b.bandTotal - a.bandTotal));

  const speedBandHintText =
    `Bands with fewer than ${ANALYSIS_MIN_SPEED_BAND_CASES} incidents are hidden. ` +
    `Shows the DX most over-represented in each speed band relative to its overall frequency (lift = band% ÷ global%). ` +
    `This removes the bias from high-frequency DX types like カットイン dominating every band by raw count.`;
  if (speedBandHint) speedBandHint.textContent = speedBandHintText;

  if (speedBandRows.length === 0){
    const elBand = document.getElementById("speedBandDxLead");
    if (elBand){
      elBand.innerHTML = `<div class="hint">No avg-speed band has at least ${ANALYSIS_MIN_SPEED_BAND_CASES} incidents in the current filter.</div>`;
    }
  } else {
    buildBarList("speedBandDxLead", speedBandRows,
      (payload)=>{
        try{
          const parsed = JSON.parse(payload || "{}");
          const a = Number(parsed.a);
          const b = Number(parsed.b);
          if (!parsed.dx || !Number.isFinite(a) || !Number.isFinite(b)) return;
          setOnlyDx(String(parsed.dx));
          setSpeedMinMax("avg", a, b);
          currentCellFilter = null;
          applyAll();
          const f2 = getFilters();
          const rows2 = filterRows(ALL_ROWS, f2);
          const pts2 = getDrawablePoints(rows2);
          zoomToPoints(pts2);
        }catch(_err){}
      },
      (r)=> `<b>${r.lift.toFixed(2)}x</b> <span class="muted">(${r.topCount}/${r.bandTotal}, ${(r.rawShare*100).toFixed(1)}% vs ${(r.baselineShare*100).toFixed(1)}% base${r.excessPp >= 0 ? ' +' : ' '}${r.excessPp.toFixed(1)}pp)</span>`
    );
  }
  }catch(_e){ console.error("[buildAnalysis:speedBand]", _e); }

  // Missing lat/lon
  const el = document.getElementById("missingLatLon");
  el.innerHTML = `
    <div style="display:flex; justify-content:space-between; align-items:center; gap:10px;">
      <div>
        <div style="font-weight:900;">missing_latlon = ${missing}</div>
        <div class="hint">（同じフィルタ条件で lat/lon が無い行。地図に表示されない。）</div>
      </div>
      <div style="text-align:right;">
        <div style="font-weight:900;">shown = ${shown}</div>
        <div class="muted">matched = ${total}</div>
      </div>
    </div>
  `;
}

/* ===== Summary (simple) ===== */
function buildDangerRanking(filteredRows, key, outId, limit=10){
  const el = document.getElementById(outId);
  const g = new Map();
  for (const p of filteredRows){
    const name = p[key] || "Unknown";
    if (name === "Unknown") continue;
    let o = g.get(name);
    if (!o){ o = {name, n:0, sum:0}; g.set(name,o); }
    o.n += 1;
    o.sum += p.score;
  }
  const rows = Array.from(g.values())
    .filter(r=>r.n>=3)
    .map(r=>{
      const avg = r.sum/r.n;
      const metric = avg * Math.log(r.n + 1);
      return {name:r.name, n:r.n, avg, metric};
    })
    .sort((a,b)=> b.metric - a.metric)
    .slice(0, limit);

  if (rows.length===0){ el.innerHTML = `<div class="hint">該当なし</div>`; return; }
  const maxV = Math.max(...rows.map(r=>r.metric));
  el.innerHTML = rows.map(r=>{
    const w = maxV>0 ? Math.round((r.metric/maxV)*100) : 0;
    return `
      <div class="barRow" data-name="${encodeURIComponent(r.name)}" title="click to filter">
        <div style="min-width:0;">
          <div style="font-weight:900; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${r.name}</div>
          <div class="muted">avg=${fmt(r.avg,1)} / n=${r.n}</div>
        </div>
        <div class="miniBar"><div style="width:${w}%;"></div></div>
        <div style="text-align:right;"><b>${fmt(r.metric,1)}</b></div>
      </div>
    `;
  }).join("");

  el.querySelectorAll(".barRow").forEach(row=>{
    row.addEventListener("click", ()=>{
      const name = decodeURIComponent(row.getAttribute("data-name"));
      if (key==="道路") document.getElementById("roadSelect").value = name;
      if (key==="ロケーション") document.getElementById("locSelect").value = name;
      currentCellFilter = null;
      applyAll();
      const f2 = getFilters();
      const rows2 = filterRows(ALL_ROWS, f2);
      const pts2 = getDrawablePoints(rows2);
      zoomToPoints(pts2);
    });
  });
}

/* ===== Main apply ===== */
function applyAll(){
  const f = getFilters();

  const filteredRows = filterRows(ALL_ROWS, f);
  const shownPoints = getDrawablePoints(filteredRows);

  // Heat is always computed from the same grouping cells (after filters)
  const cells = updateHeatLayers(shownPoints, f.groupMeters);
  const cellCount = cells.length;

  // Points: zoom ≥ 15 => individual points, zoom ≤ 14 => grouped cell markers
  const showPoints = document.getElementById("togglePoints").classList.contains("active");
  if (!showPoints){
    pointsLayer.clearLayers();
    if (aggLayer) aggLayer.clearLayers();
    if (map.hasLayer(pointsLayer)) map.removeLayer(pointsLayer);
    if (aggLayer && map.hasLayer(aggLayer)) map.removeLayer(aggLayer);
  } else {
    if (f.zoom >= 15){
      if (!map.hasLayer(pointsLayer)) pointsLayer.addTo(map);
      if (aggLayer && map.hasLayer(aggLayer)) map.removeLayer(aggLayer);
      updatePointsLayer(shownPoints);
    } else {
      pointsLayer.clearLayers();
      if (map.hasLayer(pointsLayer)) map.removeLayer(pointsLayer);
      if (!aggLayer) aggLayer = createMarkerGroup("agg");
      if (!map.hasLayer(aggLayer)) aggLayer.addTo(map);
      updateAggLayer(cells, f.groupMeters);
    }
  }

  updateStatus(f, filteredRows.length, shownPoints, cellCount);

  // update analysis/summary if open
  if (document.getElementById("panelAnalysis")?.classList.contains("active")){
    // Wrap in try-catch: an Analysis panel error must NEVER abort applyAll
    // (map markers have already been added above, so the map is always rendered).
    try{
      buildAnalysis(filteredRows, shownPoints, f);
    }catch(err){
      const msg = err && err.message ? err.message : String(err || "buildAnalysis error");
      lastWindowError = msg;
      renderStatusText(null);
      const el = document.getElementById("analysisMeta");
      if (el) el.textContent = "Analysis error: " + msg;
    }
  }
  if (document.getElementById("panelSummary")?.classList.contains("active")){
    try{
      buildDangerRanking(shownPoints, "道路", "sumTopRoads", 10);
      buildDangerRanking(shownPoints, "ロケーション", "sumTopLocs", 10);
    }catch(_err){}
  }

  // heat opacity (always runs — must not be blocked by analysis errors above)
  setHeatOpacity(f.heatOpacity);
  document.getElementById("opVal").textContent = String(f.heatOpacity);
}


function resetAll(){
  currentCellFilter = null;

  // Core filters
  const _v = (id, val)=>{ const e = document.getElementById(id); if (e) e.value = val; };
  _v("roadSelect", "");
  _v("locSelect", "");
  _v("scoreMin", 0);
  _v("scoreMax", 45);
  document.querySelectorAll("#dxList input[type=checkbox]").forEach(cb => cb.checked = true);

  // DX search: clear text AND un-hide any items that were filtered out
  const dxSearch = document.getElementById("dxSearch");
  if (dxSearch) dxSearch.value = '';
  document.querySelectorAll("#dxList .dx-item").forEach(item => item.style.display = '');

  // Heat opacity
  _v("heatOpacity", "35");
  const opVal = document.getElementById("opVal"); if (opVal) opVal.textContent = "35";

  // Speed filters
  ['speedAvgMin','speedAvgMax','speedMaxMin','speedMaxMax','speedMinMin','speedMinMax'].forEach(id=>{
    const el = document.getElementById(id); if (el) el.value = '';
  });

  // Dismiss cell details card
  clearCellDetails();

  // grouping is auto-by-zoom, so nothing to reset there
  applyAll();
}


function openPanel(panelId){
  ["panelFilter","panelSummary","panelAnalysis","panelRoutes"].forEach(id=>{
    const el = document.getElementById(id);
    if (el) el.classList.toggle("active", id===panelId);
  });
  ["chipFilter","chipSummary","chipAnalysis","chipRoutes"].forEach(id=>{
    const el = document.getElementById(id);
    if (!el) return;
    const map = {"chipFilter":"panelFilter","chipSummary":"panelSummary","chipAnalysis":"panelAnalysis","chipRoutes":"panelRoutes"};
    el.classList.toggle("active", map[id]===panelId);
  });
  // refresh contents when opening
  const f = getFilters();
  const filteredRows = filterRows(ALL_ROWS, f);
  const shownPoints = getDrawablePoints(filteredRows);
  if (panelId==="panelAnalysis"){
    try{ buildAnalysis(filteredRows, shownPoints, f); }
    catch(err){
      const msg = err && err.message ? err.message : String(err || "buildAnalysis error");
      lastWindowError = msg;
      renderStatusText(null);
      const el = document.getElementById("analysisMeta");
      if (el) el.textContent = "Analysis error: " + msg;
    }
  }
  if (panelId==="panelSummary"){
    try{
      buildDangerRanking(shownPoints, "道路", "sumTopRoads", 10);
      buildDangerRanking(shownPoints, "ロケーション", "sumTopLocs", 10);
    }catch(_err){}
  }
}

function closePanels(){
  ["panelFilter","panelSummary","panelAnalysis","panelRoutes"].forEach(id=>{
    const el = document.getElementById(id);
    if (el) el.classList.remove("active");
  });
  ["chipFilter","chipSummary","chipAnalysis","chipRoutes"].forEach(id=>{
    const el = document.getElementById(id);
    if (el) el.classList.remove("active");
  });
}



/* ===== Routes (v67) ===== */
let routeMainLayer = null;
let routeSnippetLayer = null;
let routeMainKind = null; // 'car_week'
let routeMainKey = null;

function _emptyFC(){ return {type:"FeatureCollection", features:[]}; }


function buildRouteIndexWeekCarFromScenes(fc){
  const idx = {};
  if (!fc || !fc.features) return idx;
  for (const f of fc.features){
    const props = f.properties || {};
    const cw = props.car_week;
    const cid = props.car_id;
    if (!cw || !cid) continue;
    const coords = (f.geometry && f.geometry.coordinates) ? f.geometry.coordinates : null;
    if (!coords || coords.length < 2) continue;
    const key = cw + "__" + cid;
    if (!idx[key]) idx[key] = [];
    idx[key].push({ coords, props, source_file: props.source_file || "" });
  }
  // Keep a stable order for drawing (by source_file if present)
  for (const k of Object.keys(idx)){
    idx[k].sort((a,b)=> String(a.source_file||"").localeCompare(String(b.source_file||""), "ja"));
  }
  return idx;
}

function buildRouteIndex(fc, keyProp){
  const idx = {};
  const feats = (fc && fc.features) ? fc.features : [];
  for (const f of feats){
    const props = f.properties || {};
    const k = props[keyProp];
    if (!k) continue;
    const g = f.geometry || {};
    if (g.type !== 'LineString' || !Array.isArray(g.coordinates)) continue;
    idx[String(k)] = {coords: g.coordinates, props};
  }
  return idx;
}

function coordsToLatLngs(coords){
  const out = [];
  for (const c of (coords||[])){
    if (!c || c.length < 2) continue;
    const lon = Number(c[0]);
    const lat = Number(c[1]);
    if (!Number.isFinite(lat) || !Number.isFinite(lon)) continue;
    if (Math.abs(lat) > 90 || Math.abs(lon) > 180) continue;
    if (Math.abs(lat) < 1e-12 && Math.abs(lon) < 1e-12) continue; // drop exact (0,0)
    // Drop GPS sensor init errors: some routes have ~6% of points near null island
    // (values like lat=0.07, lon=-2.05) — clearly not Japan, causes fitBounds to zoom out globally
    if (Math.abs(lat) < 2.0 && Math.abs(lon) < 15.0) continue;
    out.push([lat, lon]);
  }
  return out; // [lon,lat] -> [lat,lon]
}

// Split a flat latlngs array into segments wherever consecutive points are more than
// maxGapDeg degrees apart (default 0.5° ≈ 55 km). This prevents straight lines from
// being drawn across GPS jumps between separate recording sessions in the same car_week.
function splitAtJumps(latlngs, maxGapDeg){
  const gap = (typeof maxGapDeg === "number" && maxGapDeg > 0) ? maxGapDeg : 0.5;
  const segs = [];
  let cur = [];
  for (const pt of latlngs){
    if (cur.length > 0){
      const prev = cur[cur.length - 1];
      const dLat = Math.abs(pt[0] - prev[0]);
      const dLon = Math.abs(pt[1] - prev[1]);
      if (dLat > gap || dLon > gap){
        if (cur.length >= 2) segs.push(cur);
        cur = [];
      }
    }
    cur.push(pt);
  }
  if (cur.length >= 2) segs.push(cur);
  // If nothing was long enough or all in one segment, return whole array
  if (segs.length === 0 && latlngs.length >= 2) segs.push(latlngs);
  return segs;
}

let ROUTE_INDEX_CW = null;
let ROUTE_INDEX_CAR = null;
let ROUTE_INDEX_SCENE = null;
let ROUTE_INDEX_CW_CAR = null; // key = `${car_week}__${car_id}` -> [{coords,source_file}]
let refreshRouteInsights = null;

function initRoutesUI(){
  const cwFC = (typeof ROUTES_CARWEEK !== 'undefined' && ROUTES_CARWEEK) ? ROUTES_CARWEEK : _emptyFC();
  const sceneFC = (typeof ROUTES_SCENE !== 'undefined' && ROUTES_SCENE) ? ROUTES_SCENE : _emptyFC();
  ROUTE_INDEX_CW = buildRouteIndex(cwFC, 'car_week');
  ROUTE_INDEX_CAR = null;
  ROUTE_INDEX_SCENE = buildRouteIndex(sceneFC, 'scene_key');
  ROUTE_INDEX_CW_CAR = null;

  const cwKeys = Object.keys(ROUTE_INDEX_CW).sort((a,b)=>a.localeCompare(b,'ja'));
  const selCW = document.getElementById('carweekRouteSelect');
  if (selCW) buildSelectOptions(selCW, cwKeys, true, 'Select...');

  const rgSel = document.getElementById('routeGroupSelect');
  const rgKmEl = document.getElementById('routeGroupKm');
  const summaryEl = document.getElementById('routeScopeSummary');
  const kmTotalsEl = document.getElementById('kmTotals');
  const btnShowCarweek = document.getElementById('btnShowCarweek');

  function fmtKm(v){ return (typeof v==='number' && Number.isFinite(v)) ? v.toFixed(2) : '-'; }
  function fmtHours(v){ return (typeof v==='number' && Number.isFinite(v)) ? v.toFixed(2) : '-'; }
  function fmtSteps(v){ return Number.isFinite(Number(v)) ? Number(v).toLocaleString('ja') : '-'; }
  function fmtPct(v){ return (typeof v==='number' && Number.isFinite(v)) ? v.toFixed(1) : '-'; }

  function routeCoverageStats(total){
    const observedKm = Number(total?.distance_km) || 0;
    const expectedKm = Math.max(observedKm, Number(total?.expected_distance_km) || 0);
    const missingKm = Math.max(0, expectedKm - observedKm);
    const frozenH = Number(total?.time_frozen_h) || 0;
    const gapH = Number(total?.time_gap_h) || 0;
    const durationH = Number(total?.duration_h) || 0;
    const unavailableH = Math.max(0, Number(total?.time_unavailable_h) || (frozenH + gapH));
    const unavailablePct = Number.isFinite(Number(total?.time_unavailable_pct))
      ? Number(total.time_unavailable_pct)
      : ((durationH > 0) ? (unavailableH / durationH * 100.0) : 0);
    const observedPct = expectedKm > 0 ? (observedKm / expectedKm * 100.0) : 0;
    const missingPct = expectedKm > 0 ? (missingKm / expectedKm * 100.0) : 0;
    return { observedKm, expectedKm, missingKm, observedPct, missingPct, unavailableH, unavailablePct };
  }

  function routeItemCoverageStats(row, scopeTotal){
    const observedKm = Number(row?.distance_km) || 0;
    let expectedKm = Math.max(observedKm, Number(row?.expected_distance_km) || 0);
    let imputedKm = Math.max(0, Number(row?.imputed_km) || 0);
    let inferred = false;
    if (!(expectedKm > observedKm + 0.01)){
      const scopeCov = routeCoverageStats(scopeTotal);
      if (scopeCov.expectedKm > scopeCov.observedKm + 0.01 && scopeCov.observedKm > 0.01 && observedKm > 0.01){
        const scale = scopeCov.expectedKm / scopeCov.observedKm;
        expectedKm = observedKm * scale;
        imputedKm = Math.max(0, expectedKm - observedKm);
        inferred = true;
      }
    }
    const missingKm = Math.max(0, expectedKm - observedKm);
    const missingPct = expectedKm > 0 ? (missingKm / expectedKm * 100.0) : 0;
    return { observedKm, expectedKm, imputedKm, missingKm, missingPct, inferred };
  }

  function getRouteScope(){
    const cw = selCW?.value || '';
    if (cw && ROUTE_SCOPE_TOTALS.by_carweek && ROUTE_SCOPE_TOTALS.by_carweek[cw]){
      return {kind:'car_week', key: cw, cw, label:`car_week ${cw}`};
    }
    return {kind:'overall', key:'__all__', cw, label:'overall'};
  }

  function getScopeTotals(scope){
    if (!scope) return ROUTE_SCOPE_TOTALS.overall || {};
    if (scope.kind === 'car_week'){
      return (ROUTE_SCOPE_TOTALS.by_carweek && ROUTE_SCOPE_TOTALS.by_carweek[scope.key]) || {};
    }
    return ROUTE_SCOPE_TOTALS.overall || {};
  }

  function getMetricRows(kind, scope){
    const bucket = (ROUTE_METRICS && ROUTE_METRICS[kind]) ? ROUTE_METRICS[kind] : {};
    if (!bucket) return [];
    if (scope && scope.kind === 'car_week' && bucket.by_carweek && bucket.by_carweek[scope.key]) return bucket.by_carweek[scope.key];
    return Array.isArray(bucket.overall) ? bucket.overall : [];
  }

  // isMissingRoadName is now at global scope (see "Shared road-name helpers" section above)

  function roadGroupForIncidentRow(row){
    const road = String(row?.["道路"] || '').trim();
    if (isMissingRoadName(road)) return '';
    const mapped = String(ROAD_GROUP_LOOKUP[road] || '').trim();
    return mapped || road;
  }

  function incidentRowsForScope(scope){
    return ALL_ROWS.filter((row)=>{
      if (scope?.kind === 'car_week'){
        return String(row?.car_week || '').trim() === scope.key;
      }
      return true;
    });
  }

  function buildScopeRateRows(scope, kind){
    const scopeTotal = getScopeTotals(scope);
    const metricKind = (kind === 'roadgroups') ? 'roadgroups' : 'labels';
    const metricRows = getMetricRows(metricKind, scope);
    const counts = new Map();
    const scoreSums = new Map();
    for (const row of incidentRowsForScope(scope)){
      const key = (kind === 'roadgroups') ? roadGroupForIncidentRow(row) : String(row?.["道路"] || '').trim();
      if (isMissingRoadName(key)) continue;
      counts.set(key, (counts.get(key) || 0) + 1);
      scoreSums.set(key, (scoreSums.get(key) || 0) + (Number(row?.score) || 0));
    }

    const out = [];
    for (const row of metricRows){
      const key = String(row?.name || '').trim();
      if (!key) continue;
      const count = Number(counts.get(key) || 0);
      if (count <= 0) continue;
      const cov = routeItemCoverageStats(row, scopeTotal);
      // Use observed km as primary denominator; fall back to expected only when observed is zero.
      const km = cov.observedKm > 0 ? cov.observedKm : cov.expectedKm;
      if (!(km > 0)) continue;
      const avgScore = count > 0 ? (Number(scoreSums.get(key) || 0) / count) : 0;
      const expText = (cov.expectedKm > cov.observedKm + 0.01) ? ` / exp ${fmtKm(cov.expectedKm)} km` : '';
      out.push({
        key,
        label: key,
        sub: [
          `${count} cases`,
          `obs ${fmtKm(km)} km${expText}`,
          `avg score ${fmt(avgScore, 1)}`,
          (Number(row?.segment_count) > 1) ? `${Number(row.segment_count)} segments` : '',
          cov.inferred ? 'scope-est.' : '',
        ].filter(Boolean).join(' / '),
        n: count / km,
        count,
        observedKm: km,
        expectedKm: cov.expectedKm,
        inferred: cov.inferred,
        avgScore,
      });
    }
    out.sort((a,b)=> (b.n - a.n) || (b.count - a.count) || String(a.label).localeCompare(String(b.label), 'ja'));
    return out; // all rows — container has max-height + overflow-y:auto
  }

  function getDisplayedRouteGeo(scope){
    if (!scope) return null;
    if (scope.kind === 'car_week' && ROUTE_INDEX_CW && ROUTE_INDEX_CW[scope.key]){
      return ROUTE_INDEX_CW[scope.key].props || null;
    }
    return null;
  }

  function updateRouteGroupDetail(scope){
    if (!rgSel || !rgKmEl) return;
    const rows = getMetricRows('roadgroups', scope);
    const keys = rows.map(r => r.name).filter(Boolean);
    const current = rgSel.value || '';
    buildSelectOptions(rgSel, keys, true, 'All');
    if (current && keys.includes(current)) rgSel.value = current;
    const selected = rgSel.value || '';
    if (!selected){
      const total = getScopeTotals(scope);
      const cov = routeCoverageStats(total);
      const expectedText = (cov.expectedKm > cov.observedKm + 0.01)
        ? ` <span class="muted">(exp ${fmtKm(cov.expectedKm)} km)</span>`
        : '';
      rgKmEl.innerHTML = `${scope.label}: <b>${fmtKm(cov.observedKm)}</b> km${expectedText} / <b>${fmtHours(Number(total.duration_h))}</b> h / ${fmtSteps(total.steps)} steps`;
      return;
    }
    const hit = rows.find(r => r.name === selected);
    if (!hit){
      rgKmEl.innerHTML = '-';
      return;
    }
    const cov = routeItemCoverageStats(hit, getScopeTotals(scope));
    const segText = (Number(hit.segment_count) > 1) ? ` / ${Number(hit.segment_count)} segments` : '';
    const expectedText = (cov.expectedKm > cov.observedKm + 0.01)
      ? ` <span class="muted">(exp ${fmtKm(cov.expectedKm)} km${cov.inferred ? ' / scope-est.' : ''})</span>`
      : '';
    rgKmEl.innerHTML = `selected: <b>${fmtKm(cov.observedKm)}</b> km${expectedText} / <b>${fmtHours(Number(hit.duration_h))}</b> h / ${fmtSteps(hit.steps)} steps${segText}`;
  }

  function renderMetricBars(containerId, rows, onClick, scopeTotal){
    const mapped = (rows || []).map(r => {
      const cov = routeItemCoverageStats(r, scopeTotal || {});
      return {
        coverage: cov,
        key: r.name,
        label: r.name,
        sub: [
          (Number(r.segment_count) > 1) ? `${Number(r.segment_count)} segments` : '',
          `${fmtHours(Number(r.duration_h))} h`,
          `${fmtSteps(r.steps)} steps`,
          (cov.expectedKm > cov.observedKm + 0.01) ? `(exp ${fmtKm(cov.expectedKm)} km${cov.inferred ? ' / scope-est.' : ''})` : '',
        ].filter(Boolean).join(' / '),
        n: cov.observedKm || Number(r.distance_km) || cov.expectedKm || 0,
        distance_km: Number(r.distance_km) || 0,
        duration_h: Number(r.duration_h) || 0,
        steps: Number(r.steps) || 0,
        segment_count: Number(r.segment_count) || 0,
      };
    }).filter(r => r.n > 0).sort((a,b)=> (b.n - a.n) || String(a.label).localeCompare(String(b.label), 'ja'));
    buildBarList(
      containerId,
      mapped,
      onClick || (()=>{}),
      (r)=> {
        const cov = r.coverage || routeItemCoverageStats(r, scopeTotal || {});
        if (cov.expectedKm > cov.observedKm + 0.01){
          return `<b>${fmtKm(cov.observedKm)}</b> <span class="muted">km</span><div class="muted">(exp ${fmtKm(cov.expectedKm)} km${cov.inferred ? ' / scope-est.' : ''})</div>`;
        }
        return `<b>${fmtKm(cov.observedKm)}</b> <span class="muted">km</span>`;
      }
    );
  }

  function renderScopeRateBars(containerId, rows, onClick){
    buildBarList(
      containerId,
      rows || [],
      onClick || (()=>{}),
      (r)=> `<b>${r.n.toFixed(3)}</b> <span class="muted">cases/km</span><div class="muted">count=${r.count}, obs=${fmtKm(r.observedKm)}km${r.expectedKm > r.observedKm + 0.01 ? ' / exp ' + fmtKm(r.expectedKm) + 'km' : ''}${r.inferred ? ' / scope-est.' : ''}</div>`
    );
  }

  function renderRouteInsights(){
    const scope = getRouteScope();
    const totals = getScopeTotals(scope);
    const overall = ROUTE_SCOPE_TOTALS.overall || {};
    const geo = getDisplayedRouteGeo(scope);
    const hasRouteGeometry = cwKeys.length > 0;

    if (kmTotalsEl){
      const lines = [];
      const overallCov = routeCoverageStats(overall);
      const overallExpectedText = (overallCov.expectedKm > overallCov.observedKm + 0.01)
        ? ` <span class="muted">(exp ${fmtKm(overallCov.expectedKm)} km)</span>`
        : '';
      lines.push(`Overall observed / usable (all routes): <b>${fmtKm(overallCov.observedKm)}</b> km${overallExpectedText} / <b>${fmtHours(Number(overall.duration_h))}</b> h`);
      if (overallCov.expectedKm > overallCov.observedKm + 0.01){
        lines.push(`Estimated HDD unavailable / invalid / inaccessible: <b>${fmtPct(overallCov.unavailablePct)}%</b> of time / <b>${fmtPct(overallCov.missingPct)}%</b> of distance (${fmtKm(overallCov.missingKm)} km compensated)`);
      }
      if (scope.kind !== 'overall'){
        const scopeCov = routeCoverageStats(totals);
        const scopeExpectedText = (scopeCov.expectedKm > scopeCov.observedKm + 0.01)
          ? ` <span class="muted">(exp ${fmtKm(scopeCov.expectedKm)} km / ${fmtPct(scopeCov.missingPct)}% compensated)</span>`
          : '';
        lines.push(`${scope.label}: <b>${fmtKm(scopeCov.observedKm)}</b> km${scopeExpectedText} / <b>${fmtHours(Number(totals.duration_h))}</b> h / ${fmtSteps(totals.steps)} steps`);
      }
      kmTotalsEl.innerHTML = lines.join('<br/>');
    }

    if (summaryEl){
      const lines = [];
      const imputedKm = Number(totals.imputed_km) || 0;
      const cov = routeCoverageStats(totals);
      const expectedKm = cov.expectedKm;
      const routeKm = Number(geo?.route_geo_km);
      lines.push(`scope: <b>${scope.label}</b>`);
      const expectedText = (expectedKm > cov.observedKm + 0.01)
        ? ` <span class="muted">(exp ${fmtKm(expectedKm)} km)</span>`
        : '';
      lines.push(`observed / usable exposure: <b>${fmtKm(cov.observedKm)}</b> km${expectedText} / <b>${fmtHours(Number(totals.duration_h))}</b> h / ${fmtSteps(totals.steps)} steps`);
      if (imputedKm > 0.01 || Number(totals.steps_imputed) > 0){
        lines.push(`estimated / imputed: <b>${fmtKm(imputedKm)}</b> km / ${fmtSteps(totals.steps_imputed)} steps (${fmtPct(Number(totals.imputed_step_pct))}%)`);
      }
      if (expectedKm > cov.observedKm + 0.01){
        lines.push(`estimated unavailable / invalid / inaccessible HDD share: <b>${fmtPct(cov.unavailablePct)}%</b> of time (${fmtHours(cov.unavailableH)} h) / <b>${fmtPct(cov.missingPct)}%</b> of distance`);
      }
      if (geo && (Number.isFinite(Number(geo.route_geo_km)) || Number.isFinite(Number(geo.duration_h)))){
        const sceneText = Number.isFinite(Number(geo.scene_count)) ? ` / ${fmtSteps(geo.scene_count)} scenes` : '';
        lines.push(`route geometry: <b>${fmtKm(Number(geo.route_geo_km))}</b> km / <b>${fmtHours(Number(geo.duration_h))}</b> h${sceneText}`);
        if (Number.isFinite(routeKm) && cov.observedKm > 0.01){
          const gapKm = Math.max(0, cov.observedKm - routeKm);
          const coveragePct = Math.min(100, Math.max(0, (routeKm / cov.observedKm) * 100));
          lines.push(`route-line coverage: <b>${fmtPct(coveragePct)}%</b> (${fmtKm(routeKm)} / ${fmtKm(cov.observedKm)} km, missing on line ${fmtKm(gapKm)} km)`);
        }
      } else if (!hasRouteGeometry){
        lines.push(`route geometry: <span class="muted">routes_by_carweek.geojson is missing or empty, so only exposure totals are shown.</span>`);
      }
      if (activeRouteFilter?.car_week){
        lines.push(`map filter: showing incidents for <b>${activeRouteFilter.car_week}</b> while the route is active`);
      }
      summaryEl.innerHTML = lines.join('<br/>');
    }

    const labelRows = getMetricRows('labels', scope);
    renderMetricBars('routeSegmentList', labelRows, (roadName)=>{
      const sel = document.getElementById('roadSelect');
      if (sel) sel.value = roadName || '';
      currentCellFilter = null;
      applyAll();
      openPanel('panelFilter');
    }, totals);

    const groupRows = getMetricRows('roadgroups', scope);
    renderMetricBars('routeRoadGroupList', groupRows, (groupName)=>{
      if (rgSel){
        rgSel.value = groupName || '';
        updateRouteGroupDetail(scope);
      }
    }, totals);

    const segRateRows = buildScopeRateRows(scope, 'segments');
    renderScopeRateBars('routeSegmentRateList', segRateRows, (roadName)=>{
      const sel = document.getElementById('roadSelect');
      if (sel) sel.value = roadName || '';
      currentCellFilter = null;
      applyAll();
      openPanel('panelFilter');
    });

    const groupRateRows = buildScopeRateRows(scope, 'roadgroups');
    renderScopeRateBars('routeRoadGroupRateList', groupRateRows, (groupName)=>{
      if (rgSel){
        rgSel.value = groupName || '';
        updateRouteGroupDetail(scope);
      }
    });

    updateRouteGroupDetail(scope);
  }

  refreshRouteInsights = renderRouteInsights;
  selCW?.addEventListener('change', renderRouteInsights);
  rgSel?.addEventListener('change', ()=> updateRouteGroupDetail(getRouteScope()));
  if (btnShowCarweek){
    btnShowCarweek.disabled = (cwKeys.length === 0);
    btnShowCarweek.title = (cwKeys.length === 0) ? 'routes_by_carweek.geojson is missing or empty.' : '';
  }
  renderRouteInsights();

  btnShowCarweek?.addEventListener('click', ()=>{
    const k = selCW?.value || '';
    if (!k) return;
    showRoute('car_week', k, true);
  });

  document.getElementById('btnClearRoutes')?.addEventListener('click', ()=>{ clearRoutes(); });
}

function clearRoutes(){
  if (routeMainLayer) { map.removeLayer(routeMainLayer); routeMainLayer = null; }
  if (routeSnippetLayer) { map.removeLayer(routeSnippetLayer); routeSnippetLayer = null; }
  activeRouteFilter = null;
  routeMainKind = null; routeMainKey = null;
  applyAll();
  if (typeof refreshRouteInsights === "function") refreshRouteInsights();
}

function getRouteCoords(kind, key){
  if (kind === 'car_week') return ROUTE_INDEX_CW && ROUTE_INDEX_CW[key] ? ROUTE_INDEX_CW[key].coords : null;
  if (kind === 'car_id') return ROUTE_INDEX_CAR && ROUTE_INDEX_CAR[key] ? ROUTE_INDEX_CAR[key].coords : null;
  return null;
}


function showRouteWeekCar(carWeek, carId, fit=true){
  const key = String(carWeek||"") + "__" + String(carId||"");
  const segs = (ROUTE_INDEX_CW_CAR && ROUTE_INDEX_CW_CAR[key]) ? ROUTE_INDEX_CW_CAR[key] : null;
  if (!segs || segs.length === 0){
    alert(`No per-scene segments found for car_id=${carId} in car_week=${carWeek}.`);
    return;
  }
  // Remove existing main route
  if (routeMainLayer) { map.removeLayer(routeMainLayer); routeMainLayer = null; }

  const polys = [];
  for (const s of segs){
    const latlngs = coordsToLatLngs(s.coords);
    if (latlngs.length < 2) continue;
    polys.push(L.polyline(latlngs, {weight: 4, opacity: 0.85}));
  }
  if (polys.length === 0){
    alert(`Segments exist but are empty after filtering invalid points for car_id=${carId} in car_week=${carWeek}.`);
    return;
  }
  routeMainLayer = L.featureGroup(polys);
  routeMainLayer.addTo(map);
  routeMainKind = "car_week+car_id";
  routeMainKey = key;

  if (fit){
    map.fitBounds(routeMainLayer.getBounds(), {padding:[40,40]});
  }
}

function showRoute(kind, key, fit){
  const coords = getRouteCoords(kind, key);
  if (!coords || !coords.length){
    alert('Route not found. Run pipeline route export (routes_by_*.geojson).');
    return;
  }
  const latlngs = coordsToLatLngs(coords);
  if (routeMainLayer) map.removeLayer(routeMainLayer);
  // Split at large GPS jumps (> 0.5° ≈ 55 km) to avoid straight lines between sessions.
  const _routeSegs = splitAtJumps(latlngs);
  if (_routeSegs.length === 1){
    routeMainLayer = L.polyline(_routeSegs[0], {weight: 4, opacity: 0.85});
  } else {
    routeMainLayer = L.featureGroup(_routeSegs.map(s => L.polyline(s, {weight: 4, opacity: 0.85})));
  }
  routeMainLayer.addTo(map);
  activeRouteFilter = (kind === 'car_week') ? {car_week: String(key || '').trim()} : null;
  routeMainKind = kind; routeMainKey = key;
  applyAll();
  if (typeof refreshRouteInsights === "function") refreshRouteInsights();

  if (fit){
    try{ map.fitBounds(routeMainLayer.getBounds().pad(0.15)); }catch(e){}
  }
}

function nearestIndex(latlngs, lat, lon){
  let bestI = 0;
  let bestD = Infinity;
  for (let i=0;i<latlngs.length;i++){
    const p = latlngs[i];
    const d = (p[0]-lat)*(p[0]-lat) + (p[1]-lon)*(p[1]-lon);
    if (d < bestD){ bestD = d; bestI = i; }
  }
  return bestI;
}

function extractSceneKeyFromStorage(s){
  if (!s) return null;
  const m = String(s).match(/(\d{2}_\d{6}-\d{6}_Scene\d+)/i);
  if (m) return m[1];
  return null;
}

function showRouteSnippet(rowIndex){
  const p = ALL_ROWS[rowIndex];
  if (!p || !isValidLatLon(Number(p.lat), Number(p.lon))) return;

  // Prefer per-incident "scene route" (start→finish of that file)
  const sk = extractSceneKeyFromStorage(p['格納先']);
  if (sk && ROUTE_INDEX_SCENE && ROUTE_INDEX_SCENE[sk]){
    const coords = ROUTE_INDEX_SCENE[sk].coords;
    const latlngs = coordsToLatLngs(coords);
    if (latlngs.length < 2){
      alert('Scene route is too short.');
      return;
    }
    if (routeSnippetLayer) map.removeLayer(routeSnippetLayer);
    routeSnippetLayer = L.polyline(latlngs, {weight: 6, opacity: 0.95});
    routeSnippetLayer.addTo(map);
    map.fitBounds(routeSnippetLayer.getBounds(), {padding:[40,40]});
    return;
  }

  // Fallback: highlight a short segment on the selected main route (car_week/car_id)
  let kind = routeMainKind;
  let key = routeMainKey;
  let coords = (kind && key) ? getRouteCoords(kind, key) : null;

  if (!coords){
    if (p.car_week && ROUTE_INDEX_CW && ROUTE_INDEX_CW[p.car_week]){ kind = 'car_week'; key = p.car_week; coords = ROUTE_INDEX_CW[p.car_week].coords; }
  }

  if (!coords){
    alert('No matching route for this point (scene/car_week).');
    return;
  }

  // Ensure main route is shown for context
  if (!routeMainLayer || routeMainKind !== kind || routeMainKey !== key){
    showRoute(kind, key, false);
  }

  const latlngs = coordsToLatLngs(coords);
  const i0 = nearestIndex(latlngs, p.lat, p.lon);
  const w = 25;
  const s = Math.max(0, i0 - w);
  const e = Math.min(latlngs.length, i0 + w);
  const seg = latlngs.slice(s, e);

  if (routeSnippetLayer) map.removeLayer(routeSnippetLayer);
  routeSnippetLayer = L.polyline(seg, {weight: 6, opacity: 0.95});
  routeSnippetLayer.addTo(map);
  map.fitBounds(routeSnippetLayer.getBounds(), {padding:[40,40]});
}


function initUI(){
  // meta
  const meta = DATA.meta || {};
  const jpSheet = meta.jp_sheet_used || "集計表";
  const legacySheet = meta.legacy_sheet_used || "AllPoints_1018";
  const jpCount = Number(meta.jp_row_count) || 0;
  const legacyCount = Number(meta.legacy_row_count) || 0;
  const totalCount = Number(meta.total_rows) || ALL_ROWS.length;
  const metaLine = `Excelから再現（JP:${jpSheet} ${jpCount}件 / Legacy:${legacySheet} ${legacyCount}件 / Total ${totalCount}件, 緯度経度あり ${meta.mappable_points_with_latlon||"-"}, なし ${meta.missing_latlon_rows||"-"}）`;
  document.getElementById("metaLine").textContent = metaLine;
  document.getElementById("metaHint").textContent =
    `score: low=20 / medium=30 / high=40 / crash=45 / unknown=0　Heatはフィルタ後データで再計算（grid集約）。`;

  // build selects
  const roads = uniqueSorted(ALL_ROWS.map(p=>p["道路"]).filter(Boolean));
  const locs = uniqueSorted(ALL_ROWS.map(p=>p["ロケーション"]).filter(Boolean));
  buildSelectOptions(document.getElementById("roadSelect"), roads, true, "全て");
  buildSelectOptions(document.getElementById("locSelect"), locs, true, "全て");

  // dx options
  const dxNorm = (DX_OPTIONS || []).map(normalizeDxOption);
  buildDxList(dxNorm);

  // dx search
  const dxSearch = document.getElementById("dxSearch");
  dxSearch.addEventListener("input", ()=>{
    const q = (dxSearch.value||"").toLowerCase().trim();
    document.querySelectorAll("#dxList .dx-item").forEach(item=>{
      const txt = item.textContent.toLowerCase();
      item.style.display = (!q || txt.includes(q)) ? "" : "none";
    });
  });

  document.getElementById("dxAll").addEventListener("click", ()=>{ document.querySelectorAll("#dxList input[type=checkbox]").forEach(cb=>cb.checked=true); });
  document.getElementById("dxNone").addEventListener("click", ()=>{ document.querySelectorAll("#dxList input[type=checkbox]").forEach(cb=>cb.checked=false); });

  // apply/reset
  document.getElementById("applyNow").addEventListener("click", ()=>{ applyAll(); });
  document.getElementById("resetAll").addEventListener("click", ()=>{ resetAll(); });

  // heat opacity live
  document.getElementById("heatOpacity").addEventListener("input", ()=>{
    const v = Number(document.getElementById("heatOpacity").value||35);
    document.getElementById("opVal").textContent = String(v);
    setHeatOpacity(v);
  });

  // layer toggles
  toggleLayer("toggleHeatAvg", heatAvgLayer);
  toggleLayer("toggleHeatFreq", heatFreqLayer);
  toggleLayer("togglePoints", null);

  // chips
  document.getElementById("chipFilter").addEventListener("click", ()=> openPanel("panelFilter"));
  document.getElementById("chipSummary").addEventListener("click", ()=> openPanel("panelSummary"));
  document.getElementById("chipAnalysis").addEventListener("click", ()=> openPanel("panelAnalysis"));
  document.getElementById("chipRoutes").addEventListener("click", ()=> openPanel("panelRoutes"));

  // close buttons
  document.getElementById("closeFilter").addEventListener("click", closePanels);
  document.getElementById("closeSummary").addEventListener("click", closePanels);
  document.getElementById("closeAnalysis").addEventListener("click", closePanels);
  document.getElementById("closeRoutes").addEventListener("click", closePanels);

  // analysis-only controls (rerender without Apply)
  const rerenderAnalysis = ()=>{
    if (!document.getElementById("panelAnalysis")?.classList.contains("active")) return;
    const f = getFilters();
    const filteredRows = filterRows(ALL_ROWS, f);
    const shownPoints = getDrawablePoints(filteredRows);
    try{ buildAnalysis(filteredRows, shownPoints, f); }catch(_err){}
  };
  document.getElementById("analysisTopN")?.addEventListener("input", rerenderAnalysis);
  document.getElementById("analysisTopN")?.addEventListener("change", rerenderAnalysis);
  document.getElementById("hideUnknownRoads")?.addEventListener("change", rerenderAnalysis);

  document.getElementById("analysisReset")?.addEventListener("click", ()=>{
    resetAll();
    openPanel("panelAnalysis");
  });

  // routes UI
  initRoutesUI();

  // initial open filter
  openPanel("panelFilter");
}

window.addEventListener("load", ()=>{
  try{
    initMap();
    initUI();
    applyAll();
  }catch(err){
    lastWindowError = err && err.message ? err.message : String(err || "load error");
    renderStatusText({ baseText: "表示中: 初期化失敗", shownPointsLen: 0, pointsLayerCount: layerCount(pointsLayer), aggLayerCount: layerCount(aggLayer) });
    throw err;
  }
});
</script>
</body>
</html>
"""


def build_rows(df: pd.DataFrame):
    """
    Build rows for the HTML:
    - Always include ALL rows (including missing lat/lon) so analysis can show missing counts.
    - Fill DX (分類 / 分類_1) using fallbacks (dx_code / dx_name) so "Unknown" does NOT dominate just because lat/lon is missing.
    - Fill road label (道路) using fallbacks; for unnamed snapped roads, use a type label instead of plain "Unknown".
    """

def normalize_row_payload(rec: Dict[str, Any], dx_name_map: Dict[str, str], dx_name_to_code: Dict[str, str]) -> Dict[str, Any]:
    normalize_row_coordinates(rec)

    raw_code = _s(rec.get("dx_main")) or _s(rec.get("dx_code"))
    raw_label = _s(rec.get("分類名")) or _s(rec.get("dx_name"))
    current_dx = _s(rec.get("分類"))

    if not raw_code and _looks_like_dx_code(current_dx):
        raw_code = current_dx
    if not raw_code and current_dx and current_dx in dx_name_to_code:
        raw_code = dx_name_to_code[current_dx]
    if not raw_label and current_dx and current_dx not in {"Unknown", "不明"} and not _looks_like_dx_code(current_dx):
        raw_label = current_dx
    if raw_code and not raw_label:
        raw_label = dx_name_map.get(raw_code) or _s(rec.get("dx_name"))
    if not raw_code and raw_label and raw_label in dx_name_to_code:
        raw_code = dx_name_to_code[raw_label]

    if not raw_code:
        raw_code = "Unknown"
    if not raw_label:
        raw_label = "不明" if raw_code == "Unknown" else (dx_name_map.get(raw_code) or raw_code)

    rec["分類"] = raw_code
    rec["分類名"] = raw_label

    sub_label = _s(rec.get("分類_1")) or _s(rec.get("分類.1"))
    dx_subs = rec.get("dx_subs")
    if not sub_label and isinstance(dx_subs, list) and dx_subs:
        sub_label = " / ".join(dx_name_map.get(dx_code, dx_code) for dx_code in dx_subs if _s(dx_code))
    rec["分類_1"] = sub_label

    score = _safe_float(rec.get("score"))
    if score is None:
        lvl = _s(rec.get("ニアミスレベル")) or _s(rec.get("risk_level"))
        score = 40.0 if lvl == "大" else 30.0 if lvl == "中" else 20.0 if lvl == "小" else 0.0
    rec["score"] = float(score)

    for key in ["平均速度", "最低速度", "最高速度", "snap_dist_m"]:
        val = _safe_float(rec.get(key))
        rec[key] = float(val) if val is not None else None

    rec["道路"] = _s(_first_present(rec, ROAD_VALUE_KEYS)) or "不明"
    rec["ロケーション"] = _s(_first_present(rec, LOC_VALUE_KEYS)) or "不明"
    return rec


# --------------------------- input path resolution ---------------------------

def _resolve_excel_path(p: str, prefer_keywords: List[str]) -> str:
    """
    Resolve Excel file path robustly.
    - If p exists, return it.
    - Else scan current directory for .xlsx and choose best match by keywords.
    This avoids mojibake / codepage issues when running from .bat on Windows.
    """
    cand = Path(p)
    if cand.exists():
        return str(cand)
    # scan current directory
    xlsx = sorted(Path(".").glob("*.xlsx"))
    if not xlsx:
        return p  # will error later with clear message
    # scoring by keyword hits
    def score(path: Path) -> int:
        name = path.name
        s = 0
        for kw in prefer_keywords:
            if kw and kw in name:
                s += 10
        return s
    ranked = sorted(xlsx, key=lambda pp: (score(pp), pp.stat().st_mtime), reverse=True)
    best = ranked[0]
    return str(best)

def _ensure_style_braces_ok(html: str) -> str:
    """
    Fix accidental double-braces '{{' '}}' inside <style> blocks (template issue).
    This prevents map div from collapsing due to invalid CSS.
    """
    def fix_block(m: re.Match) -> str:
        block = m.group(0)
        # Only replace inside the block
        inner = m.group(1)
        inner = inner.replace("{{", "{").replace("}}", "}")
        return "<style>" + inner + "</style>"
    return re.sub(r"<style>(.*?)</style>", fix_block, html, flags=re.DOTALL | re.IGNORECASE)

# --------------------------- main ---------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--jp-excel", default="ニアミス抽出_進捗管理表.xlsx")
    ap.add_argument("--jp-sheet", default="集計表")
    ap.add_argument("--legacy-excel", default="route_nearmiss_analysis_with_dx_and_unknown_GROUPED_v12_all1018.xlsx")
    ap.add_argument("--legacy-sheet", default="AllPoints_1018")
    ap.add_argument("--logs-root", default=".", help="Root folder containing car_week folders with logs")
    ap.add_argument("--logs-glob", default="**/*.csv", help="Glob under logs-root to find log CSVs")
    ap.add_argument("--logs-manifest", default=None, help="Optional text file listing log CSV paths, one per line")
    ap.add_argument("--tolerance-s", type=float, default=12.0)
    ap.add_argument("--cache-dir", default="pipeline_out/_cache")
    ap.add_argument("--osm-label-cache", default="osm_label_cache.json")
    ap.add_argument("--osm-reverse-cache", default="osm_reverse_cache.json")
    ap.add_argument("--osm-overpass-cache", default="osm_overpass_cache.json")
    ap.add_argument("--osm-fetch-missing", action="store_true")
    ap.add_argument("--osm-max-requests", type=int, default=500)
    ap.add_argument("--osm-sleep", type=float, default=1.0)
    ap.add_argument("--osm-overpass-radius-m", type=float, default=40.0)
    ap.add_argument("--osm-overpass-timeout-s", type=float, default=20.0)
    ap.add_argument("--routes-by-carweek", default=None, help="Optional routes_by_carweek.geojson (from route export)")
    ap.add_argument("--routes-by-car", default=None, help="Optional routes_by_car.geojson (from route export)")
    ap.add_argument("--routes-by-scene", default=None, help="Optional routes_by_scene.geojson (from route export)")
    ap.add_argument("--routes-dir", default="pipeline_out/routes_out", help="Auto-detect routes_by_*.geojson under this dir")
    ap.add_argument("--out", default="pipeline_out/nearmiss_points_map_v67_richui_both_excels.html")
    args = ap.parse_args()

    jp_excel = Path(_resolve_excel_path(args.jp_excel, ["ニアミス", "進捗", "抽出", "管理表"]))
    legacy_excel = Path(_resolve_excel_path(args.legacy_excel, ["all1018", "1018", "GROUPED", "route_nearmiss"]))
    logs_root = Path(args.logs_root)
    cache_dir = Path(args.cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)

    # Load OSM caches
    osm_cache = load_osm_label_cache(Path(args.osm_label_cache))
    rev_cache_path = Path(args.osm_reverse_cache)
    rev_cache = load_osm_reverse_cache(rev_cache_path)
    overpass_cache_path = Path(args.osm_overpass_cache)
    overpass_cache = load_json_object_cache(overpass_cache_path)

    # Parse JP events
    jp_rows, dx_name_map, jp_sheet_used, jp_parse_stats = parse_jp_progress_clean(jp_excel, args.jp_sheet)
    dx_name_to_code = {name: code for code, name in dx_name_map.items() if name}

    # Load legacy
    legacy_df, legacy_sheet_used = load_legacy_1018(legacy_excel, args.legacy_sheet)

    # Build map from legacy for lookup by (car_week, No)
    legacy_lookup: Dict[Tuple[str, int], Dict[str, Any]] = {}
    for _, r in legacy_df.iterrows():
        cw_raw = _s(r.get("car_week"))
        cw = normalize_carweek_token(cw_raw) or cw_raw
        no = _safe_int(r.get("No"))
        if cw and no is not None:
            legacy_lookup[(cw, no)] = r.to_dict()
        if cw_raw and cw_raw != cw and no is not None:
            legacy_lookup[(cw_raw, no)] = r.to_dict()

    # Build log file index (for jp events)
    index_cache = cache_dir / "log_files_by_carweek.json"
    logs_manifest = Path(args.logs_manifest) if args.logs_manifest else None
    log_index = build_carweek_file_index(logs_root, args.logs_glob, index_cache, logs_manifest)

    # Load latlon cache (same key style as v67)
    latlon_cache_path = cache_dir / "latlon_cache.json"
    latlon_cache: Dict[str, Any] = {}
    if latlon_cache_path.exists():
        try:
            latlon_cache = json.loads(latlon_cache_path.read_text(encoding="utf-8"))
        except Exception:
            latlon_cache = {}
    latlon_cache_by_video_sec = build_video_sec_latlon_cache_index(latlon_cache)

    def cache_key(car_week: str, video: str, sec: Optional[float]) -> str:
        sec_s = f"{sec:.1f}" if sec is not None else ""
        return f"{car_week}|{video}|{sec_s}"

    # Enrich JP rows with lat/lon
    osm_requests = 0
    for ev in jp_rows:
        ev["has_latlon"] = False
        ev["lat"] = None
        ev["lon"] = None

        cw = normalize_carweek_token(_s(ev.get("car_week"))) or _s(ev.get("car_week"))
        ev["car_week"] = cw
        video = _s(ev.get("video"))
        sec = ev.get("sec")
        no_int = _safe_int(ev.get("No"))

        # Attempt: legacy lookup by (car_week, No)
        legacy_hit = None
        if cw and no_int is not None and (cw, no_int) in legacy_lookup:
            legacy_hit = legacy_lookup[(cw, no_int)]

        # Attempt: cached latlon by key
        if video and isinstance(sec, (int,float)):
            if cw:
                k = cache_key(cw, video, float(sec))
                if k in latlon_cache:
                    got = latlon_cache[k]
                    ev["lat"] = got.get("lat"); ev["lon"] = got.get("lon")
                    ev["speed"] = got.get("speed")
                    normalize_row_coordinates(ev)
            if not ev["has_latlon"]:
                got = latlon_cache_by_video_sec.get(_video_sec_cache_key(video, float(sec)))
                if got:
                    ev["lat"] = got.get("lat"); ev["lon"] = got.get("lon")
                    ev["speed"] = got.get("speed")
                    normalize_row_coordinates(ev)

        # Attempt: log search
        if not ev["has_latlon"] and video and isinstance(sec, (int,float)):
            ts0 = parse_timestamp_from_video_name(video)
            if ts0 is not None:
                target_ts = ts0 + datetime.timedelta(seconds=float(sec))
                files = candidate_log_files_for_event(log_index, cw, _s(ev.get("car_id")), target_ts)
                got = find_latlon_in_logfiles(files, target_ts, args.tolerance_s)
                if got:
                    ev["lat"] = got["lat"]; ev["lon"] = got["lon"]; ev["speed"] = got.get("speed")
                    ev["has_latlon"] = True
                    cache_k = cache_key(cw, video, float(sec))
                    latlon_cache[cache_k] = got
                    latlon_cache_by_video_sec[_video_sec_cache_key(video, float(sec))] = got

        # Attempt: legacy fallback for coordinates + labels
        if not ev["has_latlon"] and legacy_hit is not None:
            lat = _pick_float(legacy_hit, LAT_VALUE_KEYS)
            lon = _pick_float(legacy_hit, LON_VALUE_KEYS)
            if _is_valid_coord(lat, lon):
                ev["lat"] = lat; ev["lon"] = lon
                ev["has_latlon"] = True
        merge_legacy_labels_into_event(ev, legacy_hit)

        # Fill speeds
        spd = ev.get("speed")
        if spd is not None:
            ev.setdefault("平均速度", spd)
            ev.setdefault("最低速度", spd)
            ev.setdefault("最高速度", spd)

        normalize_row_payload(ev, dx_name_map, dx_name_to_code)
        osm_requests = enrich_row_location_labels(
            ev,
            osm_cache,
            rev_cache,
            False,
            args.osm_max_requests,
            osm_requests,
            args.osm_sleep,
            overpass_cache=overpass_cache,
            overpass_radius_m=args.osm_overpass_radius_m,
            overpass_timeout_s=args.osm_overpass_timeout_s,
        )

        # Store 格納先 as folder/video for convenience
        if not _s(ev.get("格納先")):
            ev["格納先"] = f"{cw}/{video}" if cw or video else ""

    # Convert legacy rows to row dicts, and append (as separate source)
    legacy_rows: List[Dict[str, Any]] = []
    for _, r in legacy_df.iterrows():
        d = r.to_dict()
        d["_src"] = "legacy1018"
        cw_norm = normalize_carweek_token(_s(d.get("car_week")))
        if cw_norm:
            d["car_week"] = cw_norm
        normalize_row_payload(d, dx_name_map, dx_name_to_code)
        d["has_latlon"] = _coerce_bool(d.get("has_latlon")) or d.get("has_latlon") is True
        normalize_row_coordinates(d)
        osm_requests = enrich_row_location_labels(
            d,
            osm_cache,
            rev_cache,
            False,
            args.osm_max_requests,
            osm_requests,
            args.osm_sleep,
            overpass_cache=overpass_cache,
            overpass_radius_m=args.osm_overpass_radius_m,
            overpass_timeout_s=args.osm_overpass_timeout_s,
        )
        legacy_rows.append(d)

    # Combine ALL rows
    rows_all: List[Dict[str, Any]] = []
    rows_all.extend(jp_rows)
    rows_all.extend(legacy_rows)

    for row in rows_all:
        normalize_row_payload(row, dx_name_map, dx_name_to_code)

    if args.osm_fetch_missing and osm_requests < args.osm_max_requests:
        def _needs_road(row: Dict[str, Any]) -> bool:
            return bool(row.get("has_latlon")) and _is_missing_label_text(row.get("道路"))

        def _needs_loc(row: Dict[str, Any]) -> bool:
            return bool(row.get("has_latlon")) and _is_missing_label_text(row.get("ロケーション"))

        prioritized_rows = [r for r in rows_all if _needs_road(r)]
        prioritized_rows.extend([r for r in rows_all if not _needs_road(r) and _needs_loc(r)])

        for row in prioritized_rows:
            if osm_requests >= args.osm_max_requests:
                break
            osm_requests = enrich_row_location_labels(
                row,
                osm_cache,
                rev_cache,
                True,
                args.osm_max_requests,
                osm_requests,
                args.osm_sleep,
                overpass_cache=overpass_cache,
                overpass_radius_m=args.osm_overpass_radius_m,
                overpass_timeout_s=args.osm_overpass_timeout_s,
            )

    # ---- Targeted OSM enrichment pass for '不明' location/road labels ----
    # Runs cache-first (always), and optionally fetches from Nominatim.
    # Uses the improved _extract_nominatim_location_label_ja() for better Japanese results.
    _osm_enrich_stats = enrich_location_with_osm(
        rows_all,
        osm_cache,
        rev_cache,
        overpass_cache,
        enable_fetch=args.osm_fetch_missing and osm_requests < args.osm_max_requests,
        max_requests=max(0, args.osm_max_requests - osm_requests),
        sleep_s=args.osm_sleep,
    )
    osm_requests += _osm_enrich_stats.get("fetch_requests_used", 0)

    # Save caches after all enrichment passes complete.
    latlon_cache_path.write_text(json.dumps(latlon_cache, ensure_ascii=False), encoding="utf-8")
    # Always save rev_cache if any enrichment pass ran (cache-only pass may have
    # re-extracted better labels from stored entries without making fetch requests).
    _rev_cache_dirty = (
        (args.osm_fetch_missing and osm_requests > 0)
        or _osm_enrich_stats.get("location_resolved_from_cache", 0) > 0
        or _osm_enrich_stats.get("road_resolved_from_cache", 0) > 0
    )
    if _rev_cache_dirty:
        save_osm_reverse_cache(rev_cache_path, rev_cache)
        save_json_object_cache(overpass_cache_path, overpass_cache)

    # assign __i for route snippet function
    for i, row in enumerate(rows_all):
        row["__i"] = i

    # dx options (main only). Keys must match JS: {"分類": code, "label": name}
    dx_options: List[Dict[str, str]] = []
    dx_option_map: Dict[str, str] = {}
    for k, name in sorted(dx_name_map.items()):
        dx_option_map[k] = name
    for row in rows_all:
        code = _s(row.get("分類"))
        label = _s(row.get("分類名"))
        if code and code != "Unknown" and code not in dx_option_map:
            dx_option_map[code] = label or code
    for code in sorted(dx_option_map):
        dx_options.append({"分類": code, "label": dx_option_map[code]})

    # always include Unknown so rows without DX are selectable and not filtered out
    if not any(o["分類"] == "Unknown" for o in dx_options):
        dx_options.append({"分類": "Unknown", "label": "不明"})

    # routes geojson (auto-detect if not provided)
    def _load_geojson(p: Path):
        try:
            return json.loads(p.read_text(encoding='utf-8'))
        except Exception:
            return {}

    routes_dir = Path(args.routes_dir) if args.routes_dir else Path('pipeline_out/routes_out')
    routes_carweek = {}
    routes_car = {}
    routes_scene = {}

    def _autofind(name_candidates):
        for pp in name_candidates:
            if pp and Path(pp).exists():
                return Path(pp)
        for nm in name_candidates:
            if nm:
                p1 = routes_dir / nm
                if p1.exists():
                    return p1
                p2 = Path(nm)
                if p2.exists():
                    return p2
        return None

    p_cw = _autofind([args.routes_by_carweek, 'routes_by_carweek.geojson', 'routes_by_car_week.geojson'])
    p_car = _autofind([args.routes_by_car, 'routes_by_car.geojson'])
    p_scene = _autofind([args.routes_by_scene, 'routes_by_scene.geojson'])

    if p_cw: routes_carweek = _load_geojson(p_cw)
    if p_car: routes_car = _load_geojson(p_car)
    if p_scene: routes_scene = _load_geojson(p_scene)

    # exposure
    exposure_dir = Path("pipeline_out/exposure_out")
    exposure_km_by_label, totals_overall, totals_by_car, totals_by_carweek = load_exposure(exposure_dir)
    route_metrics, route_scope_totals = load_route_metrics(exposure_dir, legacy_df, top_n=12)
    road_group_lookup = build_legacy_route_group_mapping(legacy_df)

    def _km_from_rec(rec: dict):
        for k in ['dist_geo_clamped_km','km','distance_km','total_km','dist_km','dist_km_geo','dist_geo_km']:
            if k in rec:
                try:
                    v = float(rec[k])
                    if v == v:
                        return v
                except Exception:
                    pass
        return None

    totals_by_car_map = {}
    totals_by_carweek_map = {}
    totals_by_carweek_car_map = {}

    # totals_by_car and totals_by_carweek currently are lists of records (from CSV). Convert to maps for JS.
    if isinstance(totals_by_car, list):
        for rec in totals_by_car:
            if not isinstance(rec, dict):
                continue
            cid = str(rec.get('car_id') or rec.get('car') or rec.get('cid') or '').strip()
            km = _km_from_rec(rec)
            if cid and km is not None:
                totals_by_car_map[cid] = km

    if isinstance(totals_by_carweek, list):
        for rec in totals_by_carweek:
            if not isinstance(rec, dict):
                continue
            cw = str(rec.get('car_week') or rec.get('carweek') or rec.get('cw') or '').strip()
            cid = str(rec.get('car_id') or rec.get('car') or rec.get('cid') or '').strip()
            km = _km_from_rec(rec)
            if cw and km is not None:
                # If file is per-carweek only, store simple map
                if cw not in totals_by_carweek_map or (totals_by_carweek_map[cw] is None):
                    totals_by_carweek_map[cw] = km
            if cw and cid and km is not None:
                totals_by_carweek_car_map.setdefault(cw, {})[cid] = km

    if isinstance(route_scope_totals, dict):
        for key, rec in (route_scope_totals.get("by_carweek_car") or {}).items():
            if not isinstance(rec, dict):
                continue
            try:
                cw, cid = (str(key).split("__", 1) + [""])[:2]
            except Exception:
                cw, cid = "", ""
            km = _safe_float(rec.get("distance_km"))
            if cw and cid and km is not None:
                totals_by_carweek_car_map.setdefault(cw, {})[cid] = float(km)


    # Compute bounds from mappable points
    mappable = [r for r in rows_all if _is_valid_coord(r.get("lat"), r.get("lon"))]
    if mappable:
        min_lat = min(r["lat"] for r in mappable); max_lat = max(r["lat"] for r in mappable)
        min_lon = min(r["lon"] for r in mappable); max_lon = max(r["lon"] for r in mappable)
    else:
        min_lat, max_lat, min_lon, max_lon = 34.0, 36.5, 138.0, 140.5

    meta = {
        "version": "v67_richui_both_excels",
        "generated_utc": datetime.datetime.utcnow().isoformat(timespec="seconds")+"Z",
        "jp_sheet_used": jp_sheet_used,
        "legacy_sheet_used": legacy_sheet_used,
        "jp_row_count": int(len(jp_rows)),
        "jp_sheet_row_span_rows": int(jp_parse_stats.get("sheet_row_span_rows", 0)),
        "jp_sheet_nonblank_rows": int(jp_parse_stats.get("sheet_nonblank_rows", len(jp_rows))),
        "jp_skipped_blank_rows": int(jp_parse_stats.get("skipped_blank_rows", 0)),
        "jp_skipped_empty_context_rows": int(jp_parse_stats.get("skipped_empty_context_rows", 0)),
        "jp_skipped_non_event_rows": int(jp_parse_stats.get("skipped_non_event_rows", 0)),
        "legacy_row_count": int(len(legacy_rows)),
        "total_rows": int(len(rows_all)),
        "mappable_points_with_latlon": int(len(mappable)),
        "missing_latlon_rows": int(len(rows_all) - len(mappable)),
        "logs_root": str(logs_root),
        "logs_glob": args.logs_glob,
        "logs_manifest": str(logs_manifest) if logs_manifest else None,
        "note": "Counts are based on ALL rows (JP parsed event rows + legacy1018). Map draws only rows with has_latlon.",
        # OSM location enrichment stats
        "location_resolved_from_cache": int(_osm_enrich_stats.get("location_resolved_from_cache", 0)),
        "location_resolved_from_fetch": int(_osm_enrich_stats.get("location_resolved_from_fetch", 0)),
        "location_still_unknown": int(_osm_enrich_stats.get("location_still_unknown", 0)),
        "road_resolved_from_cache": int(_osm_enrich_stats.get("road_resolved_from_cache", 0)),
        "road_resolved_from_fetch": int(_osm_enrich_stats.get("road_resolved_from_fetch", 0)),
        "osm_fetch_requests_used": int(_osm_enrich_stats.get("fetch_requests_used", 0)),
        "osm_fetch_errors": int(_osm_enrich_stats.get("fetch_errors", 0)),
    }

    data_blob = json.dumps({
        "meta": meta,
        "rows": rows_all,
        "dx_options": dx_options,
        "exposure": exposure_km_by_label,
        "totals_overall": totals_overall,
        "totals_by_car": totals_by_car_map,
        "totals_by_carweek": totals_by_carweek_map,
        "totals_by_carweek_car": totals_by_carweek_car_map,
        "totals_by_car_raw": totals_by_car,
        "totals_by_carweek_raw": totals_by_carweek,
        "route_metrics": route_metrics,
        "route_scope_totals": route_scope_totals,
        "road_group_lookup": road_group_lookup,
    }, ensure_ascii=False)

    html = (
        HTML_TEMPLATE
        .replace("__DATA_JSON__", data_blob)
        .replace("__ROUTES_CARWEEK_JSON__", json.dumps(routes_carweek, ensure_ascii=False))
        .replace("__ROUTES_CAR_JSON__", json.dumps(routes_car, ensure_ascii=False))
        .replace("__ROUTES_SCENE_JSON__", json.dumps(routes_scene, ensure_ascii=False))
        .replace("__MIN_LAT__", f"{min_lat:.10f}")
        .replace("__MAX_LAT__", f"{max_lat:.10f}")
        .replace("__MIN_LON__", f"{min_lon:.10f}")
        .replace("__MAX_LON__", f"{max_lon:.10f}")
    )
    html = _ensure_style_braces_ok(html)

    outp = Path(args.out)
    outp.parent.mkdir(parents=True, exist_ok=True)
    outp.write_text(html, encoding="utf-8")
    print(f"Wrote: {outp.resolve()}")
    print(
        "JP rows: "
        f"parsed={len(jp_rows)} / "
        f"sheet_nonblank={jp_parse_stats.get('sheet_nonblank_rows', len(jp_rows))} / "
        f"skipped_empty_context={jp_parse_stats.get('skipped_empty_context_rows', 0)} / "
        f"skipped_non_event={jp_parse_stats.get('skipped_non_event_rows', 0)} / "
        f"skipped_blank={jp_parse_stats.get('skipped_blank_rows', 0)}"
    )
    print(f"Legacy rows: {len(legacy_rows)}")
    print(f"ALL rows (JP parsed + legacy): {len(rows_all)} | mappable: {len(mappable)} | missing_latlon: {len(rows_all)-len(mappable)}")
    if args.osm_fetch_missing:
        print(f"OSM network lookups used: {osm_requests} (reverse cache: {rev_cache_path}, overpass cache: {overpass_cache_path})")
    _loc_cache  = _osm_enrich_stats.get("location_resolved_from_cache", 0)
    _loc_fetch  = _osm_enrich_stats.get("location_resolved_from_fetch", 0)
    _loc_unk    = _osm_enrich_stats.get("location_still_unknown", 0)
    _road_cache = _osm_enrich_stats.get("road_resolved_from_cache", 0)
    _road_fetch = _osm_enrich_stats.get("road_resolved_from_fetch", 0)
    print(
        f"OSM location enrichment: "
        f"location cache={_loc_cache} fetch={_loc_fetch} still_unknown={_loc_unk} | "
        f"road cache={_road_cache} fetch={_road_fetch} | "
        f"fetch_requests={_osm_enrich_stats.get('fetch_requests_used', 0)} errors={_osm_enrich_stats.get('fetch_errors', 0)}"
    )

if __name__ == "__main__":
    main()
